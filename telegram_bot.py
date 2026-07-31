# -*- coding: utf-8 -*-
"""
Bot do Telegram — arquivo para iniciar o robo (Run ▶).

1. Rode run_project.py antes (prepara ambiente).
2. Configure TELEGRAM_BOT_TOKEN no .env.
3. Clique em Run neste arquivo.
"""
from __future__ import annotations

import sys
from pathlib import Path

_PROJECT_DIR = Path(__file__).resolve().parent
if str(_PROJECT_DIR) not in sys.path:
    sys.path.insert(0, str(_PROJECT_DIR))

from src.bootstrap import relaunch_in_project_venv

relaunch_in_project_venv()

import atexit
import argparse
import json
import os
import re
import sys
import tempfile
import threading
import time
from collections.abc import Callable
from datetime import date, datetime
from pathlib import Path

import requests

from src.dates import today_local, now_local
from src.telegram_images import (
    THEME,
    format_brl,
    pil_available,
    render_data_table,
    render_kpi_dashboard,
    render_text_card,
)
from src.bot_processor import (
    apply_parse_result,
    build_missing_fields_message,
    build_preview,
    get_default_workbook,
    process_command,
    sale_id_from_parse_result,
    spreadsheet_setup_hint,
    _spreadsheet_error_message,
)
from src.parser import parse_message, _parse_pt_date, _is_service_delivery_finalized, should_replace_pending_preview, _extract_total_value, _currency_candidates, enrich_with_last_sale_context, recalculate_payments_for_total, parse_money_value, _extract_customer, apply_multi_field_corrections, apply_preview_corrections, apply_batch_preview_corrections, extract_supplemental_sale_id, _extract_target_sale_id_for_updates, extract_sale_ids_list_for_updates, extract_delete_sale_ids_list, _is_sale_delete_request, parse_multi_sales_message, detect_intent, sanitize_new_sale_identity, _extract_preview_correction_fields
from src.transcription import TranscriptionError, transcribe_audio
from src.llm_parser import llm_fallback_enabled, try_llm_parse
from src.monthly_report_pdf import build_planilha_report_pdf, collect_planilha_stats

# Carregar .env
def _load_dotenv() -> None:
    for base in (Path(__file__).resolve().parent, Path.cwd()):
        env_file = base / ".env"
        if env_file.is_file():
            for line in env_file.read_text(encoding="utf-8").splitlines():
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if "=" in line:
                    k, _, v = line.partition("=")
                    k, v = k.strip(), v.strip()
                    if v.startswith('"') and v.endswith('"') or v.startswith("'") and v.endswith("'"):
                        v = v[1:-1]
                    # Se o arquivo tiver chaves repetidas, queremos que a última
                    # prevaleça (evita problema de TELEGRAM_BOT_TOKEN ficar vazio).
                    os.environ[k] = v
            break


_load_dotenv()

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
BASE_URL = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}"


def _lock_path() -> Path:
    return Path(__file__).resolve().parent / ".telegram_bot.lock"


def _acquire_single_instance_lock() -> None:
    """Evita rodar duas instancias do bot ao mesmo tempo (causa previas/erros duplicados)."""
    from src.bootstrap import stop_old_bot_instances

    stop_old_bot_instances(prefix="Telegram")

    lock_path = _lock_path()
    lock_path.write_text(str(os.getpid()), encoding="utf-8")

    def _cleanup() -> None:
        try:
            lock_path.unlink(missing_ok=True)
        except Exception:
            pass

    atexit.register(_cleanup)

def _escape_md(text: str) -> str:
    """Evita que _, *, ` ou [ no nome quebrem o Markdown do Telegram."""
    return (
        str(text)
        .replace("\\", "\\\\")
        .replace("_", "\\_")
        .replace("*", "\\*")
        .replace("`", "\\`")
        .replace("[", "\\[")
    )


def _user_display_name(user: dict | None) -> str:
    """Nome do perfil Telegram (first_name), sem @username."""
    if not user:
        return ""
    return str(user.get("first_name") or "").strip()


def _help_text(user_name: str = "") -> str:
    name = _escape_md((user_name or "").strip())
    hello = f"Olá, {name}!" if name else "Olá!"
    return f"""👋 *{hello} Eu sou o assistente da sua planilha.*

⚙️ *Como funciona*
1️⃣ Você manda *áudio*, *texto* ou o *formulário* abaixo
2️⃣ Eu mostro uma *prévia*
3️⃣ Você confirma com *SIM* ou *OK*
4️⃣ Salvo na planilha ✅

🔹🔹🔹

🆕 *Exemplo 1 — venda nova (copie e preencha)*

Siga o roteiro — uma linha por campo:

```
Cliente: nome do cliente
Produto: o que foi vendido
Data da venda: hoje
Valor total: 3000
Pago: entrada 1500
Pendente: 1500
Data de entrega: 25/06
```

💡 *Pagou tudo?* Troque as duas linhas de pagamento por:
`Pago: a vista` ou `Status: pago`

💡 *Só material na hora da venda?* Adicione no final:
`Material: 800`

💬 *Ou fale naturalmente:*
_"Vendi um banner para a Lucia hoje por 2200. Deu 700 de entrada, falta 1500 pro dia 25."_

Ao salvar, a planilha gera o 🧾 *ID VENDA* (ex.: 012). *Guarde esse código.*

🔹🔹🔹

📂 *Exemplo 2 — venda já salva (sempre com ID VENDA)*

💬 *Cliente pagou:*
`ID VENDA 012 pagou tudo hoje`

💬 *Alterar entrega:*
`ID 012 data de entrega 06/07`

💬 *Alterar custo de material (substitui o anterior):*
`ID 012 altera custo material para 897,80`

💬 *Adicionar mais material (soma ao que já tem):*
`ID 012 adiciona mais 200 de material` ou `Gastei 350 na ID 012`

💬 *Alterar valor pendente:*
`ID 012 altera pendente para 1500`

🔹🔹🔹

🔧 *Exemplo 3 — outras atualizações*

💬 *Gasto com material:*
`Gastei 350 na ID VENDA 012`

💬 *Vários de uma vez:*
`ID 004, 005 pagou` ou `ID 004, 005 foi entregue`

🔹🔹🔹

⚡ *Botões rápidos*
📋 *Prévia* — pendências e entregas
📊 *Resumo* — totais gerais (use *Ver detalhes* na mensagem para quem pagou/deve)
📄 *Relatório PDF* — relatório com gráfico da planilha
✏️ *Corrigir* — alterar ou apagar um lançamento
🔄 *Nova conversa* — limpa o ID VENDA ativo

🔁 *ID VENDA na conversa*
Depois que informar um ID, o bot mantém para custos, pagamentos e entregas até *Nova conversa*.

✅ *Depois da prévia*
👍 Certo → *SIM* ou *OK*
✏️ Ajustar → mande só o que mudou: `Entrada: 2000`"""


# Compatível com imports/testes que ainda esperam HELP_TEXT.
HELP_TEXT = _help_text()
# Teclado de menu (botões que aparecem abaixo do campo de digitação)
MAIN_MENU_KEYBOARD = {
    "keyboard": [
        ["📋 Prévia", "📊 Resumo"],
        ["📄 Relatório PDF", "✏️ Corrigir"],
        ["❓ Ajuda", "🔄 Nova conversa"],
    ],
    "resize_keyboard": True,
    "one_time_keyboard": False,
}

RESUMO_DETAILS_INLINE_KEYBOARD = {
    "inline_keyboard": [
        [{"text": "🔍 Ver detalhes", "callback_data": "resumo_detalhes"}],
    ],
}

CORRECT_CATEGORIES_KEYBOARD = {
    "inline_keyboard": [
        [
            {"text": "🧾 Vendas", "callback_data": "corr:cat:vendas"},
            {"text": "🧱 Materiais", "callback_data": "corr:cat:materiais"},
        ],
        [
            {"text": "🏠 Gastos fixos", "callback_data": "corr:cat:fixos"},
        ],
        [{"text": "❌ Cancelar", "callback_data": "corr:cancel"}],
    ],
}


def _reset_chat_session(
    chat_id: int | str,
    *,
    pending_preview: dict,
    pending_delivery: dict,
    pending_overdue_delivery: dict,
    last_sale_id_by_chat: dict,
    last_customer_by_chat: dict | None = None,
    pending_correct: dict | None = None,
) -> None:
    """Limpa prévias e contexto de ID VENDA do chat."""
    pending_preview.pop(chat_id, None)
    pending_delivery.pop(chat_id, None)
    pending_overdue_delivery.pop(chat_id, None)
    last_sale_id_by_chat.pop(chat_id, None)
    if last_customer_by_chat is not None:
        last_customer_by_chat.pop(chat_id, None)
    if pending_correct is not None:
        pending_correct.pop(chat_id, None)


def _correct_intro_text(user_name: str = "") -> str:
    name = _escape_md((user_name or "").strip())
    hello = f"Beleza, {name}! " if name else "Beleza! "
    return (
        f"{hello}Vamos *corrigir ou apagar* um lançamento.\n\n"
        "1. Escolha a *categoria* abaixo\n"
        "2. Toque no *lançamento*\n"
        "3. Escolha *alterar* ou *apagar*\n"
        "4. Confirme com *SIM* quando pedir\n\n"
        "Para cancelar a qualquer momento: *NÃO* ou 🔄 *Nova conversa*."
    )


def _normalize_menu_button(text: str) -> str:
    """Remove emojis dos botões do menu para reconhecer o comando."""
    cleaned = (text or "").strip().lower()
    for ch in ("📄", "📊", "📋", "✏️", "🔄", "❓", "💡"):
        cleaned = cleaned.replace(ch, "")
    return cleaned.strip()


def _copyable_codeblock(body: str) -> str:
    return f"```\n{body.strip()}\n```"


def _build_correct_copy_template(category: str, item: dict[str, str]) -> str:
    if category == "vendas":
        sid = str(item.get("sale_id") or "003").strip()
        cliente = (item.get("cliente") or "NOVO NOME").strip()
        pendente_raw = item.get("pendente") or "0"
        try:
            pendente_txt = f"{float(pendente_raw):.0f}".replace(".0", "")
        except Exception:
            pendente_txt = "800"
        delivery = (item.get("delivery") or "20/07").strip()
        if delivery == "-":
            delivery = "20/07"
        return (
            f"ID {sid} cliente: {cliente}\n"
            f"ID {sid} pendente: {pendente_txt}\n"
            f"ID {sid} data de entrega {delivery}\n"
            f"ID {sid} pagou"
        )
    if category == "materiais":
        sid = (item.get("sale_id") or "012").strip()
        valor = item.get("valor") or "900"
        try:
            valor_txt = f"{float(valor):.0f}".replace(".0", "")
        except Exception:
            valor_txt = "900"
        if sid and sid != "-":
            return f"ID {sid} altera custo material para {valor_txt}"
        return f"Gastei {valor_txt} na ID VENDA 012"
    desc = (item.get("desc") or "Descrição").strip()
    valor = item.get("valor") or "280"
    try:
        valor_txt = f"{float(valor):.0f}".replace(".0", "")
    except Exception:
        valor_txt = "280"
    return f"Gasto fixo descrição: {desc}\nGasto fixo valor: {valor_txt}"


def _correct_copy_keyboard(
    template_body: str,
    category: str,
    item_key: str,
) -> dict:
    copy_text = (template_body or "").strip()[:256]
    return {
        "inline_keyboard": [
            [{"text": "📋 Copiar modelo", "copy_text": {"text": copy_text}}],
            [
                {"text": "⬅️ Voltar", "callback_data": f"corr:item:{category}:{item_key}"},
                {"text": "❌ Cancelar", "callback_data": "corr:cancel"},
            ],
        ]
    }


def _correct_ask_alter_message(category: str, item: dict[str, str]) -> tuple[str, str]:
    template_body = _build_correct_copy_template(category, item)
    copy_block = _copyable_codeblock(template_body)
    if category == "vendas":
        guide = (
            "✏️ *Copie o modelo abaixo*, edite no celular e envie.\n"
            "🎤 *Ou mande por áudio* falando o que quer mudar.\n\n"
            f"{copy_block}\n\n"
            "_Envie uma linha por vez ou o bloco inteiro editado._"
        )
    elif category == "materiais":
        guide = (
            "✏️ *Copie o modelo*, edite o valor/ID e envie.\n"
            "🎤 *Ou mande por áudio* descrevendo a correção.\n\n"
            f"{copy_block}"
        )
    else:
        guide = (
            "✏️ *Copie o modelo*, edite e envie.\n"
            "🎤 *Ou mande por áudio* com o ajuste do gasto fixo.\n\n"
            f"{copy_block}"
        )
    return f"{_correct_item_summary(category, item)}\n\n{guide}", template_body


def _try_handle_correct_alter_input(
    chat_id: int | str,
    text: str,
    *,
    pending_correct: dict,
    pending_preview: dict,
    last_sale_id_by_chat: dict,
    origin: str = "telegram",
) -> bool:
    pending = pending_correct.get(chat_id)
    if not pending or pending.get("step") != "await_alter":
        return False

    category = str(pending.get("category") or "")
    item_key = str(pending.get("item_key") or "")
    sale_id = str(pending.get("sale_id") or item_key or "").strip()
    alter_text = (text or "").strip()
    if not alter_text:
        return True

    if sale_id:
        last_sale_id_by_chat[chat_id] = sale_id

    lower = alter_text.lower()
    if "pagou" in lower and sale_id and category == "vendas":
        if not re.search(r"\bid\s+venda\b", lower) and not re.search(r"\bid\s+\d", lower):
            alter_text = f"ID {sale_id} {alter_text}"

    fields = _extract_preview_correction_fields(alter_text)
    applied_labels: list[str] = []
    if category == "vendas" and sale_id and fields:
        try:
            svc = _open_spreadsheet_for_bot()
            applied_labels = svc.apply_sale_field_corrections(
                sale_id,
                fields,
                reference_date=today_local(),
            )
        except Exception as exc:
            send_message(
                chat_id,
                f"⚠️ Não consegui aplicar a correção: {_spreadsheet_error_message(exc)}",
                reply_markup=MAIN_MENU_KEYBOARD,
            )
            return True

    if applied_labels and "pagou" not in lower:
        pending_correct.pop(chat_id, None)
        send_message(
            chat_id,
            "✅ *Planilha atualizada!*\n\n"
            + "\n".join(f"• {label}" for label in applied_labels),
            parse_mode="Markdown",
            reply_markup=MAIN_MENU_KEYBOARD,
        )
        return True

    parse_text = enrich_with_last_sale_context(alter_text, sale_id or last_sale_id_by_chat.get(chat_id))
    parse_result = parse_message(parse_text, reference_date=today_local())
    _normalize_material_only_sale(parse_result)
    parse_result = _maybe_enhance_parse_with_llm(
        alter_text,
        parse_result,
        last_sale_id=sale_id or last_sale_id_by_chat.get(chat_id),
    )

    if parse_result.missing_fields:
        send_message(
            chat_id,
            build_missing_fields_message(parse_result.missing_fields, parse_result.intent)
            + "\n\n_Tente copiar uma linha do modelo e editar só o que precisa._",
            parse_mode="Markdown",
            reply_markup=MAIN_MENU_KEYBOARD,
        )
        return True

    preview_intents = {
        "sale",
        "mixed_update",
        "refund",
        "status_update",
        "delivery_update",
        "delivery_finalize",
        "material_update",
        "sale_delete",
        "sale_value_update",
    }
    if parse_result.intent in preview_intents:
        pending_correct.pop(chat_id, None)
        preview_text = build_preview(parse_result)
        send_message(chat_id, preview_text, parse_mode="Markdown")
        pending_preview[chat_id] = {
            "parse_result": parse_result,
            "original_text": alter_text,
            "origin": origin,
        }
        _remember_sale_id_from_parse(parse_result, last_sale_id_by_chat, chat_id)
        return True

    pending_correct.pop(chat_id, None)
    with _spreadsheet_saving_timer(chat_id, getattr(parse_result, "intent", "sale")):
        reply = apply_parse_result(
            parse_result,
            origin=origin,
            original_text=alter_text,
            chat_id=str(chat_id),
        )
    send_message(chat_id, reply, parse_mode="Markdown", reply_markup=MAIN_MENU_KEYBOARD)
    _remember_sale_id_from_parse(parse_result, last_sale_id_by_chat, chat_id)
    return True


def _try_handle_correct_wizard_text(
    chat_id: int | str,
    text: str,
    *,
    pending_correct: dict,
) -> bool:
    """Evita que texto solto quebre o wizard Corrigir (antes de alterar/apagar)."""
    if chat_id not in pending_correct:
        return False
    step = pending_correct[chat_id].get("step")
    if step == "await_alter":
        return False
    send_message(
        chat_id,
        "Use os *botões* da mensagem acima para escolher o lançamento,\n"
        "ou toque em *❌ Cancelar* / 🔄 *Nova conversa*.",
        parse_mode="Markdown",
    )
    return True


    try:
        return format_brl(float(value_txt or 0))
    except Exception:
        return format_brl(0)


def _correct_sale_button_label(item: dict[str, str]) -> str:
    sid = item.get("sale_id") or "?"
    cliente = (item.get("cliente") or "Sem cliente").strip()
    produto = (item.get("produto") or "").strip()
    label = f"{sid} · {cliente}"
    if produto:
        label += f" · {produto}"
    return label[:60]


def _correct_material_button_label(item: dict[str, str]) -> str:
    desc = (item.get("desc") or "Material").strip()
    valor = _correct_money(item.get("valor") or "0")
    sid = (item.get("sale_id") or "").strip()
    label = f"{desc} · {valor}"
    if sid:
        label = f"ID {sid} · {label}"
    return label[:60]


def _correct_fixed_button_label(item: dict[str, str]) -> str:
    desc = (item.get("desc") or "Gasto").strip()
    valor = _correct_money(item.get("valor") or "0")
    return f"{desc} · {valor}"[:60]


def _correct_entries_keyboard(category: str, items: list[dict[str, str]]) -> dict:
    rows: list[list[dict[str, str]]] = []
    for item in items:
        if category == "vendas":
            sid = item.get("sale_id") or ""
            if not sid:
                continue
            rows.append(
                [{"text": _correct_sale_button_label(item), "callback_data": f"corr:item:vendas:{sid}"}]
            )
        elif category == "materiais":
            row_id = item.get("row") or ""
            if not row_id:
                continue
            rows.append(
                [{"text": _correct_material_button_label(item), "callback_data": f"corr:item:materiais:{row_id}"}]
            )
        else:
            row_id = item.get("row") or ""
            if not row_id:
                continue
            rows.append(
                [{"text": _correct_fixed_button_label(item), "callback_data": f"corr:item:fixos:{row_id}"}]
            )
    rows.append(
        [
            {"text": "⬅️ Categorias", "callback_data": "corr:back:cats"},
            {"text": "❌ Cancelar", "callback_data": "corr:cancel"},
        ]
    )
    return {"inline_keyboard": rows}


def _correct_actions_keyboard(category: str, item_key: str) -> dict:
    return {
        "inline_keyboard": [
            [
                {"text": "✏️ Alterar", "callback_data": f"corr:act:alt:{category}:{item_key}"},
                {"text": "🗑️ Apagar", "callback_data": f"corr:act:del:{category}:{item_key}"},
            ],
            [
                {"text": "⬅️ Voltar", "callback_data": f"corr:cat:{category}"},
                {"text": "❌ Cancelar", "callback_data": "corr:cancel"},
            ],
        ]
    }


def _start_correct_flow(
    chat_id: int | str,
    *,
    pending_correct: dict,
    pending_preview: dict,
    user_name: str = "",
) -> None:
    pending_preview.pop(chat_id, None)
    pending_correct[chat_id] = {"step": "category"}
    send_message(
        chat_id,
        _correct_intro_text(user_name),
        parse_mode="Markdown",
        reply_markup=CORRECT_CATEGORIES_KEYBOARD,
    )


def _correct_show_category(
    chat_id: int | str,
    category: str,
    *,
    pending_correct: dict,
    message_id: int | None = None,
) -> None:
    if not _spreadsheet_ready_for_bot():
        send_message(chat_id, "Planilha não encontrada.", reply_markup=MAIN_MENU_KEYBOARD)
        return
    try:
        svc = _open_spreadsheet_for_bot()
        if category == "vendas":
            items = svc.list_recent_sales(limit=10)
            title = "🧾 *Vendas recentes*\n\nToque no lançamento que deseja corrigir ou apagar:"
        elif category == "materiais":
            items = svc.list_recent_materials(limit=10)
            title = "🧱 *Materiais recentes*\n\nToque no lançamento:"
        else:
            items = svc.list_recent_fixed_costs(limit=10)
            title = "🏠 *Gastos fixos recentes*\n\nToque no lançamento:"
    except Exception as exc:
        send_message(chat_id, _spreadsheet_error_message(exc), reply_markup=MAIN_MENU_KEYBOARD)
        return

    pending_correct[chat_id] = {"step": "list", "category": category, "items": items}
    if not items:
        text = (
            f"{title}\n\n_Nenhum lançamento encontrado nesta categoria._"
        )
        keyboard = {
            "inline_keyboard": [
                [{"text": "⬅️ Categorias", "callback_data": "corr:back:cats"}],
                [{"text": "❌ Cancelar", "callback_data": "corr:cancel"}],
            ]
        }
    else:
        text = title
        keyboard = _correct_entries_keyboard(category, items)

    if message_id and edit_chat_message(
        chat_id, message_id, text, parse_mode="Markdown", reply_markup=keyboard
    ):
        return
    send_message(chat_id, text, parse_mode="Markdown", reply_markup=keyboard)


def _correct_find_item(pending: dict, category: str, item_key: str) -> dict[str, str] | None:
    items = pending.get("items") or []
    for item in items:
        if category == "vendas" and str(item.get("sale_id") or "") == item_key:
            return item
        if category != "vendas" and str(item.get("row") or "") == item_key:
            return item
    return None


def _correct_item_summary(category: str, item: dict[str, str]) -> str:
    if category == "vendas":
        lines = [
            f"🧾 *Venda ID {item.get('sale_id') or '-'}*",
            f"• Cliente: *{item.get('cliente') or '-'}*",
            f"• Produto: *{item.get('produto') or '-'}*",
            f"• Data venda: *{item.get('sale_date') or '-'}*",
            f"• Data entrega: *{item.get('delivery') or '-'}*",
            f"• Pago: *{_correct_money(item.get('pago') or '0')}*",
            f"• Pendente: *{_correct_money(item.get('pendente') or '0')}*",
            f"• Status: *{item.get('status') or '-'}*",
        ]
        return "\n".join(lines)
    if category == "materiais":
        sid = item.get("sale_id") or "-"
        return "\n".join(
            [
                "🧱 *Material*",
                f"• Descrição: *{item.get('desc') or '-'}*",
                f"• Valor: *{_correct_money(item.get('valor') or '0')}*",
                f"• ID VENDA: *{sid}*",
                f"• Data: *{item.get('data') or '-'}*",
            ]
        )
    return "\n".join(
        [
            "🏠 *Gasto fixo*",
            f"• Descrição: *{item.get('desc') or '-'}*",
            f"• Valor: *{_correct_money(item.get('valor') or '0')}*",
            f"• Data: *{item.get('data') or '-'}*",
        ]
    )


def _correct_show_item(
    chat_id: int | str,
    category: str,
    item_key: str,
    *,
    pending_correct: dict,
    message_id: int | None = None,
) -> None:
    pending = pending_correct.get(chat_id) or {}
    item = _correct_find_item(pending, category, item_key)
    if item is None:
        # Lista pode ter expirado — recarrega categoria.
        _correct_show_category(chat_id, category, pending_correct=pending_correct, message_id=message_id)
        return
    pending_correct[chat_id] = {
        "step": "item",
        "category": category,
        "item_key": item_key,
        "items": pending.get("items") or [item],
        "item": item,
    }
    text = (
        f"{_correct_item_summary(category, item)}\n\n"
        "O que deseja fazer com este lançamento?"
    )
    keyboard = _correct_actions_keyboard(category, item_key)
    if message_id and edit_chat_message(
        chat_id, message_id, text, parse_mode="Markdown", reply_markup=keyboard
    ):
        return
    send_message(chat_id, text, parse_mode="Markdown", reply_markup=keyboard)


def _correct_ask_alter(
    chat_id: int | str,
    category: str,
    item_key: str,
    *,
    pending_correct: dict,
    last_sale_id_by_chat: dict,
    message_id: int | None = None,
) -> None:
    pending = pending_correct.get(chat_id) or {}
    item = pending.get("item") or _correct_find_item(pending, category, item_key)
    if item is None:
        _correct_show_category(chat_id, category, pending_correct=pending_correct, message_id=message_id)
        return

    if category == "vendas":
        sale_id = item.get("sale_id") or item_key
        last_sale_id_by_chat[chat_id] = sale_id
        pending_correct[chat_id] = {
            "step": "await_alter",
            "category": category,
            "item_key": item_key,
            "item": item,
            "sale_id": sale_id,
        }
        text, template_body = _correct_ask_alter_message(category, item)
    elif category == "materiais":
        sale_id = (item.get("sale_id") or "").strip()
        if sale_id:
            last_sale_id_by_chat[chat_id] = sale_id
        pending_correct[chat_id] = {
            "step": "await_alter",
            "category": category,
            "item_key": item_key,
            "item": item,
            "sale_id": sale_id,
        }
        text, template_body = _correct_ask_alter_message(category, item)
    else:
        pending_correct[chat_id] = {
            "step": "await_alter",
            "category": category,
            "item_key": item_key,
            "item": item,
        }
        text, template_body = _correct_ask_alter_message(category, item)

    keyboard = _correct_copy_keyboard(template_body, category, item_key)
    if message_id and edit_chat_message(
        chat_id, message_id, text, parse_mode="Markdown", reply_markup=keyboard
    ):
        return
    send_message(chat_id, text, parse_mode="Markdown", reply_markup=keyboard)


def _correct_ask_delete(
    chat_id: int | str,
    category: str,
    item_key: str,
    *,
    pending_correct: dict,
    pending_preview: dict,
    message_id: int | None = None,
) -> None:
    pending = pending_correct.get(chat_id) or {}
    item = pending.get("item") or _correct_find_item(pending, category, item_key)
    if item is None:
        _correct_show_category(chat_id, category, pending_correct=pending_correct, message_id=message_id)
        return

    if category != "vendas":
        text = (
            f"{_correct_item_summary(category, item)}\n\n"
            "Por enquanto a exclusão guiada está disponível para *Vendas*.\n"
            "Para materiais/gastos, envie o ajuste por texto ou apague direto na planilha.\n\n"
            "Toque em *⬅️ Voltar* para escolher outra ação."
        )
        keyboard = _correct_actions_keyboard(category, item_key)
        if message_id and edit_chat_message(
            chat_id, message_id, text, parse_mode="Markdown", reply_markup=keyboard
        ):
            return
        send_message(chat_id, text, parse_mode="Markdown", reply_markup=keyboard)
        return

    sale_id = item.get("sale_id") or item_key
    pr = parse_message(f"apagar id venda {sale_id}", reference_date=today_local())
    pending_correct.pop(chat_id, None)
    pending_preview[chat_id] = {
        "parse_result": pr,
        "original_text": f"apagar id venda {sale_id}",
        "origin": "telegram",
    }
    preview = build_preview(pr)
    text = f"{_correct_item_summary(category, item)}\n\n{preview}"
    if message_id and edit_chat_message(
        chat_id,
        message_id,
        text,
        parse_mode="Markdown",
        reply_markup=MAIN_MENU_KEYBOARD,
    ):
        return
    send_message(chat_id, text, parse_mode="Markdown", reply_markup=MAIN_MENU_KEYBOARD)


def _handle_correct_callback(
    chat_id: int | str,
    callback_query_id: str,
    callback_data: str,
    *,
    pending_correct: dict,
    pending_preview: dict,
    last_sale_id_by_chat: dict,
    message_id: int | None = None,
) -> bool:
    """Callbacks do fluxo Corrigir. Retorna True se tratou."""
    if not callback_data.startswith("corr:"):
        return False

    answer_callback_query(callback_query_id)
    parts = callback_data.split(":")
    action = parts[1] if len(parts) > 1 else ""

    if action == "cancel":
        pending_correct.pop(chat_id, None)
        text = "❌ Correção cancelada. Pode continuar normalmente."
        if message_id:
            edit_chat_message(chat_id, message_id, text, parse_mode="Markdown", reply_markup={"inline_keyboard": []})
        send_message(chat_id, text, parse_mode="Markdown", reply_markup=MAIN_MENU_KEYBOARD)
        return True

    if action == "back" and len(parts) > 2 and parts[2] == "cats":
        pending_correct[chat_id] = {"step": "category"}
        text = _correct_intro_text()
        if message_id and edit_chat_message(
            chat_id, message_id, text, parse_mode="Markdown", reply_markup=CORRECT_CATEGORIES_KEYBOARD
        ):
            return True
        send_message(chat_id, text, parse_mode="Markdown", reply_markup=CORRECT_CATEGORIES_KEYBOARD)
        return True

    if action == "cat" and len(parts) > 2:
        _correct_show_category(
            chat_id, parts[2], pending_correct=pending_correct, message_id=message_id
        )
        return True

    if action == "item" and len(parts) > 3:
        _correct_show_item(
            chat_id,
            parts[2],
            parts[3],
            pending_correct=pending_correct,
            message_id=message_id,
        )
        return True

    if action == "act" and len(parts) > 4:
        kind = parts[2]  # alt | del
        category = parts[3]
        item_key = parts[4]
        if kind == "alt":
            _correct_ask_alter(
                chat_id,
                category,
                item_key,
                pending_correct=pending_correct,
                last_sale_id_by_chat=last_sale_id_by_chat,
                message_id=message_id,
            )
        elif kind == "del":
            _correct_ask_delete(
                chat_id,
                category,
                item_key,
                pending_correct=pending_correct,
                pending_preview=pending_preview,
                message_id=message_id,
            )
        return True

    return True


def _is_correct_menu_command(text: str) -> bool:
    norm = _normalize_menu_button(text)
    return norm in {
        "corrigir",
        "corrigir lançamento",
        "corrigir lancamento",
        "alterar lançamento",
        "alterar lancamento",
    }


def _remember_sale_id_from_parse(
    parse_result,
    last_sale_id_by_chat: dict,
    chat_id: int | str,
) -> None:
    remembered = sale_id_from_parse_result(parse_result)
    if remembered:
        last_sale_id_by_chat[chat_id] = remembered


def _normalize_material_only_sale(parse_result) -> None:
    """Prévia de custo/material com ID não deve exigir Cliente como venda nova."""
    cmd = parse_result.command
    sale_id = (getattr(cmd, "sale_id", None) or "").strip()
    material_cost = float(getattr(cmd, "material_cost", None) or 0)
    total_value = float(getattr(cmd, "total_value", None) or 0)
    if not sale_id or material_cost <= 0 or total_value > 0.01:
        return
    if parse_result.intent in ("sale", "mixed_update"):
        parse_result.intent = "material_update"
    parse_result.missing_fields = [
        f for f in parse_result.missing_fields if f not in ("Cliente", "Valor total da venda")
    ]


def _load_customer_from_workbook(sale_id: str) -> str:
    sale_id = str(sale_id or "").strip()
    if not sale_id:
        return ""
    if not _spreadsheet_ready_for_bot():
        return ""
    try:
        svc = _open_spreadsheet_for_bot()
        rows = svc.get_pending_sale_by_sale_id(sale_id, max_rows_scan=500, include_paid=True)
        if rows:
            return str(rows[0].get("customer") or "").strip()
    except Exception:
        pass
    return ""


QUICK_DASHBOARD_KEYWORDS = (
    "resumo",
    "status",
    "planilha",
    "prévia",
    "previa",
    "detalhes",
    "detalhe",
)


def send_chat_action(chat_id: int | str, action: str = "typing") -> None:
    """Indica no Telegram que o bot está processando (melhora percepção de velocidade)."""
    if not TELEGRAM_BOT_TOKEN:
        return
    try:
        requests.post(
            f"{BASE_URL}/sendChatAction",
            json={"chat_id": str(chat_id), "action": action},
            timeout=5,
        )
    except Exception:
        pass


def _pulse_chat_action(chat_id: int | str, action: str, stop_event: threading.Event) -> None:
    """Mantém o indicador de digitação/envio ativo enquanto uma operação demorada roda."""
    send_chat_action(chat_id, action)
    while not stop_event.wait(4.0):
        send_chat_action(chat_id, action)


def _run_with_chat_action(chat_id: int | str, action: str, func):
    """Executa `func` renovando o chat action a cada poucos segundos."""
    stop = threading.Event()
    worker = threading.Thread(
        target=_pulse_chat_action,
        args=(chat_id, action, stop),
        daemon=True,
    )
    worker.start()
    try:
        return func()
    finally:
        stop.set()
        worker.join(timeout=1.0)


def _is_quick_dashboard_command(text: str) -> bool:
    cmd_strip = _normalize_menu_button(text)
    return any(cmd_strip == kw or cmd_strip.startswith(kw) for kw in QUICK_DASHBOARD_KEYWORDS)


def _maybe_build_text_image_png(text: str) -> str | None:
    """Gera imagem legível do resumo em texto para o Telegram."""
    if not pil_available():
        return None
    try:
        return render_text_card(text, title="Resumo da planilha")
    except Exception:
        return None


def _maybe_build_status_table_image(
    workbook_path: str,
    wb=None,
    svc=None,
) -> str | None:
    """Gera card de lucros e totais (Status / Resumo)."""
    if not pil_available():
        return None
    try:
        from src.excel_store import SHEET_SALES, SHEET_MATERIAL, SHEET_FIXED, DATA_START_ROW
    except Exception:
        return None

    try:
        own_wb = wb is None
        if own_wb:
            svc = _open_spreadsheet_for_bot()
            wb = svc._open_workbook(data_only=True)
        try:
            ws_sales = wb[svc._resolve_sheet_name(wb, SHEET_SALES)]
            ws_mat = wb[svc._resolve_sheet_name(wb, SHEET_MATERIAL)]
            ws_fix = wb[svc._resolve_sheet_name(wb, SHEET_FIXED)]

            sales_cols = svc._sales_columns(ws_sales)
            col_id = sales_cols.get("id venda")
            col_paid = sales_cols.get("total de vendas (pago)")
            col_pending = sales_cols.get("valor (pendente)")

            sales_end = svc._effective_data_end_row(ws_sales, col_id) if col_id else svc._scan_row_cap(ws_sales)
            total_paid = svc._sum_column_range(ws_sales, col_paid, DATA_START_ROW, sales_end)
            total_pending = svc._sum_column_range(ws_sales, col_pending, DATA_START_ROW, sales_end)

            mat_cols = svc._material_columns(ws_mat)
            col_mat_desc = mat_cols.get("descricao")
            col_mat_val = mat_cols.get("valor")
            mat_end = svc._effective_data_end_row(ws_mat, col_mat_desc) if col_mat_desc else svc._scan_row_cap(ws_mat)
            total_mat = svc._sum_column_range(ws_mat, col_mat_val, DATA_START_ROW, mat_end)

            fix_end = svc._scan_row_cap(ws_fix)
            total_fix = svc._sum_column_range(ws_fix, "D", DATA_START_ROW, fix_end)
        finally:
            if own_wb:
                wb.close()
    except Exception:
        return None

    total_sales = round(total_paid + total_pending, 2)
    lucro = round(total_sales - total_mat - total_fix, 2)
    kpi = [
        ("Total de vendas (geral)", format_brl(total_sales), False),
        ("Vendas pagas", format_brl(total_paid), False),
        ("Valor pendente", format_brl(total_pending), False),
        ("Compras materia-prima", format_brl(total_mat), False),
        ("Gastos fixos", format_brl(total_fix), False),
        ("Lucro estimado", format_brl(lucro), True),
    ]
    try:
        return render_kpi_dashboard(kpi, profit_value=lucro)
    except Exception:
        return None


def _maybe_build_sales_snippet_image(
    workbook_path: str,
    *,
    wb=None,
    svc=None,
    sale_id: str | None = None,
    limit: int = 7,
    title_main: str = "TOTAL DE VENDAS - Visão atual",
    meta_line: str | None = None,
    snippet_mode: str = "preview",
) -> str | None:
    """Gera tabela visual da aba de vendas (Prévia / Detalhes)."""
    if not pil_available():
        return None
    try:
        from src.excel_store import SHEET_SALES, DATA_START_ROW
    except Exception:
        return None

    rows_txt: list[list[str]] = []
    headers: list[str] = []
    delivery_cell_states: list[bool | None] = []

    try:
        own_wb = wb is None
        if own_wb:
            svc = _open_spreadsheet_for_bot()
            wb = svc._open_workbook(data_only=False)
        try:
            ws = wb[svc._resolve_sheet_name(wb, SHEET_SALES)]
            cols = svc._sales_columns(ws)
            material_totals = svc.material_costs_by_sale_id(wb)
            if snippet_mode == "details":
                display_cols = [
                    ("cliente", "Cliente"),
                    ("id produto", "Produto"),
                    ("total de vendas (pago)", "Entrada"),
                    ("valor (pendente)", "Pendente"),
                    ("__material__", "Material"),
                ]
            else:
                display_cols = [
                    ("data de venda", "Data da venda"),
                    ("data de entrega", "Data entrega"),
                    ("cliente", "Cliente"),
                    ("id produto", "Produto"),
                    ("__material__", "Custo material"),
                    ("total de vendas (pago)", "Pago"),
                    ("valor (pendente)", "Pendente"),
                    ("id venda", "ID VENDA"),
                    ("status de valor", "Status"),
                ]
            col_pairs: list[tuple[str | None, str, str]] = []
            for key, label in display_cols:
                if key == "__material__":
                    col_pairs.append((None, label, key))
                elif cols.get(key):
                    col_pairs.append((cols.get(key), label, key))
            if not any(c for c, _label, _key in col_pairs if c):
                return None

            id_col = cols.get("id venda")
            data_rows: list[int] = []
            if id_col:
                sales_end = svc._effective_data_end_row(ws, id_col)
                if getattr(ws, "read_only", False):
                    from openpyxl.utils import column_index_from_string

                    col_idx = column_index_from_string(id_col)
                    for row in ws.iter_rows(
                        min_row=DATA_START_ROW,
                        max_row=sales_end,
                        min_col=col_idx,
                        max_col=col_idx,
                    ):
                        if row[0].value not in (None, ""):
                            data_rows.append(row[0].row)
                else:
                    for r in range(DATA_START_ROW, sales_end + 1):
                        if ws[f"{id_col}{r}"].value not in (None, ""):
                            data_rows.append(r)
            if not data_rows:
                return None

            highlight_row = data_rows[-1]
            if sale_id and id_col:
                for r in reversed(data_rows[-250:]):
                    if str(ws[f"{id_col}{r}"].value or "").strip() == str(sale_id).strip():
                        highlight_row = r
                        break

            last_rows = data_rows[-max(limit, 3):]
            if snippet_mode != "details" and highlight_row not in last_rows:
                last_rows = (last_rows + [highlight_row])[-max(limit, 3):]

            headers = [label for _c, label, _key in col_pairs]

            def _cell_txt(c: str | None, key: str, r: int) -> str:
                if key == "__material__":
                    if not id_col:
                        return "-"
                    sid = str(ws[f"{id_col}{r}"].value or "").strip()
                    amount = material_totals.get(sid, 0.0)
                    if amount <= 0.01:
                        return "-"
                    return format_brl(amount)
                v = ws[f"{c}{r}"].value
                if v in (None, ""):
                    return ""
                if key in ("total de vendas (pago)", "valor (pendente)"):
                    return format_brl(svc._to_float(v))
                if key in ("data de venda", "data de entrega"):
                    return svc._format_date_br(v) or ""
                if hasattr(v, "strftime"):
                    try:
                        return v.strftime("%d/%m/%Y")
                    except Exception:
                        pass
                return str(v)

            rows_txt = [[_cell_txt(c, key, r) for c, _label, key in col_pairs] for r in last_rows]
            if snippet_mode != "details":
                reminder_meta = svc._reminder_meta_by_sale_id(wb)
                for r in last_rows:
                    delivery_col = cols.get("data de entrega")
                    state: bool | None = None
                    sale_key = (
                        svc._format_sale_id(ws[f"{id_col}{r}"].value) if id_col else ""
                    )
                    rem_status = (reminder_meta.get(sale_key) or {}).get("status_servico", "")
                    if svc._normalize_name(rem_status) == "finalizado":
                        state = False
                    elif delivery_col:
                        cell = ws[f"{delivery_col}{r}"]
                        if svc._cell_fill_matches(cell, "C8E6C9"):
                            state = False
                        elif svc._cell_fill_matches(cell, "FFF59D"):
                            state = True
                        elif cell.value not in (None, ""):
                            # Sem fill na nuvem: enquanto não confirmou entrega, fica amarelo
                            # (mesmo se o status financeiro já for "pago").
                            state = True
                    delivery_cell_states.append(state)
            else:
                delivery_cell_states = [None] * len(last_rows)
        finally:
            if own_wb:
                wb.close()
    except Exception:
        return None

    if not rows_txt:
        return None

    def _row_has_pending(rvals: list[str]) -> bool:
        try:
            pi = headers.index("Pendente")
            pt = rvals[pi]
            if pt.startswith("R$"):
                raw = pt.replace("R$", "").strip().replace(".", "").replace(",", ".")
                if float(raw) > 0.01:
                    return True
        except Exception:
            pass
        try:
            si = headers.index("Status")
            if "pendente" in (rvals[si] or "").strip().lower():
                return True
        except Exception:
            pass
        return False

    def _cell_style(ri: int, ci: int, header: str, row: list[str]) -> tuple[tuple[int, int, int], tuple[int, int, int]]:
        pending_row = _row_has_pending(row)
        delivery_state = delivery_cell_states[ri] if ri < len(delivery_cell_states) else None
        base_fill = THEME.highlight if pending_row else (THEME.zebra if ri % 2 else THEME.card_bg)
        cell_ink = THEME.ink

        if header == "Data entrega":
            if delivery_state is True:
                return THEME.status_pending, THEME.ink
            if delivery_state is False:
                return THEME.delivery_done, THEME.ink
            return base_fill, cell_ink

        if header == "Status":
            sl = (row[ci] or "").strip().lower()
            if sl == "pago" or (sl.startswith("pago") and "pendente" not in sl):
                return THEME.status_paid, THEME.ink
            if "pendente" in sl:
                return THEME.status_pending, THEME.ink
            return base_fill, cell_ink

        if header == "Pendente":
            txt = row[ci] or ""
            if txt.startswith("R$"):
                try:
                    raw = txt.replace("R$", "").strip().replace(".", "").replace(",", ".")
                    if float(raw) > 0.01:
                        return base_fill, THEME.pending
                except Exception:
                    pass

        return base_fill, cell_ink

    meta_text = (meta_line or "").strip()
    try:
        return render_data_table(
            title=title_main,
            meta=meta_text or None,
            headers=headers,
            rows=rows_txt,
            cell_style=_cell_style,
        )
    except Exception:
        return None


def send_photo(
    chat_id: int | str,
    image_path: str,
    caption: str = "",
    parse_mode: str | None = None,
    reply_markup: dict | None = None,
) -> bool:
    """Envia uma foto para o chat (sendPhoto), com legenda opcional."""
    if not TELEGRAM_BOT_TOKEN:
        return False
    try:
        with open(image_path, "rb") as f:
            files = {"photo": f}
            data: dict[str, str] = {"chat_id": str(chat_id)}
            if caption:
                data["caption"] = caption[:1024]
            if parse_mode:
                data["parse_mode"] = parse_mode
            if reply_markup:
                data["reply_markup"] = json.dumps(reply_markup)
            r = requests.post(f"{BASE_URL}/sendPhoto", data=data, files=files, timeout=30)
            return r.status_code == 200
    except Exception as e:
        print(f"[Telegram] Erro ao enviar foto: {e}")
        return False


def send_document(
    chat_id: int | str,
    file_path: str,
    caption: str = "",
    filename: str | None = None,
    parse_mode: str | None = None,
    reply_markup: dict | None = None,
) -> bool:
    """Envia um documento (ex.: PDF) para o chat."""
    if not TELEGRAM_BOT_TOKEN:
        return False
    try:
        doc_name = filename or Path(file_path).name
        with open(file_path, "rb") as f:
            files = {"document": (doc_name, f)}
            data: dict[str, str] = {"chat_id": str(chat_id)}
            if caption:
                data["caption"] = caption[:1024]
            if parse_mode:
                data["parse_mode"] = parse_mode
            if reply_markup:
                data["reply_markup"] = json.dumps(reply_markup)
            r = requests.post(f"{BASE_URL}/sendDocument", data=data, files=files, timeout=60)
            return r.status_code == 200
    except Exception as e:
        print(f"[Telegram] Erro ao enviar documento: {e}")
        return False


def send_reply(
    chat_id: int | str,
    text: str,
    *,
    parse_mode: str | None = None,
    reply_markup: dict | None = None,
    image_path: str | None = None,
) -> bool:
    """Envia texto ou foto com legenda (uma única mensagem quando há imagem)."""
    if image_path:
        return send_photo(
            chat_id,
            image_path,
            caption=text,
            parse_mode=parse_mode,
            reply_markup=reply_markup,
        )
    return send_message(chat_id, text, parse_mode=parse_mode, reply_markup=reply_markup)


def send_message(
    chat_id: int | str,
    text: str,
    parse_mode: str | None = None,
    reply_markup: dict | None = None,
) -> int | None:
    """Envia mensagem de texto para o chat. Retorna message_id em caso de sucesso."""
    if not TELEGRAM_BOT_TOKEN:
        return None
    text = (text or "")[:4096]
    payload: dict = {"chat_id": chat_id, "text": text}
    if parse_mode:
        payload["parse_mode"] = parse_mode
    if reply_markup:
        payload["reply_markup"] = reply_markup
    try:
        r = requests.post(f"{BASE_URL}/sendMessage", json=payload, timeout=15)
        if r.status_code != 200:
            return None
        data = r.json()
        result = data.get("result") or {}
        message_id = result.get("message_id")
        return int(message_id) if isinstance(message_id, int) else None
    except Exception as e:
        print(f"[Telegram] Erro ao enviar: {e}")
        return None


def answer_callback_query(callback_query_id: str, text: str = "") -> None:
    """Confirma o toque em botão inline (remove o relógio de carregamento)."""
    if not TELEGRAM_BOT_TOKEN or not callback_query_id:
        return
    try:
        payload: dict = {"callback_query_id": str(callback_query_id)}
        if text:
            payload["text"] = text[:200]
        requests.post(f"{BASE_URL}/answerCallbackQuery", json=payload, timeout=10)
    except Exception as e:
        print(f"[Telegram] Erro ao responder callback: {e}")


def delete_chat_message(chat_id: int | str, message_id: int | None) -> None:
    """Remove mensagem temporária (ex.: aviso de aguarde)."""
    if not TELEGRAM_BOT_TOKEN or not message_id:
        return
    try:
        requests.post(
            f"{BASE_URL}/deleteMessage",
            json={"chat_id": str(chat_id), "message_id": int(message_id)},
            timeout=10,
        )
    except Exception:
        pass


def edit_chat_message(
    chat_id: int | str,
    message_id: int | None,
    text: str,
    *,
    parse_mode: str | None = None,
    reply_markup: dict | None = None,
) -> bool:
    """Atualiza texto de uma mensagem já enviada (ex.: cronômetro de aguarde)."""
    if not TELEGRAM_BOT_TOKEN or not message_id:
        return False
    payload: dict = {
        "chat_id": str(chat_id),
        "message_id": int(message_id),
        "text": (text or "")[:4096],
    }
    if parse_mode:
        payload["parse_mode"] = parse_mode
    if reply_markup is not None:
        payload["reply_markup"] = reply_markup
    try:
        r = requests.post(f"{BASE_URL}/editMessageText", json=payload, timeout=10)
        return r.status_code == 200
    except Exception:
        return False


def _format_elapsed_timer(elapsed_sec: float) -> str:
    total = max(0, int(elapsed_sec))
    minutes, seconds = divmod(total, 60)
    return f"{minutes}:{seconds:02d}"


def _spreadsheet_saving_wait_message(
    intent: str = "",
    *,
    batch: bool = False,
    elapsed_sec: float | None = None,
) -> str:
    if batch:
        body = (
            "⏳ *Salvando na planilha...*\n"
            "Atualizando vários registros. Aguarde um instante."
        )
    else:
        intent = (intent or "").strip().lower()
        action_labels = {
            "sale": "a venda",
            "material_update": "o material",
            "status_update": "o status da venda",
            "payment_update": "o pagamento",
            "delivery_update": "a data de entrega",
            "delivery_finalize": "a entrega",
            "sale_delete": "a exclusão",
            "refund": "o estorno",
            "mixed_update": "a atualização",
        }
        action = action_labels.get(intent, "o lançamento")
        body = (
            "⏳ *Salvando na planilha...*\n"
            f"Registrando {action}. Aguarde um instante."
        )
    if elapsed_sec is not None:
        body += f"\n⏱️ Tempo: *{_format_elapsed_timer(elapsed_sec)}*"
    return body


class _LiveWaitTimer:
    """Envia mensagem de aguarde e atualiza o cronômetro a cada segundo."""

    def __init__(
        self,
        chat_id: int | str,
        build_message: Callable[[float], str],
        *,
        chat_action: str = "typing",
    ) -> None:
        self.chat_id = chat_id
        self.build_message = build_message
        self.chat_action = chat_action
        self.message_id: int | None = None
        self._stop = threading.Event()
        self._started = 0.0
        self._worker: threading.Thread | None = None

    def __enter__(self) -> "_LiveWaitTimer":
        self._started = time.monotonic()
        self.message_id = send_message(
            self.chat_id,
            self.build_message(0),
            parse_mode="Markdown",
        )
        send_chat_action(self.chat_id, self.chat_action)
        if self.message_id:
            self._worker = threading.Thread(target=self._run_ticks, daemon=True)
            self._worker.start()
        return self

    def __exit__(self, exc_type, exc, tb) -> bool:
        self._stop.set()
        if self._worker:
            self._worker.join(timeout=2.0)
        delete_chat_message(self.chat_id, self.message_id)
        return False

    def _run_ticks(self) -> None:
        while not self._stop.wait(1.0):
            if not self.message_id:
                continue
            elapsed = time.monotonic() - self._started
            edit_chat_message(
                self.chat_id,
                self.message_id,
                self.build_message(elapsed),
                parse_mode="Markdown",
            )
            send_chat_action(self.chat_id, self.chat_action)


def _spreadsheet_saving_timer(
    chat_id: int | str,
    intent: str = "",
    *,
    batch: bool = False,
) -> _LiveWaitTimer:
    def build_message(elapsed: float) -> str:
        return _spreadsheet_saving_wait_message(intent, batch=batch, elapsed_sec=elapsed)

    return _LiveWaitTimer(chat_id, build_message)


def _notify_spreadsheet_saving(chat_id: int | str, intent: str = "", *, batch: bool = False) -> None:
    """Compatibilidade: prefira `with _spreadsheet_saving_timer(...):`."""
    send_message(
        chat_id,
        _spreadsheet_saving_wait_message(intent, batch=batch),
        parse_mode="Markdown",
    )
    send_chat_action(chat_id, "typing")


def _audio_download_read_timeout(*, file_size: int = 0, duration_sec: int = 0) -> int:
    """Timeout de leitura do download — escala com tamanho e duração do áudio."""
    try:
        env_max = int(os.getenv("TELEGRAM_AUDIO_DOWNLOAD_TIMEOUT", "600") or 600)
    except ValueError:
        env_max = 600
    env_max = max(120, min(env_max, 900))
    by_size = 60 + (max(file_size, 0) // 2048)
    by_duration = max(duration_sec, 0) * 4
    return min(env_max, max(180, by_size, by_duration))


def _audio_processing_wait_message(duration_sec: int, *, elapsed_sec: float | None = None) -> str:
    if duration_sec >= 120:
        wait_hint = "cerca de 2 a 4 minutos"
    elif duration_sec >= 60:
        wait_hint = "cerca de 1 a 2 minutos"
    elif duration_sec >= 30:
        wait_hint = "30 segundos a 1 minuto"
    else:
        wait_hint = "alguns segundos"
    body = (
        f"⏳ *Processando áudio...*\n"
        f"Pode levar {wait_hint}. Aguarde — não envie outra mensagem ainda."
    )
    if elapsed_sec is not None:
        body += f"\n⏱️ Tempo: *{_format_elapsed_timer(elapsed_sec)}*"
    return body


def _audio_processing_timer(chat_id: int | str, duration_sec: int) -> _LiveWaitTimer:
    def build_message(elapsed: float) -> str:
        return _audio_processing_wait_message(duration_sec, elapsed_sec=elapsed)

    return _LiveWaitTimer(chat_id, build_message)


def _dashboard_wait_message(cmd_strip: str) -> str:
    if cmd_strip.startswith(("prévia", "previa")):
        title = "a prévia"
        detail = "pendências e entregas"
    elif cmd_strip.startswith(("detalhes", "detalhe")):
        title = "os detalhes"
        detail = "quem pagou, quem deve e de quê"
    elif cmd_strip.startswith("status"):
        title = "o status financeiro"
        detail = "lucro e totais"
    else:
        title = "o resumo"
        detail = "lucro e totais da planilha"
    return (
        f"⏳ *Gerando {title}...*\n"
        f"Lendo a planilha ({detail}). Aguarde — não precisa clicar de novo."
    )


def _notify_dashboard_loading(chat_id: int | str, cmd_strip: str) -> int | None:
    send_chat_action(chat_id, "typing")
    send_chat_action(chat_id, "upload_photo")
    return send_message(chat_id, _dashboard_wait_message(cmd_strip), parse_mode="Markdown")


def _requests_get_with_retry(
    url: str,
    *,
    params: dict | None = None,
    timeout: float | tuple[float, float] = (15, 120),
    max_attempts: int = 3,
    label: str = "requisicao",
) -> requests.Response:
    """GET com retentativas em falhas de rede/timeout (comum em áudios longos)."""
    last_exc: Exception | None = None
    for attempt in range(1, max_attempts + 1):
        try:
            response = requests.get(url, params=params, timeout=timeout)
            response.raise_for_status()
            return response
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as exc:
            last_exc = exc
            if attempt >= max_attempts:
                break
            wait_s = min(2 ** attempt, 8)
            print(f"[Telegram] {label} falhou (tentativa {attempt}/{max_attempts}): {exc}. Retentando em {wait_s}s...")
            time.sleep(wait_s)
    raise RuntimeError(
        f"Falha ao conectar com o Telegram apos {max_attempts} tentativas ({label}). "
        "Verifique sua internet e tente enviar o audio novamente."
    ) from last_exc


def _transcribe_voice_file_id(
    file_id: str,
    *,
    duration_sec: int = 0,
    hinted_file_size: int = 0,
) -> str:
    """Baixa e transcreve áudio do Telegram. Levanta TranscriptionError."""
    audio_path = download_telegram_voice(
        file_id,
        duration_sec=duration_sec,
        hinted_file_size=hinted_file_size,
    )
    try:
        whisper_model = os.getenv("WHISPER_MODEL", "small").strip() or "small"
        return (transcribe_audio(audio_path, model_size=whisper_model) or "").strip()
    finally:
        try:
            Path(audio_path).unlink(missing_ok=True)
        except Exception:
            pass


def download_telegram_voice(
    file_id: str,
    *,
    duration_sec: int = 0,
    hinted_file_size: int = 0,
) -> str:
    """Baixa o áudio de uma mensagem de voz do Telegram; retorna caminho do arquivo temporário (.ogg)."""
    r = _requests_get_with_retry(
        f"{BASE_URL}/getFile",
        params={"file_id": file_id},
        timeout=(15, 60),
        label="getFile",
    )
    data = r.json()
    if not data.get("ok"):
        raise RuntimeError("getFile falhou")
    result = data.get("result", {}) or {}
    file_path = result.get("file_path")
    if not file_path:
        raise RuntimeError("file_path nao retornado")
    file_size = int(result.get("file_size") or hinted_file_size or 0)
    download_read_timeout = _audio_download_read_timeout(
        file_size=file_size,
        duration_sec=duration_sec,
    )
    url = f"https://api.telegram.org/file/bot{TELEGRAM_BOT_TOKEN}/{file_path}"
    audio_r = _requests_get_with_retry(
        url,
        timeout=(20, download_read_timeout),
        label="download do audio",
    )
    suffix = ".ogg" if "ogg" in (file_path or "").lower() else ".oga"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp.write(audio_r.content)
    tmp.close()
    return tmp.name


def get_updates(offset: int | None = None) -> list[dict]:
    """Long polling: busca novas mensagens (timeout 30s)."""
    if not TELEGRAM_BOT_TOKEN:
        return []
    params = {"timeout": 30}
    if offset is not None:
        params["offset"] = offset
    try:
        r = requests.get(f"{BASE_URL}/getUpdates", params=params, timeout=35)
        if r.status_code != 200:
            return []
        data = r.json()
        if not data.get("ok"):
            return []
        return data.get("result", [])
    except Exception as e:
        print(f"[Telegram] Erro getUpdates: {e}")
        return []


def _print_bot_identity() -> None:
    """Mostra no terminal qual bot esta conectado."""
    try:
        r = requests.get(f"{BASE_URL}/getMe", timeout=10)
        if r.status_code != 200:
            print(f"[Telegram] Aviso: API retornou HTTP {r.status_code}", flush=True)
            return
        data = r.json()
        if not data.get("ok"):
            print(f"[Telegram] Aviso: token invalido ou bot inacessivel.", flush=True)
            return
        bot = data.get("result", {})
        name = bot.get("first_name", "?")
        username = bot.get("username", "")
        label = f"@{username}" if username else name
        print(f"[Telegram] Conectado: {label} ({name})", flush=True)
    except Exception as e:
        print(f"[Telegram] Aviso: nao foi possivel validar bot na API: {e}", flush=True)


def _format_currency_pt(value: float) -> str:
    txt = f"{float(value):,.2f}"
    txt = txt.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {txt}"


def _build_scheduled_reminder_message(item: dict) -> tuple[str, list[str]]:
    """Monta texto do lembrete e lista de tipos enviados (entrega/cobranca)."""
    sale_id = item.get("id venda") or ""
    cliente = item.get("cliente", "")
    desc = item.get("descricao", "")
    pending_amount = float(item.get("pending_amount_num") or 0.0)
    needs_delivery = bool(item.get("needs_delivery_reminder"))
    needs_payment = bool(item.get("needs_payment_reminder"))
    kinds: list[str] = []
    if needs_delivery:
        kinds.append("entrega")
    if needs_payment:
        kinds.append("cobranca")

    if needs_delivery and needs_payment:
        header = "⏰ *Lembrete do dia — entrega e cobrança*"
    elif needs_payment:
        header = "💰 *Cobrança — valor pendente hoje*"
    else:
        header = "🚚 *Lembrete de entrega hoje*"

    lines = [
        header,
        "",
        f"🧾 ID VENDA: *{sale_id}*",
        f"👤 Cliente: *{cliente}*",
        f"📦 Produto: *{desc}*",
        "",
    ]
    if needs_delivery:
        lines.append("🚚 *Entrega:* ainda não marcada como entregue.")
        lines.append(f"Quando entregar, envie: `FINALIZAR ID VENDA {sale_id}`")
        lines.append("")
    elif needs_payment and item.get("service_finalized"):
        lines.append("✅ *Entrega:* já marcada como entregue.")
        lines.append("")

    if needs_payment:
        lines.append(f"💰 *Pendente:* {_format_currency_pt(pending_amount)} — receber hoje.")
        lines.append(f"Quando receber, envie: `ID VENDA {sale_id} pagou`")

    return "\n".join(lines), kinds


def _build_overdue_delivery_message(item: dict) -> str:
    sale_id = item.get("id venda") or ""
    cliente = item.get("cliente", "") or "-"
    desc = item.get("descricao", "") or "-"
    delivery_txt = item.get("data entrega") or "-"
    today_txt = item.get("reference_today") or today_local().strftime("%d/%m/%Y")
    days = int(item.get("days_overdue") or 0)
    if days <= 1:
        days_txt = "1 dia"
    else:
        days_txt = f"{days} dias"
    return (
        "🚚 *Confirme a entrega — planilha pendente*\n\n"
        f"📆 *Hoje:* {today_txt}\n"
        f"A *Data de Entrega* de *{cliente}* (ID VENDA *{sale_id}*) era *{delivery_txt}* "
        f"— já passou há {days_txt} — e ainda está *amarela* (não entregue).\n\n"
        f"📦 Produto: *{desc}*\n\n"
        "⚠️ *O que aconteceu? Preciso que você confirme:*\n"
        "✅ *Já entregou?* → responda *SIM* (marco *verde* na planilha)\n"
        f"   ou envie: `ID VENDA {sale_id} foi entregue`\n"
        "❌ *Ainda não entregou?* → responda *NÃO*\n\n"
        "_Enquanto estiver amarelo e a data tiver passado, continuo cobrando._"
    )


def _register_overdue_pending(
    pending_overdue_delivery: dict,
    chat_id: int | str,
    sale_id: str,
    customer: str,
) -> None:
    bucket = pending_overdue_delivery.setdefault(chat_id, {})
    bucket[str(sale_id).strip()] = {"customer": customer or ""}


def _resolve_overdue_sale_id(text: str, bucket: dict[str, dict]) -> str | None:
    if not bucket:
        return None
    hinted = _extract_target_sale_id_for_updates(text or "")
    if hinted:
        hint = str(hinted).strip()
        for key in bucket:
            if key == hint or key.lstrip("0") == hint.lstrip("0"):
                return key
    if len(bucket) == 1:
        return next(iter(bucket))
    digits = re.findall(r"\b\d{2,4}\b", text or "")
    for raw in digits:
        for candidate in (raw, raw.zfill(3)):
            if candidate in bucket:
                return candidate
    return None


def _try_handle_overdue_delivery_reply(
    chat_id: int | str,
    text: str,
    pending_overdue_delivery: dict,
    pending_preview: dict | None = None,
) -> bool:
    bucket = pending_overdue_delivery.get(chat_id)
    if not bucket or not (text or "").strip():
        return False
    lower = text.strip().lower()
    preview_confirm = ("sim", "confirmar", "ok", "confirmo", "pode salvar", "salvar")
    if pending_preview is not None and chat_id in pending_preview and lower in preview_confirm:
        return False

    negative = any(
        tok in lower
        for tok in ("nao", "não", "ainda nao", "ainda não", "ainda nao entreg", "ainda não entreg")
    )
    positive = any(
        tok in lower
        for tok in (
            "sim",
            "ok",
            "confirmo",
            "pode atualizar",
            "pode marcar",
            "foi entregue",
            "foi entreg",
            "ja entregue",
            "já entregue",
            "entregue",
            "entregou",
            "finalizado",
            "finalizada",
        )
    )
    if not negative and not positive:
        return False

    if positive:
        multi_ids = extract_sale_ids_list_for_updates(text)
        targets = [sid for sid in multi_ids if sid in bucket]
        if len(targets) > 1:
            if not _spreadsheet_ready_for_bot():
                send_message(chat_id, "Planilha não encontrada.")
                return True
            try:
                svc = _open_spreadsheet_for_bot()
                with _spreadsheet_saving_timer(chat_id, "delivery_finalize", batch=True):
                    ok_ids: list[str] = []
                    fail_ids: list[str] = []
                    for sid in targets:
                        if svc.finalize_service(sid):
                            ok_ids.append(sid)
                            bucket.pop(sid, None)
                        else:
                            fail_ids.append(sid)
                if not bucket:
                    pending_overdue_delivery.pop(chat_id, None)
                lines = []
                if ok_ids:
                    lines.append(
                        f"✅ *Entregas confirmadas:* {', '.join(f'*{s}*' for s in ok_ids)} — verde na planilha."
                    )
                if fail_ids:
                    lines.append(f"⚠️ Não achei na planilha: {', '.join(fail_ids)}")
                send_message(
                    chat_id,
                    "\n".join(lines) or "Nenhuma entrega atualizada.",
                    parse_mode="Markdown",
                    reply_markup=MAIN_MENU_KEYBOARD,
                )
            except Exception as exc:
                send_message(chat_id, f"Erro ao atualizar: {exc}", reply_markup=MAIN_MENU_KEYBOARD)
            return True

    sale_id = _resolve_overdue_sale_id(text, bucket)
    if not sale_id:
        lines = [
            "Você tem *várias entregas* amarelas aguardando confirmação:",
            "",
        ]
        for sid, info in sorted(bucket.items()):
            nome = info.get("customer") or "-"
            lines.append(f"• ID VENDA *{sid}* — {nome}")
        lines.extend(
            [
                "",
                "Responda, por exemplo:",
                "`SIM ID VENDA 002` — já entregue",
                "`NÃO ID VENDA 003` — ainda não entregue",
            ]
        )
        send_message(chat_id, "\n".join(lines), parse_mode="Markdown", reply_markup=MAIN_MENU_KEYBOARD)
        return True

    if negative and not positive:
        send_message(
            chat_id,
            f"Ok, ID VENDA *{sale_id}* continua *amarelo* (não entregue).\n"
            f"Quando entregar, responda *SIM* ou `ID VENDA {sale_id} foi entregue`.",
            parse_mode="Markdown",
            reply_markup=MAIN_MENU_KEYBOARD,
        )
        return True

    if not _spreadsheet_ready_for_bot():
        send_message(chat_id, "Planilha não encontrada.")
        return True
    try:
        svc = _open_spreadsheet_for_bot()
        with _spreadsheet_saving_timer(chat_id, "delivery_finalize"):
            ok = svc.finalize_service(sale_id)
            bucket.pop(sale_id, None)
        if not bucket:
            pending_overdue_delivery.pop(chat_id, None)
        if ok:
            send_message(
                chat_id,
                f"✅ *Entrega confirmada!*\n\nID VENDA *{sale_id}* marcado como entregue — *verde* na planilha.",
                parse_mode="Markdown",
                reply_markup=MAIN_MENU_KEYBOARD,
            )
        else:
            send_message(
                chat_id,
                f"Não achei ID VENDA *{sale_id}* na planilha.",
                parse_mode="Markdown",
                reply_markup=MAIN_MENU_KEYBOARD,
            )
    except Exception as exc:
        send_message(chat_id, f"Erro ao atualizar: {exc}", reply_markup=MAIN_MENU_KEYBOARD)
    return True


def _workbook_path_for_bot() -> str:
    from src.spreadsheet_factory import get_workbook_path

    return get_workbook_path()


def _spreadsheet_ready_for_bot() -> bool:
    from src.spreadsheet_factory import spreadsheet_is_ready

    return spreadsheet_is_ready()


def _open_spreadsheet_for_bot():
    from src.spreadsheet_factory import get_spreadsheet_service

    return get_spreadsheet_service()


def _is_sale_already_delivered(sale_id: str | None) -> bool:
    """Consulta a planilha: Data de Entrega verde = serviço já entregue."""
    sale_id = str(sale_id or "").strip()
    if not sale_id:
        return False
    if not _spreadsheet_ready_for_bot():
        return False
    try:
        return _open_spreadsheet_for_bot().is_sale_delivered(sale_id)
    except Exception:
        return False


def _should_ask_delivery_date(parse_result, cmd) -> bool:
    """
    Pergunta data de entrega só em venda nova com valor.
    Atualizações (custo, pagamento, status etc.) e vendas já entregues não perguntam.
    """
    if parse_result.intent in (
        "status_update",
        "payment_update",
        "delivery_update",
        "delivery_finalize",
        "sale_delete",
        "material_update",
        "sale_value_update",
        "mixed_update",
    ):
        return False

    has_material = bool(
        (getattr(cmd, "material_cost", None) or 0) > 0
        or getattr(cmd, "material_allocations", None)
    )
    sale_id = getattr(cmd, "sale_id", None)
    if has_material and sale_id:
        return False

    if cmd.service_due_date is not None:
        return False
    if (cmd.total_value or 0.0) <= 0.01:
        return False

    pending_amount = 0.0
    for p in cmd.payments:
        if (p.status or "").strip().lower() != "pago":
            pending_amount += float(p.value or 0.0)
    if pending_amount > 0.01:
        return False

    if sale_id and _is_sale_already_delivered(str(sale_id)):
        return False

    return True


def _build_dashboard_reply(cmd_strip: str) -> tuple[str, str | None]:
    """
    Gera texto + imagem abrindo a planilha uma única vez (Prévia/Status/Resumo).
    Retorna (legenda, caminho_imagem).
    """
    from src.excel_store import SpreadsheetService

    if not _spreadsheet_ready_for_bot():
        return process_command(cmd_strip, origin="telegram"), None

    try:
        svc = _open_spreadsheet_for_bot()
    except Exception as exc:
        return _spreadsheet_error_message(exc), None

    workbook_path = _workbook_path_for_bot()
    wb = svc._open_workbook(data_only=False)
    img_path: str | None = None
    try:
        if cmd_strip.startswith(("prévia", "previa")):
            reply = svc.get_sales_preview(wb=wb)
            img_path = _maybe_build_sales_snippet_image(
                workbook_path,
                wb=wb,
                svc=svc,
                title_main="TOTAL DE VENDAS - Visão atual",
            )
        elif cmd_strip.startswith(("detalhes", "detalhe")):
            # Detalhes vão como mensagem de texto (não imagem).
            return svc.get_planilha_details(wb=wb), None
        elif cmd_strip.startswith(("status", "resumo", "planilha")):
            reply = svc.get_planilha_summary(wb=wb)
            img_path = _maybe_build_status_table_image(workbook_path, wb=wb, svc=svc)
        else:
            return process_command(cmd_strip, origin="telegram"), None
    finally:
        wb.close()

    if not img_path:
        img_path = _maybe_build_text_image_png(reply)
        if img_path:
            return "", img_path
    return reply, img_path


def _is_pdf_report_request(text: str) -> bool:
    normalized = _normalize_menu_button(text)
    return normalized in {
        "/relatorio",
        "/relatório",
        "relatorio",
        "relatório",
        "relatorio pdf",
        "relatório pdf",
        "pdf do mes",
        "pdf do mês",
        "relatorio mensal",
        "relatório mensal",
    }


def _send_planilha_pdf(chat_id: int | str, display_name: str = "") -> None:
    if not _spreadsheet_ready_for_bot():
        send_message(
            chat_id,
            spreadsheet_setup_hint(),
            reply_markup=MAIN_MENU_KEYBOARD,
        )
        return
    send_chat_action(chat_id, "typing")
    try:
        stats = collect_planilha_stats()
    except Exception as exc:
        send_message(
            chat_id,
            _spreadsheet_error_message(exc),
            reply_markup=MAIN_MENU_KEYBOARD,
        )
        print(f"[Telegram] Erro no relatório PDF: {exc}")
        return
    if not stats:
        send_message(chat_id, "Planilha não encontrada.", reply_markup=MAIN_MENU_KEYBOARD)
        return
    pdf_path = build_planilha_report_pdf(stats, user_name=display_name)
    if not pdf_path:
        send_message(
            chat_id,
            "⚠️ Não consegui gerar o PDF. Rode `python run_project.py` para instalar fpdf2 e matplotlib.",
            reply_markup=MAIN_MENU_KEYBOARD,
        )
        return
    month_label = str(stats.get("mes_label") or "mês atual")
    try:
        send_document(
            chat_id,
            pdf_path,
            caption=f"📄 *Relatório da planilha — {month_label}*",
            filename=f"relatorio_{month_label.replace(' ', '_').lower()}.pdf",
            parse_mode="Markdown",
            reply_markup=MAIN_MENU_KEYBOARD,
        )
    finally:
        try:
            Path(pdf_path).unlink(missing_ok=True)
        except Exception:
            pass


def _maybe_enhance_parse_with_llm(
    text: str,
    parse_result,
    *,
    last_sale_id: str | None = None,
):
    if not parse_result.missing_fields or not llm_fallback_enabled():
        return parse_result
    partial: dict[str, str] = {}
    if last_sale_id:
        partial["sale_id"] = last_sale_id
    cmd = getattr(parse_result, "command", None)
    if cmd and getattr(cmd, "customer", None):
        partial["customer"] = str(cmd.customer)
    if cmd and getattr(cmd, "product_id", None):
        partial["product"] = str(cmd.product_id)
    llm_result = try_llm_parse(
        text,
        reference_date=today_local(),
        partial=partial or None,
        last_sale_id=last_sale_id,
    )
    if llm_result and not llm_result.missing_fields:
        print("[Telegram] Parser LLM fallback interpretou a frase.")
        return llm_result
    if llm_result and len(llm_result.missing_fields) < len(parse_result.missing_fields):
        print("[Telegram] Parser LLM fallback reduziu campos faltantes.")
        return llm_result
    return parse_result


def _handle_quick_dashboard_command(
    chat_id: int | str,
    text: str,
    *,
    pending_preview: dict,
    quick_menu_processing: set[int | str],
    quick_menu_cooldown: dict[int | str, float],
) -> bool:
    """Prévia / Resumo do menu rápido. Retorna True se tratou o comando."""
    cmd_strip = _normalize_menu_button(text)
    if not _is_quick_dashboard_command(cmd_strip):
        return False
    if not cmd_strip.startswith(("status", "resumo", "planilha", "prévia", "previa", "detalhes", "detalhe")):
        return False

    now = time.time()
    if chat_id in quick_menu_processing or now < quick_menu_cooldown.get(chat_id, 0.0):
        send_message(
            chat_id,
            "⏳ *Ainda preparando a resposta anterior...*\n"
            "Aguarde alguns segundos — não precisa tocar no botão de novo.",
            parse_mode="Markdown",
            reply_markup=MAIN_MENU_KEYBOARD,
        )
        send_chat_action(chat_id, "typing")
        return True

    quick_menu_processing.add(chat_id)
    pending_preview.pop(chat_id, None)
    wait_message_id: int | None = None
    img_path: str | None = None
    is_details = cmd_strip.startswith(("detalhes", "detalhe"))
    try:
        if is_details:
            send_chat_action(chat_id, "typing")
            wait_message_id = send_message(
                chat_id,
                _dashboard_wait_message(cmd_strip),
                parse_mode="Markdown",
            )
            caption, img_path = _run_with_chat_action(
                chat_id,
                "typing",
                lambda: _build_dashboard_reply(cmd_strip),
            )
            send_message(
                chat_id,
                caption or "Nenhum detalhe disponível.",
                parse_mode="Markdown",
                reply_markup=MAIN_MENU_KEYBOARD,
            )
        else:
            wait_message_id = _notify_dashboard_loading(chat_id, cmd_strip)
            caption, img_path = _run_with_chat_action(
                chat_id,
                "upload_photo",
                lambda: _build_dashboard_reply(cmd_strip),
            )
            send_reply(
                chat_id,
                caption,
                parse_mode="Markdown",
                reply_markup=(
                    RESUMO_DETAILS_INLINE_KEYBOARD
                    if cmd_strip.startswith(("status", "resumo", "planilha"))
                    else MAIN_MENU_KEYBOARD
                ),
                image_path=img_path,
            )
    except Exception as exc:
        send_message(
            chat_id,
            _spreadsheet_error_message(exc),
            reply_markup=MAIN_MENU_KEYBOARD,
        )
    finally:
        delete_chat_message(chat_id, wait_message_id)
        if img_path:
            try:
                Path(img_path).unlink(missing_ok=True)
            except Exception:
                pass
        quick_menu_processing.discard(chat_id)
        quick_menu_cooldown[chat_id] = time.time() + 3.0
    return True


def _handle_resumo_detalhes_callback(
    chat_id: int | str,
    callback_query_id: str,
    *,
    pending_preview: dict,
    quick_menu_processing: set[int | str],
    quick_menu_cooldown: dict[int | str, float],
) -> bool:
    """Botão inline 'Ver detalhes' no card do Resumo — responde em texto."""
    now = time.time()
    if chat_id in quick_menu_processing or now < quick_menu_cooldown.get(chat_id, 0.0):
        answer_callback_query(
            callback_query_id,
            "Ainda preparando a resposta anterior...",
        )
        return True

    quick_menu_processing.add(chat_id)
    pending_preview.pop(chat_id, None)
    wait_message_id: int | None = None
    try:
        answer_callback_query(callback_query_id, "Gerando detalhes...")
        send_chat_action(chat_id, "typing")
        wait_message_id = send_message(
            chat_id,
            _dashboard_wait_message("detalhes"),
            parse_mode="Markdown",
        )
        caption, _img = _run_with_chat_action(
            chat_id,
            "typing",
            lambda: _build_dashboard_reply("detalhes"),
        )
        send_message(
            chat_id,
            caption or "Nenhum detalhe disponível.",
            parse_mode="Markdown",
            reply_markup=MAIN_MENU_KEYBOARD,
        )
    except Exception as exc:
        send_message(
            chat_id,
            f"Erro ao gerar detalhes: {exc}",
            reply_markup=MAIN_MENU_KEYBOARD,
        )
    finally:
        delete_chat_message(chat_id, wait_message_id)
        quick_menu_processing.discard(chat_id)
        quick_menu_cooldown[chat_id] = time.time() + 3.0
    return True


def _send_batch_update_preview(
    chat_id: int | str,
    batch_results: list,
    *,
    original_text: str,
    pending_preview: dict,
    not_found: list[str] | None = None,
) -> None:
    from src.bot_processor import build_preview

    missing = not_found or []
    if len(batch_results) == 1:
        parse_result = batch_results[0]
        preview_text = build_preview(parse_result)
        if missing:
            preview_text += f"\n\nIDs não encontrados: {', '.join(missing)}"
        send_message(chat_id, preview_text, parse_mode="Markdown")
        pending_preview[chat_id] = {
            "parse_result": parse_result,
            "original_text": original_text,
            "origin": "telegram",
        }
        return

    parts: list[str] = []
    for idx, pr in enumerate(batch_results, start=1):
        parts.append(f"*Item {idx}:*\n{build_preview(pr)}")
    if missing:
        parts.append(f"IDs não encontrados: {', '.join(missing)}")
    parts.append("\nResponda *SIM* para confirmar o lote ou *NÃO* para cancelar.")
    send_message(chat_id, "\n\n".join(parts[:8]), parse_mode="Markdown")
    pending_preview[chat_id] = {
        "batch_parse_results": batch_results,
        "original_text": original_text,
        "origin": "telegram",
    }


def _try_handle_batch_sale_updates(
    chat_id: int | str,
    text: str,
    pending_preview: dict,
) -> bool:
    """Vários ID VENDA na mesma mensagem — entrega ou pagamento total."""
    lower = text.strip().lower()
    sale_tokens = (
        "vendi",
        "fechei",
        "fechamos",
        "acabei de fazer uma venda",
        "fiz uma venda",
        "fiz um",
        "comprou",
    )
    if any(tok in lower for tok in sale_tokens):
        return False

    ids = extract_sale_ids_list_for_updates(text)
    if len(ids) < 2:
        return False

    nums = [int(n) for n in re.findall(r"\b\d{1,6}\b", lower)]
    has_amount = any(n >= 100 for n in nums)

    is_payment = ("pagou" in lower or "pago" in lower or "quitou" in lower) and not has_amount
    is_delivery = (
        _is_service_delivery_finalized(text)
        or bool(re.search(r"\b(?:foi\s+)?entreg", lower))
        or bool(re.search(r"\bentregues?\b", lower))
        or "finaliz" in lower
        or bool(re.search(r"\batualiz\w+.*entreg", lower))
    ) and not is_payment

    if not is_payment and not is_delivery:
        return False

    if not _spreadsheet_ready_for_bot():
        return False

    batch_results: list = []
    not_found: list[str] = []
    today = today_local()

    if is_delivery:
        for sale_id in ids:
            pr = parse_message(f"ID VENDA {sale_id} foi entregue hoje", reference_date=today)
            if pr.missing_fields:
                not_found.append(sale_id)
            else:
                batch_results.append(pr)
        kind = "entrega"
    else:
        try:
            svc = _open_spreadsheet_for_bot()
        except Exception:
            return False
        for cliente_id in ids:
            target_sale_id = None
            sale_rows = svc.get_pending_sale_by_sale_id(
                cliente_id, max_rows_scan=500, include_paid=True
            )
            if sale_rows:
                target_sale_id = sale_rows[0].get("sale_id")
            else:
                pend_rows = svc.get_pending_sales_by_customer(
                    cliente_id, max_rows_scan=500, include_paid=True
                )
                if pend_rows:
                    target_sale_id = pend_rows[0].get("sale_id")
            if target_sale_id:
                pr = parse_message(f"ID VENDA {target_sale_id} pagou", reference_date=today)
                batch_results.append(pr)
            else:
                not_found.append(cliente_id)
        kind = "pagamento"

    if not batch_results:
        send_message(
            chat_id,
            f"Não consegui montar o lote de {kind} para os IDs: {', '.join(ids)}.",
            reply_markup=MAIN_MENU_KEYBOARD,
        )
        return True

    _send_batch_update_preview(
        chat_id,
        batch_results,
        original_text=text,
        pending_preview=pending_preview,
        not_found=not_found,
    )
    return True


def _try_handle_batch_delete(
    chat_id: int | str,
    text: str,
    pending_preview: dict,
) -> bool:
    """Vários ID VENDA na mesma mensagem — exclusão em lote."""
    if not _is_sale_delete_request(text):
        return False

    ids = extract_delete_sale_ids_list(text)
    if len(ids) < 2:
        return False

    today = today_local()
    batch_results: list = []
    not_found: list[str] = []

    for sale_id in ids:
        pr = parse_message(f"apagar id venda {sale_id}", reference_date=today)
        if pr.missing_fields or pr.intent != "sale_delete":
            not_found.append(sale_id)
            continue
        batch_results.append(pr)

    if not batch_results:
        send_message(
            chat_id,
            f"Não consegui montar a exclusão em lote para os IDs: {', '.join(ids)}.",
            reply_markup=MAIN_MENU_KEYBOARD,
        )
        return True

    _send_batch_update_preview(
        chat_id,
        batch_results,
        original_text=text,
        pending_preview=pending_preview,
        not_found=not_found,
    )
    return True


def _try_handle_multi_sales_same_customer(
    chat_id: int | str,
    text: str,
    pending_preview: dict,
    last_sale_id_by_chat: dict | None = None,
) -> bool:
    """Várias vendas (produtos distintos) para o mesmo cliente na mesma mensagem."""
    if detect_intent(text) != "sale":
        return False

    batch = parse_multi_sales_message(text, reference_date=today_local())
    if not batch or len(batch) < 2:
        return False

    if any(pr.missing_fields for pr in batch):
        return False

    if last_sale_id_by_chat is not None:
        last_sale_id_by_chat.pop(chat_id, None)

    from src.bot_processor import build_preview

    customer = (batch[0].command.customer or "").strip()
    intro = (
        f"📋 *Entendi {len(batch)} vendas para o mesmo cliente"
        + (f" ({customer})*" if customer else "*")
        + " — confira cada uma:"
    )
    parts: list[str] = [intro, ""]
    for idx, pr in enumerate(batch, start=1):
        parts.append(f"*Venda {idx}:*\n{build_preview(pr)}")
    parts.append("\nResponda *SIM* para registrar todas ou *NÃO* para cancelar.")
    parts.append(
        "\n💡 *Para corrigir uma venda:* "
        "*Venda 1: Entrada: 3000* ou *na venda um deu 3000 de entrada*."
    )
    parts.append(
        "\n🔢 *Cada venda receberá um ID VENDA único* ao confirmar (ex.: 003, 004)."
    )
    send_message(chat_id, "\n\n".join(parts[:12]), parse_mode="Markdown")
    pending_preview[chat_id] = {
        "batch_parse_results": batch,
        "original_text": text,
        "origin": "telegram",
    }
    return True


def _resend_multi_sales_batch_preview(
    chat_id: int | str,
    batch: list,
    pending: dict,
    pending_preview: dict,
) -> None:
    from src.bot_processor import build_preview

    customer = (batch[0].command.customer or "").strip() if batch else ""
    intro = (
        f"📋 *Prévia atualizada — {len(batch)} vendas"
        + (f" para {customer}*" if customer else "*")
        + ":"
    )
    parts: list[str] = [intro, ""]
    for idx, pr in enumerate(batch, start=1):
        parts.append(f"*Venda {idx}:*\n{build_preview(pr)}")
    parts.append("\nResponda *SIM* para registrar todas ou *NÃO* para cancelar.")
    parts.append(
        "\n💡 *Para corrigir outra venda:* "
        "*Venda 2: Entrada: 800* ou *na venda dois valor total 1500*."
    )
    send_message(chat_id, "\n\n".join(parts[:12]), parse_mode="Markdown")
    pending["batch_parse_results"] = batch
    pending_preview[chat_id] = pending


def _process_scheduled_reminders(tracker) -> None:
    """Envia lembretes no dia da entrega: a partir das 6h, a cada 2h."""
    from src.excel_store import SpreadsheetService
    from src.reminder_scheduler import current_reminder_slot
    from src.workbook_paths import default_workbook_path

    slot = current_reminder_slot()
    if not slot:
        return
    today, slot_hour = slot

    if not _spreadsheet_ready_for_bot():
        return

    try:
        svc = _open_spreadsheet_for_bot()
    except Exception:
        return

    wb = svc._open_workbook(data_only=True)
    try:
        due = svc.list_due_reminders(wb, today)
    finally:
        wb.close()

    if not due:
        return

    admin_chat = os.getenv("ADMIN_CHAT_ID", "").strip()
    for item in due[:20]:
        sale_id = str(item.get("id venda") or "").strip()
        if not sale_id:
            continue

        send_delivery = bool(item.get("needs_delivery_reminder")) and not tracker.was_sent(
            "entrega", sale_id, today, slot_hour
        )
        send_payment = bool(item.get("needs_payment_reminder")) and not tracker.was_sent(
            "cobranca", sale_id, today, slot_hour
        )
        if not send_delivery and not send_payment:
            continue

        preview_item = dict(item)
        preview_item["needs_delivery_reminder"] = send_delivery
        preview_item["needs_payment_reminder"] = send_payment
        msg_txt, kinds = _build_scheduled_reminder_message(preview_item)
        if not kinds:
            continue

        chat_target = (item.get("chat id") or "").strip()
        sent = False
        if chat_target:
            sent = send_message(chat_target, msg_txt, parse_mode="Markdown")
        if admin_chat:
            send_message(admin_chat, msg_txt, parse_mode="Markdown")
            sent = True
        if not sent:
            continue

        for kind in kinds:
            tracker_key = "entrega" if kind == "entrega" else "cobranca"
            tracker.mark_sent(tracker_key, sale_id, today, slot_hour)
        print(
            f"[Telegram] Lembrete enviado ID VENDA {sale_id} "
            f"({', '.join(kinds)}) slot {slot_hour:02d}h"
        )


def _process_overdue_deliveries(tracker, pending_overdue_delivery: dict) -> None:
    """
    Cobra entregas amarelas cuja data já passou em relação a hoje (today_local()).
    Roda a cada ~1 min; na subida do bot já verifica na hora.
    """
    from src.excel_store import SpreadsheetService
    from src.reminder_scheduler import overdue_reminder_slot

    now = now_local()
    today, slot_hour = overdue_reminder_slot(now)

    if not _spreadsheet_ready_for_bot():
        return

    try:
        svc = _open_spreadsheet_for_bot()
    except Exception:
        return

    wb = svc._open_workbook(data_only=False)
    try:
        overdue = svc.list_overdue_deliveries(wb, today)
    finally:
        wb.close()

    if not overdue:
        return

    print(
        f"[Telegram] {len(overdue)} entrega(s) amarela(s) com data passada "
        f"(referência hoje: {today.strftime('%d/%m/%Y')})"
    )

    admin_chat = os.getenv("ADMIN_CHAT_ID", "").strip()
    for item in overdue[:15]:
        sale_id = str(item.get("id venda") or "").strip()
        if not sale_id or tracker.was_sent("atraso", sale_id, today, slot_hour):
            continue

        msg_txt = _build_overdue_delivery_message(item)
        chat_target = (item.get("chat id") or "").strip()
        sent = False
        if chat_target:
            if send_message(chat_target, msg_txt, parse_mode="Markdown"):
                _register_overdue_pending(
                    pending_overdue_delivery,
                    chat_target,
                    sale_id,
                    str(item.get("cliente") or ""),
                )
                sent = True
        if admin_chat:
            send_message(admin_chat, msg_txt, parse_mode="Markdown")
            _register_overdue_pending(
                pending_overdue_delivery,
                admin_chat,
                sale_id,
                str(item.get("cliente") or ""),
            )
            sent = True
        if not sent:
            continue

        tracker.mark_sent("atraso", sale_id, today, slot_hour)
        print(
            f"[Telegram] Cobrança entrega amarela ID VENDA {sale_id} slot {slot_hour:02d}h"
        )


def run_polling() -> None:
    """Loop principal: processa mensagens e responde."""
    _acquire_single_instance_lock()
    if not TELEGRAM_BOT_TOKEN:
        print("[Telegram] Configure TELEGRAM_BOT_TOKEN no .env (token do @BotFather).")
        return

    workbook = get_default_workbook()
    print("[Telegram] Bot iniciado. Aguardando mensagens... (Ctrl+C para parar)", flush=True)
    if workbook:
        print(f"[Telegram] Planilha ativa: {workbook}", flush=True)
    _print_bot_identity()
    print(f"[Telegram] {time.strftime('%H:%M:%S')} — polling ativo (atualiza a cada ~2 min)", flush=True)
    if llm_fallback_enabled():
        model = os.getenv("LLM_MODEL", "gpt-4o-mini").strip() or "gpt-4o-mini"
        print(f"[Telegram] Parser LLM fallback ligado ({model}) — usado quando o regex falha.", flush=True)
    else:
        print(
            "[Telegram] Parser LLM fallback desligado "
            "(defina LLM_FALLBACK_ENABLED=1 e LLM_API_KEY no .env para ativar).",
            flush=True,
        )

    next_offset = None
    seen_update_ids: set[int] = set()
    seen_message_keys: set[tuple[int | str, int]] = set()
    # Prévia pendente por chat: só salva na planilha após o usuário confirmar (SIM)
    pending_preview: dict = {}
    pending_delivery: dict[int | str, dict] = {}
    # Cache simples por chat: último nome de Cliente informado com sucesso.
    last_customer_by_chat: dict[int | str, str] = {}
    last_sale_id_by_chat: dict[int | str, str] = {}
    pending_overdue_delivery: dict[int | str, dict] = {}
    pending_correct: dict[int | str, dict] = {}
    quick_menu_processing: set[int | str] = set()
    quick_menu_cooldown: dict[int | str, float] = {}
    last_reminder_check = 0.0
    last_status_ping = time.time()
    status_interval_s = 120.0
    from src.reminder_scheduler import ReminderSlotTracker

    reminder_tracker = ReminderSlotTracker()

    print(
        f"[Telegram] Verificação inicial de entregas amarelas "
        f"(data de hoje: {today_local().strftime('%d/%m/%Y')})..."
    )
    try:
        _process_overdue_deliveries(reminder_tracker, pending_overdue_delivery)
    except Exception as e:
        print(f"[Telegram] Erro na verificação inicial de entregas: {e}")

    while True:
        updates = get_updates(offset=next_offset)
        now = time.time()
        if not updates and now - last_status_ping >= status_interval_s:
            print(
                f"[Telegram] {time.strftime('%H:%M:%S')} — online, aguardando mensagens...",
                flush=True,
            )
            last_status_ping = now
        elif updates:
            last_status_ping = now
        for update in updates:
            uid = update.get("update_id", 0)
            if isinstance(uid, int) and uid in seen_update_ids:
                continue
            if isinstance(uid, int):
                seen_update_ids.add(uid)
                # Evitar crescimento infinito (mantém só os últimos ~2000)
                if len(seen_update_ids) > 2000:
                    for old in sorted(seen_update_ids)[:500]:
                        seen_update_ids.discard(old)
            next_offset = uid + 1

            callback = update.get("callback_query")
            if callback:
                callback_id = callback.get("id")
                callback_data = str(callback.get("data") or "").strip()
                callback_chat_id = (callback.get("message") or {}).get("chat", {}).get("id")
                if callback_data == "resumo_detalhes" and callback_chat_id:
                    _handle_resumo_detalhes_callback(
                        callback_chat_id,
                        str(callback_id or ""),
                        pending_preview=pending_preview,
                        quick_menu_processing=quick_menu_processing,
                        quick_menu_cooldown=quick_menu_cooldown,
                    )
                elif callback_data.startswith("corr:") and callback_chat_id:
                    msg_obj = callback.get("message") or {}
                    _handle_correct_callback(
                        callback_chat_id,
                        str(callback_id or ""),
                        callback_data,
                        pending_correct=pending_correct,
                        pending_preview=pending_preview,
                        last_sale_id_by_chat=last_sale_id_by_chat,
                        message_id=msg_obj.get("message_id"),
                    )
                elif callback_id:
                    answer_callback_query(str(callback_id))
                continue

            msg = update.get("message") or update.get("edited_message")
            if not msg:
                continue
            chat_id = msg.get("chat", {}).get("id")
            message_id = msg.get("message_id")
            if chat_id and isinstance(message_id, int):
                key = (chat_id, message_id)
                if key in seen_message_keys:
                    continue
                seen_message_keys.add(key)
                if len(seen_message_keys) > 4000:
                    # Remove a parte mais antiga (ordenação estável pelo tuple)
                    for old in sorted(seen_message_keys)[:1000]:
                        seen_message_keys.discard(old)
            text = (msg.get("text") or "").strip()
            from_voice = False

            # Áudio: baixar, transcrever e usar o texto como se fosse mensagem digitada
            voice_payload = msg.get("voice") or msg.get("audio")
            in_correct_alter = (
                chat_id
                and chat_id in pending_correct
                and pending_correct[chat_id].get("step") == "await_alter"
            )
            if not text and voice_payload:
                file_id = voice_payload.get("file_id")
                if file_id:
                    try:
                        duration_sec = int(voice_payload.get("duration") or 0)
                        hinted_size = int(voice_payload.get("file_size") or 0)
                        with _audio_processing_timer(chat_id, duration_sec):
                            if duration_sec >= 60:
                                print(
                                    f"[Telegram] Transcrevendo audio longo ({duration_sec}s) "
                                    f"com modelo {os.getenv('WHISPER_MODEL', 'small')}..."
                                )
                            text = _transcribe_voice_file_id(
                                str(file_id),
                                duration_sec=duration_sec,
                                hinted_file_size=hinted_size,
                            )
                            from_voice = True
                            preview_text = text
                            if len(preview_text) > 3500:
                                preview_text = preview_text[:3500] + "\n\n… _(texto truncado na exibição)_"
                            if in_correct_alter:
                                send_message(
                                    chat_id,
                                    f"🎤 *Áudio entendido para correção:*\n_{preview_text}_",
                                    parse_mode="Markdown",
                                )
                            else:
                                send_message(
                                    chat_id,
                                    "🎤 *Áudio entendido!*\n\n"
                                    f"{preview_text}\n\n"
                                    "Agora vou montar a prévia do lançamento. Confira os campos e responda *SIM* para salvar, "
                                    "ou envie a correção por texto.",
                                    parse_mode="Markdown",
                                    reply_markup=MAIN_MENU_KEYBOARD,
                                )
                    except TranscriptionError as e:
                        send_message(chat_id, f"Não consegui transcrever o áudio: {e}")
                        continue
                    except Exception as e:
                        err = str(e).lower()
                        if "timeout" in err or "timed out" in err:
                            hint = (
                                "A conexão com o Telegram demorou demais para baixar o áudio.\n"
                                "Tente de novo com internet estável, ou envie um áudio mais curto / digite o texto."
                            )
                        else:
                            hint = str(e)
                        send_message(chat_id, f"Erro ao processar o áudio: {hint}")
                        print(f"[Telegram] Erro áudio: {e}")
                        continue

            if not chat_id:
                continue

            if text and _try_handle_overdue_delivery_reply(
                chat_id, text, pending_overdue_delivery, pending_preview
            ):
                continue

            if not text:
                send_message(
                    chat_id,
                    "Envie um texto (ex.: vendi placa para o João por 2000), um áudio, ou toque em 📊 *Resumo*.",
                    reply_markup=MAIN_MENU_KEYBOARD,
                )
                continue

            lower_cmd = _normalize_menu_button(text)

            # Comandos especiais do Telegram: boas-vindas e ajuda (limpam prévia pendente) + menu
            if text in ("/start", "/help") or _normalize_menu_button(text) in ("ajuda", "start"):
                pending_preview.pop(chat_id, None)
                pending_correct.pop(chat_id, None)
                user_name = _user_display_name(msg.get("from"))
                send_message(
                    chat_id,
                    _help_text(user_name),
                    parse_mode="Markdown",
                    reply_markup=MAIN_MENU_KEYBOARD,
                )
                continue

            # Nova conversa: limpa prévia e contexto de ID VENDA
            if lower_cmd in ("nova conversa", "novo atendimento", "novo chat", "reiniciar"):
                _reset_chat_session(
                    chat_id,
                    pending_preview=pending_preview,
                    pending_delivery=pending_delivery,
                    pending_overdue_delivery=pending_overdue_delivery,
                    last_sale_id_by_chat=last_sale_id_by_chat,
                    last_customer_by_chat=last_customer_by_chat,
                    pending_correct=pending_correct,
                )
                user_name = _escape_md(_user_display_name(msg.get("from")))
                hello = f"Pronto, {user_name}! " if user_name else ""
                if lower_cmd.startswith("nova"):
                    msg_txt = (
                        f"🔄 {hello}*Nova conversa iniciada.*\n\n"
                        "O contexto do ID VENDA foi limpo. Pode registrar uma venda nova "
                        "ou informar outro ID quando quiser."
                    )
                else:
                    msg_txt = (
                        f"🔄 {hello}*Nova conversa iniciada.*\n\n"
                        "Contexto limpo. Pode enviar um novo comando ou tocar em *Prévia*."
                    )
                send_message(chat_id, msg_txt, parse_mode="Markdown", reply_markup=MAIN_MENU_KEYBOARD)
                continue

            # Menu Corrigir: alterar ou apagar lançamentos
            if _is_correct_menu_command(text):
                user_name = _user_display_name(msg.get("from"))
                _start_correct_flow(
                    chat_id,
                    pending_correct=pending_correct,
                    pending_preview=pending_preview,
                    user_name=user_name,
                )
                continue

            # Cancelar fluxo Corrigir com NÃO
            if lower_cmd in ("não", "nao", "cancelar", "cancela") and chat_id in pending_correct:
                pending_correct.pop(chat_id, None)
                send_message(
                    chat_id,
                    "❌ Correção cancelada. Pode continuar normalmente.",
                    parse_mode="Markdown",
                    reply_markup=MAIN_MENU_KEYBOARD,
                )
                continue

            # Correção por texto ou áudio (copiar modelo e editar)
            if _try_handle_correct_alter_input(
                chat_id,
                text,
                pending_correct=pending_correct,
                pending_preview=pending_preview,
                last_sale_id_by_chat=last_sale_id_by_chat,
                origin="audio" if from_voice else "telegram",
            ):
                continue

            if _try_handle_correct_wizard_text(chat_id, text, pending_correct=pending_correct):
                continue

            user = msg.get("from", {})
            username = user.get("username") or user.get("first_name") or "?"
            print(f"[Telegram] Mensagem de {username} ({chat_id}): {text[:60]!r}")

            try:
                # Finalizar serviço
                if text.strip().lower().startswith("finalizar") and "id venda" in text.strip().lower():
                    import re
                    sale_digits = re.findall(r"\d+", text)
                    if sale_digits:
                        sale_id = sale_digits[-1].zfill(3) if len(sale_digits[-1]) <= 3 else sale_digits[-1]
                        if not _spreadsheet_ready_for_bot():
                            send_message(chat_id, "Planilha não encontrada.", reply_markup=MAIN_MENU_KEYBOARD)
                            continue
                        try:
                            svc = _open_spreadsheet_for_bot()
                            ok = svc.finalize_service(sale_id)
                        except Exception as exc:
                            send_message(
                                chat_id,
                                _spreadsheet_error_message(exc),
                                reply_markup=MAIN_MENU_KEYBOARD,
                            )
                            continue
                        if ok:
                            send_message(
                                chat_id,
                                f"✅ *Entrega confirmada!*\n\nID VENDA *{sale_id}* salvo na planilha — marcado como *entregue* (verde).",
                                parse_mode="Markdown",
                                reply_markup=MAIN_MENU_KEYBOARD,
                            )
                        else:
                            send_message(
                                chat_id,
                                f"❌ Não achei ID VENDA *{sale_id}* na planilha.",
                                parse_mode="Markdown",
                            )
                        continue

                # Confirmação da prévia: salvar na planilha
                confirm_words = ("sim", "confirmar", "ok", "confirmo", "pode salvar", "salvar")
                if text.strip().lower() in confirm_words and chat_id in pending_preview:
                    pending = pending_preview.pop(chat_id)
                    # Fluxo em lote: múltiplos updates (ex.: "cliente id 003 e id 002 pagou")
                    batch_results = pending.get("batch_parse_results")
                    if batch_results:
                        is_delete_batch = all(
                            getattr(pr, "intent", "") == "sale_delete" for pr in batch_results
                        )
                        is_multi_sale_batch = (
                            not is_delete_batch
                            and len(batch_results) >= 2
                            and all(getattr(pr, "intent", "") == "sale" for pr in batch_results)
                        )
                        _notify_intent = (
                            "sale_delete" if is_delete_batch else ("sale" if is_multi_sale_batch else "")
                        )
                        with _spreadsheet_saving_timer(chat_id, _notify_intent, batch=True):
                            applied = 0
                            replies: list[str] = []
                            for pr in batch_results:
                                if getattr(pr, "intent", "") == "sale":
                                    sanitize_new_sale_identity(pr.command)
                                try:
                                    reply_item = apply_parse_result(
                                        pr,
                                        origin=pending.get("origin", "telegram"),
                                        original_text=pending.get("original_text", ""),
                                        chat_id=str(chat_id),
                                    )
                                    replies.append(reply_item)
                                    applied += 1
                                except Exception as e:
                                    replies.append(f"Falha em um item do lote: {e}")
                        if is_delete_batch:
                            msg_out = (
                                f"✅ Exclusão em lote concluída: {applied}/{len(batch_results)} registro(s).\n\n"
                                + "\n".join(replies[:6])
                            )
                        elif is_multi_sale_batch:
                            msg_out = (
                                f"✅ {applied}/{len(batch_results)} vendas registradas na planilha!\n\n"
                                + "\n".join(replies[:6])
                            )
                        else:
                            msg_out = (
                                f"Atualização em lote concluída: {applied}/{len(batch_results)} item(ns).\n\n"
                                + "\n".join(replies[:6])
                            )
                        send_message(chat_id, msg_out, parse_mode="Markdown", reply_markup=MAIN_MENU_KEYBOARD)
                        continue
                    parse_result = pending["parse_result"]
                    original_text = pending.get("original_text", "")
                    origin = pending.get("origin", "telegram")
                    _normalize_material_only_sale(parse_result)
                    cmd = parse_result.command
                    if not (cmd.customer or "").strip() and getattr(cmd, "sale_id", None):
                        loaded_customer = _load_customer_from_workbook(str(cmd.sale_id))
                        if loaded_customer:
                            cmd.customer = loaded_customer
                    # Se ainda faltar Cliente (nome), não salva em fluxos de venda nova.
                    # Para atualização de status, entrega, material etc., usa-se o ID VENDA.
                    if parse_result.intent not in (
                        "status_update",
                        "delivery_update",
                        "delivery_finalize",
                        "payment_update",
                        "sale_delete",
                        "material_update",
                        "sale_value_update",
                    ):
                        if "Cliente" in parse_result.missing_fields or not (cmd.customer or "").strip():
                            pending_preview[chat_id] = pending
                            send_message(
                                chat_id,
                                "Faltou o *Cliente* para salvar.\nEnvie por exemplo: `Cliente: Macdonald`.",
                                parse_mode="Markdown",
                            )
                            continue
                    # Se ainda faltar ID VENDA, não salva: pede o ID e recoloca a prévia pendente.
                    if "ID VENDA" in parse_result.missing_fields:
                        pending_preview[chat_id] = pending
                        send_message(
                            chat_id,
                            "Faltou o *ID VENDA* para salvar.\nEnvie por exemplo: `id venda 002`.",
                            parse_mode="Markdown",
                        )
                        continue
                    # Se não informou data de entrega e for um fluxo de venda (tem valor/parcelas),
                    # e não houver pendência de pagamento, perguntar a data.
                    needs_delivery_date = _should_ask_delivery_date(parse_result, cmd)
                    if needs_delivery_date:
                        pending_delivery[chat_id] = pending
                        send_message(
                            chat_id,
                            "Qual a *Data de Entrega* deste serviço? (ex.: 20/03, 20/03/2026 ou *hoje*)\n"
                            "Se já foi entregue, pode responder por exemplo: *foi finalizado hoje*.",
                            parse_mode="Markdown",
                        )
                        continue

                    with _spreadsheet_saving_timer(chat_id, getattr(parse_result, "intent", "")):
                        reply = apply_parse_result(
                            parse_result,
                            origin=origin,
                            original_text=original_text,
                            chat_id=str(chat_id),
                        )
                    remembered = sale_id_from_parse_result(parse_result)
                    if remembered:
                        last_sale_id_by_chat[chat_id] = remembered
                    send_chat_action(chat_id, "upload_photo")
                    workbook_path = os.getenv("WORKBOOK_PATH", "").strip()
                    if not workbook_path:
                        from src.workbook_paths import default_workbook_path
                        workbook_path = default_workbook_path([Path.cwd(), Path.cwd().parent])
                    img_path = None
                    if workbook_path:
                        img_path = _maybe_build_sales_snippet_image(
                            workbook_path,
                            sale_id=getattr(parse_result.status_update_command, "sale_id", None)
                            if getattr(parse_result, "intent", "") == "status_update"
                            else getattr(parse_result.command, "sale_id", None),
                        )
                    try:
                        send_reply(
                            chat_id,
                            reply,
                            parse_mode="Markdown",
                            reply_markup=MAIN_MENU_KEYBOARD,
                            image_path=img_path,
                        )
                    finally:
                        if img_path:
                            try:
                                Path(img_path).unlink(missing_ok=True)
                            except Exception:
                                pass
                    print(f"[Telegram] Planilha atualizada para {chat_id}")
                    continue

                # Cancelar prévia
                cancel_words = ("não", "nao", "cancelar", "cancela")
                if text.strip().lower() in cancel_words and chat_id in pending_preview:
                    pending_preview.pop(chat_id, None)
                    send_message(chat_id, "❌ Cancelado. Pode enviar os dados de novo quando quiser.", reply_markup=MAIN_MENU_KEYBOARD)
                    continue

                # Relatório PDF
                if _is_pdf_report_request(text):
                    display_name = (msg.get("from") or {}).get("first_name") or ""
                    _send_planilha_pdf(chat_id, display_name)
                    continue

                # Menu rápido (Prévia/Resumo): feedback imediato + anti clique duplo
                if _handle_quick_dashboard_command(
                    chat_id,
                    text,
                    pending_preview=pending_preview,
                    quick_menu_processing=quick_menu_processing,
                    quick_menu_cooldown=quick_menu_cooldown,
                ):
                    continue

                # Correção pontual da prévia (Cliente/Produto/Valor) antes de confirmar
                if chat_id in pending_preview and text.strip() and not _is_quick_dashboard_command(text):
                    if should_replace_pending_preview(text):
                        pending_preview.pop(chat_id, None)
                    else:
                        lower = text.strip().lower()
                        pending = pending_preview.get(chat_id)
                        if pending:
                            batch_results = pending.get("batch_parse_results")
                            if batch_results:
                                updated, batch_results = apply_batch_preview_corrections(
                                    text,
                                    batch_results,
                                    reference_date=today_local(),
                                )
                                if updated:
                                    _resend_multi_sales_batch_preview(
                                        chat_id,
                                        batch_results,
                                        pending,
                                        pending_preview,
                                    )
                                    continue
                                send_message(
                                    chat_id,
                                    "Não entendi o ajuste no lote.\n\n"
                                    "Diga por exemplo:\n"
                                    "• *Venda 1: Entrada: 3000*\n"
                                    "• *na venda um deu 3000 de entrada*\n"
                                    "• *Venda 2: Valor total: 1500*",
                                    parse_mode="Markdown",
                                )
                                continue
                            repl = parse_message(text, reference_date=today_local())
                            if (
                                repl.intent == "sale_delete"
                                and repl.delete_command
                                and repl.delete_command.sale_id
                                and not repl.missing_fields
                            ):
                                pending_preview[chat_id] = {
                                    "parse_result": repl,
                                    "original_text": text,
                                    "origin": pending.get("origin", "telegram"),
                                }
                                send_message(chat_id, build_preview(repl), parse_mode="Markdown")
                                continue
                        if pending and pending.get("parse_result"):
                            parse_result = pending["parse_result"]
                            cmd = parse_result.command
                            updated = apply_multi_field_corrections(
                                text, cmd, parse_result, reference_date=today_local()
                            )
                            if not updated:
                                updated = apply_preview_corrections(text, cmd)
                            if not updated and "ID VENDA" in parse_result.missing_fields:
                                supplemental_id = extract_supplemental_sale_id(text)
                                if supplemental_id:
                                    cmd.sale_id = supplemental_id
                                    dc = getattr(parse_result, "delete_command", None)
                                    if dc is not None:
                                        dc.sale_id = supplemental_id
                                    parse_result.missing_fields = [
                                        f for f in parse_result.missing_fields if f != "ID VENDA"
                                    ]
                                    updated = True
                            if updated:
                                if (cmd.customer or "").strip() and "Cliente" in parse_result.missing_fields:
                                    parse_result.missing_fields = [
                                        f for f in parse_result.missing_fields if f != "Cliente"
                                    ]
                                # Se o usuário forneceu o ID VENDA, remover da lista de missing.
                                if getattr(cmd, "sale_id", None) and "ID VENDA" in parse_result.missing_fields:
                                    parse_result.missing_fields = [
                                        f for f in parse_result.missing_fields if f != "ID VENDA"
                                    ]
                                dc = getattr(parse_result, "delete_command", None)
                                if dc is not None and getattr(cmd, "sale_id", None):
                                    dc.sale_id = str(cmd.sale_id).strip()
                                _normalize_material_only_sale(parse_result)
                                if not (cmd.customer or "").strip() and getattr(cmd, "sale_id", None):
                                    loaded_customer = _load_customer_from_workbook(str(cmd.sale_id))
                                    if loaded_customer:
                                        cmd.customer = loaded_customer
                                pending["parse_result"] = parse_result
                                _remember_sale_id_from_parse(parse_result, last_sale_id_by_chat, chat_id)
                                preview_text = build_preview(parse_result)
                                send_message(chat_id, preview_text, parse_mode="Markdown")
                                continue
                        if pending and pending.get("parse_result"):
                            send_message(
                                chat_id,
                                "Não entendi o ajuste na prévia.\n\n"
                                "Exemplos:\n"
                                "• `Custo: 1780`\n"
                                "• `Cliente: Ana`\n"
                                "• `id venda 004`\n\n"
                                "Para *confirmar* responda *SIM*. Para *cancelar* responda *NÃO*.",
                                parse_mode="Markdown",
                            )
                            continue

                # Resposta de data de entrega pendente
                if chat_id in pending_delivery:
                    try:
                        pending = pending_delivery.get(chat_id)
                        if not pending:
                            continue
                        raw = text.strip()
                        parsed = _parse_pt_date(raw, today_local(), full_text=raw)
                        if parsed is None:
                            from datetime import datetime
                            for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d/%m"):
                                try:
                                    dt = datetime.strptime(raw, fmt)
                                    if fmt == "%d/%m":
                                        dt = dt.replace(year=today_local().year)
                                    parsed = dt.date()
                                    break
                                except Exception:
                                    continue
                        if parsed is None:
                            if _is_service_delivery_finalized(raw):
                                parsed = today_local()
                            else:
                                raise ValueError("Data inválida.")
                        parse_result = pending["parse_result"]
                        parse_result.command.service_due_date = parsed
                        if _is_service_delivery_finalized(raw):
                            parse_result.command.service_status = "finalizado"
                        with _spreadsheet_saving_timer(chat_id, getattr(parse_result, "intent", "sale")):
                            reply = apply_parse_result(
                                parse_result,
                                origin=pending.get("origin", "telegram"),
                                original_text=pending.get("original_text", ""),
                                chat_id=str(chat_id),
                            )
                            if _is_service_delivery_finalized(raw):
                                sale_id = getattr(parse_result.command, "sale_id", None)
                                if sale_id and _spreadsheet_ready_for_bot():
                                    try:
                                        _open_spreadsheet_for_bot().finalize_service(str(sale_id))
                                    except Exception:
                                        pass
                        send_message(chat_id, reply, parse_mode="Markdown", reply_markup=MAIN_MENU_KEYBOARD)
                        pending_delivery.pop(chat_id, None)
                        continue
                    except Exception:
                        # Se a resposta não for uma data válida, segue sem data de entrega.
                        pending = pending_delivery.pop(chat_id, None)
                        if pending:
                            parse_result = pending["parse_result"]
                            with _spreadsheet_saving_timer(chat_id, getattr(parse_result, "intent", "sale")):
                                reply = apply_parse_result(
                                    parse_result,
                                    origin=pending.get("origin", "telegram"),
                                    original_text=pending.get("original_text", ""),
                                    chat_id=str(chat_id),
                                )
                            send_message(
                                chat_id,
                                "Não entendi a data, então vou salvar sem Data de Entrega.",
                            )
                            send_message(chat_id, reply, parse_mode="Markdown", reply_markup=MAIN_MENU_KEYBOARD)
                        continue

                # Lote: vários ID VENDA — exclusão, entrega ou pagamento na mesma mensagem
                if _try_handle_batch_delete(chat_id, text, pending_preview):
                    continue
                if _try_handle_multi_sales_same_customer(chat_id, text, pending_preview, last_sale_id_by_chat):
                    continue
                if _try_handle_batch_sale_updates(chat_id, text, pending_preview):
                    continue

                # Interpretar mensagem: se tiver dados completos, mostrar prévia; senão, pedir o que falta
                parse_text = enrich_with_last_sale_context(text, last_sale_id_by_chat.get(chat_id))
                if parse_text != text:
                    print(f"[Telegram] Contexto: último ID VENDA {last_sale_id_by_chat.get(chat_id)} aplicado no chat {chat_id}")
                parse_result = parse_message(parse_text, reference_date=today_local())
                _normalize_material_only_sale(parse_result)
                parse_result = _maybe_enhance_parse_with_llm(
                    text,
                    parse_result,
                    last_sale_id=last_sale_id_by_chat.get(chat_id),
                )

                # Atualiza cache quando conseguir extrair cliente.
                parsed_customer = (getattr(parse_result.command, "customer", "") or "").strip()
                explicit_customer = (_extract_customer(text) or _extract_customer(parse_text) or "").strip()
                if explicit_customer:
                    parse_result.command.customer = explicit_customer
                    parse_result.missing_fields = [
                        f for f in parse_result.missing_fields if f != "Cliente"
                    ]
                    parsed_customer = explicit_customer
                if parsed_customer and parsed_customer != "-":
                    last_customer_by_chat[chat_id] = parsed_customer

                # Fallback: se faltou somente "Cliente", tente preencher com o último nome conhecido.
                if "Cliente" in parse_result.missing_fields:
                    cached_customer = None if should_replace_pending_preview(text) else last_customer_by_chat.get(chat_id)
                    if cached_customer:
                        parse_result.command.customer = cached_customer
                        parse_result.missing_fields = [
                            f for f in parse_result.missing_fields if f != "Cliente"
                        ]
                        print(f"[Telegram] Fallback de 'Cliente' usado no chat {chat_id}: {cached_customer}")

                # Fallback: atualização de entrega/status logo após outra ação ("para ele", "entrega hoje").
                if "ID VENDA" in parse_result.missing_fields and parse_result.intent in (
                    "status_update",
                    "delivery_update",
                    "delivery_finalize",
                    "payment_update",
                ):
                    cached_sale_id = last_sale_id_by_chat.get(chat_id)
                    if cached_sale_id:
                        cmd.sale_id = cached_sale_id
                        parse_result.missing_fields = [
                            f for f in parse_result.missing_fields if f != "ID VENDA"
                        ]
                        print(f"[Telegram] Fallback de 'ID VENDA' usado no chat {chat_id}: {cached_sale_id}")

                remembered_sale = sale_id_from_parse_result(parse_result)
                if remembered_sale:
                    last_sale_id_by_chat[chat_id] = remembered_sale
                else:
                    hinted_sale = _extract_target_sale_id_for_updates(text) or extract_supplemental_sale_id(text)
                    if hinted_sale:
                        last_sale_id_by_chat[chat_id] = hinted_sale

                # Inferência: gasto de material sem ID VENDA explícito — tenta achar a venda na planilha.
                cmd = parse_result.command
                if "ID VENDA" in parse_result.missing_fields:
                    try:
                        material_cost = getattr(cmd, "material_cost", None)
                        is_material_expense = bool(material_cost and float(material_cost) > 0 and (cmd.total_value or 0.0) <= 0.01 and not cmd.payments)
                        if is_material_expense:
                            customer_id = (getattr(cmd, "customer", "") or "").strip()
                            if customer_id and _spreadsheet_ready_for_bot():
                                try:
                                    svc = _open_spreadsheet_for_bot()
                                    customer_digits = re.sub(r"\D", "", customer_id)
                                    is_numeric_ref = bool(customer_digits) and customer_id.strip().isdigit()
                                    if is_numeric_ref:
                                        sale_rows = svc.get_pending_sale_by_sale_id(
                                            customer_id, max_rows_scan=500, include_paid=True
                                        )
                                    else:
                                        sale_rows = svc.get_pending_sales_by_customer(
                                            customer_id, max_rows_scan=500, include_paid=True
                                        )
                                    if not sale_rows and customer_digits:
                                        sale_rows = svc.get_pending_sale_by_sale_id(
                                            customer_id, max_rows_scan=500, include_paid=True
                                        )
                                    if sale_rows:
                                        inferred_sale_id = sale_rows[0].get("sale_id")
                                        if inferred_sale_id:
                                            cmd.sale_id = inferred_sale_id
                                            # Se ainda nao houver material_allocations, crie agora para vincular no Excel.
                                            if not getattr(cmd, "material_allocations", None):
                                                from src.models import MaterialAllocation
                                                alloc_date = getattr(cmd, "material_date", None) or today_local()
                                                cmd.material_allocations = [
                                                    MaterialAllocation(
                                                        sale_id=str(inferred_sale_id).strip(),
                                                        amount=float(material_cost),
                                                        material_date=alloc_date,
                                                    )
                                                ]
                                            cmd.description = f"Material da venda {inferred_sale_id}"
                                            parse_result.missing_fields = [
                                                f for f in parse_result.missing_fields if f != "ID VENDA"
                                            ]
                                            print(f"[Telegram] Inferido ID VENDA {inferred_sale_id} (chat {chat_id}).")
                                except Exception:
                                    pass
                    except Exception as e:
                        print(f"[Telegram] Falha na inferencia de ID VENDA: {e}")

                if parse_result.missing_fields:
                    missing_set = set(parse_result.missing_fields)

                    # Mostra prévia mesmo sem Cliente, para o usuário validar os demais campos.
                    if missing_set == {"Cliente"} and parse_result.intent in (
                        "sale",
                        "mixed_update",
                        "refund",
                        "status_update",
                        "delivery_update",
                        "delivery_finalize",
                        "material_update",
                        "sale_value_update",
                        "sale_delete",
                    ):
                        preview_text = build_preview(parse_result)
                        send_message(
                            chat_id,
                            preview_text
                            + "\n\nFaltou apenas o *Cliente* (nome).\n"
                            + "Envie por exemplo: `Cliente: Macdonald`.",
                            parse_mode="Markdown",
                        )
                        pending_preview[chat_id] = {
                            "parse_result": parse_result,
                            "original_text": text,
                            "origin": "audio" if from_voice else "telegram",
                        }
                        continue

                    if missing_set == {"ID VENDA"} and parse_result.intent in (
                        "sale",
                        "mixed_update",
                        "refund",
                        "status_update",
                        "delivery_update",
                        "delivery_finalize",
                        "material_update",
                        "sale_value_update",
                        "sale_delete",
                    ):
                        preview_text = build_preview(parse_result)
                        send_message(
                            chat_id,
                            preview_text
                            + "\n\nFaltou apenas o *ID VENDA* para vincular na planilha.\n"
                            + "Envie por exemplo: `id venda 002`.",
                            parse_mode="Markdown",
                        )
                        pending_preview[chat_id] = {
                            "parse_result": parse_result,
                            "original_text": text,
                            "origin": "audio" if from_voice else "telegram",
                        }
                        continue

                    if chat_id in pending_preview:
                        send_message(
                            chat_id,
                            "Para *confirmar* a prévia responda SIM. Para *cancelar* responda NÃO. "
                            "Ou envie os dados corrigidos (ex.: Cliente é X, produto Y, valor Z).",
                            parse_mode="Markdown",
                        )
                    else:
                        send_message(
                            chat_id,
                            build_missing_fields_message(
                                parse_result.missing_fields,
                                parse_result.intent,
                            ),
                            parse_mode="Markdown",
                        )
                    continue

                # Dados completos: mostrar prévia e guardar até o usuário confirmar
                if parse_result.intent in (
                    "sale",
                    "mixed_update",
                    "refund",
                    "status_update",
                    "delivery_update",
                    "delivery_finalize",
                    "material_update",
                    "sale_delete",
                ):
                    preview_text = build_preview(parse_result)
                    send_message(chat_id, preview_text, parse_mode="Markdown")
                    pending_preview[chat_id] = {
                        "parse_result": parse_result,
                        "original_text": text,
                        "origin": "audio" if from_voice else "telegram",
                    }
                    remembered = sale_id_from_parse_result(parse_result)
                    if remembered:
                        last_sale_id_by_chat[chat_id] = remembered
                    continue

                # Outros casos (ex.: só texto que não é venda/estorno): processar direto
                reply = process_command(text, origin="telegram")
                send_message(chat_id, reply, parse_mode="Markdown", reply_markup=MAIN_MENU_KEYBOARD)
            except Exception as e:
                from src.bot_processor import _spreadsheet_error_message

                reply = f"⚠️ Não consegui processar essa mensagem:\n{_spreadsheet_error_message(e)}"
                send_message(chat_id, reply, reply_markup=MAIN_MENU_KEYBOARD)
                print(f"[Telegram] Erro ao processar: {e}")

        if not updates:
            time.sleep(0.5)

        # Lembretes: checagem a cada ~1 min; envio às 6h e a cada 2h no dia da entrega.
        if time.time() - last_reminder_check >= 60:
            last_reminder_check = time.time()
            try:
                _process_scheduled_reminders(reminder_tracker)
                _process_overdue_deliveries(reminder_tracker, pending_overdue_delivery)
            except Exception as e:
                print(f"[Telegram] Erro nos lembretes: {e}")


def _run_local_test(args: argparse.Namespace) -> None:
    """
    Modo de teste para validar parser + preview + escrita na planilha
    sem depender de TELEGRAM_BOT_TOKEN nem do Telegram em si.
    """
    if args.workbook:
        os.environ["WORKBOOK_PATH"] = args.workbook

    text = args.test_message or ""
    if not text.strip():
        raise SystemExit("Informe --test-message com um texto para testar.")

    # Mantemos a mesma lógica do bot.
    parse_result = parse_message(text, reference_date=today_local())

    if parse_result.missing_fields:
        print("Missing fields:", ", ".join(parse_result.missing_fields))
        return

    if parse_result.intent in (
        "sale",
        "mixed_update",
        "refund",
        "status_update",
        "delivery_update",
        "delivery_finalize",
        "material_update",
        "sale_value_update",
        "sale_delete",
    ):
        preview_text = build_preview(parse_result)
        print(preview_text)

    if args.apply:
        reply = apply_parse_result(
            parse_result,
            origin=args.origin,
            original_text=text,
            chat_id=args.chat_id,
        )
        print("\n--- Apply result ---")
        print(reply)


def main() -> None:
    print("[Telegram] Iniciando bot...", flush=True)
    parser = argparse.ArgumentParser(description="Telegram bot (planilha) com modo local de teste.")
    parser.add_argument("--test-message", dest="test_message", default="", help="Texto para simular uma mensagem do Telegram.")
    parser.add_argument("--apply", action="store_true", help="Aplica a escrita na planilha durante o modo de teste.")
    parser.add_argument("--workbook", dest="workbook", default="", help="Sobrescreve WORKBOOK_PATH no modo de teste.")
    parser.add_argument("--origin", dest="origin", default="local-test", help="Origem usada no log durante o modo de teste.")
    parser.add_argument("--chat-id", dest="chat_id", default="local-test", help="Chat ID usado para lembretes (modo teste).")
    args, _unknown = parser.parse_known_args()

    if args.test_message:
        # Windows console pode estar em cp1252 e falhar ao imprimir emojis (prévia usa emoji).
        if hasattr(sys.stdout, "reconfigure"):
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[call-arg]
        return _run_local_test(args)

    try:
        run_polling()
    except KeyboardInterrupt:
        # Encerramento amigável no Ctrl+C (sem traceback assustando o usuário).
        print("\n[Telegram] Bot encerrado por Ctrl+C.", flush=True)
        return


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n[Telegram] Bot encerrado por Ctrl+C.", flush=True)
    except Exception as exc:
        print(f"\n[Telegram] Erro ao iniciar o bot: {exc}", flush=True)
        raise SystemExit(1) from exc
