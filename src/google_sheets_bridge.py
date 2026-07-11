# -*- coding: utf-8 -*-
"""
Ponte Google Sheets ↔ openpyxl.

Leitura: 1 values.batchGet para todas as abas (em vez de 1 get_all_values por aba).
Gravação: batch_update agrupado (em vez de update_acell por célula).
"""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any
import re
import time
import unicodedata

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .google_sheets_api import GoogleSheetsQuotaError, sheets_call

# Mesmo limite usado pelo excel_store.
_MAX_SYNC_ROW = 1048575
_MAX_SYNC_COL = 26  # A..Z

# Abas auxiliares: se falharem no sync, a venda principal ainda deve ser salva.
_OPTIONAL_SHEETS = frozenset({"Log_Agente", "Lembretes"})
# Reusa a planilha em memória entre leituras/gravações do bot (bem mais rápido).
_CACHE_TTL_SEC = 90.0
# NÃO baixar a planilha inteira: a linha-âncora 1048576 faz o batchGet demorar ~1 min.
_MAX_FETCH_ROW = 800
_MAX_FETCH_COL = "Z"
# Só baixa abas usadas pelo robô (igual à ideia do Line: não varrer a planilha inteira).
_CORE_SHEET_HINTS = (
    "total de vendas",
    "compras",
    "gastos fixos",
    "log agente",
    "lembretes",
    "_tmpl_",
)


class GoogleSheetsBridge:
    def __init__(self, sheet_id: str, credentials_path: str | Path) -> None:
        self.sheet_id = sheet_id.strip()
        self.credentials_path = Path(credentials_path)
        if not self.credentials_path.is_file():
            raise FileNotFoundError(f"Credenciais Google nao encontradas: {self.credentials_path}")
        self._client = None
        self._spreadsheet = None
        self._worksheets: dict[str, Any] = {}
        self._baseline: dict[str, dict[str, Any]] = {}
        self._aux_sheets_ready = False
        # Cache em memória: evita baixar a planilha inteira de novo a cada operação.
        self._memory_wb: Workbook | None = None
        self._cache_until = 0.0

    def _authorize(self):
        if self._client is not None:
            return self._client
        import gspread
        from google.oauth2.service_account import Credentials

        scopes = ["https://www.googleapis.com/auth/spreadsheets"]
        creds = Credentials.from_service_account_file(str(self.credentials_path), scopes=scopes)
        self._client = gspread.authorize(creds)
        return self._client

    def _spreadsheet_obj(self):
        if self._spreadsheet is None:
            self._spreadsheet = sheets_call(
                lambda: self._authorize().open_by_key(self.sheet_id),
            )
        return self._spreadsheet

    @staticmethod
    def _normalize_sheet_title(title: str) -> str:
        clean = unicodedata.normalize("NFKD", (title or "").strip().lower())
        clean = "".join(ch for ch in clean if not unicodedata.combining(ch))
        return " ".join(clean.replace("_", " ").split())

    def _find_worksheet(self, spreadsheet, title: str):
        try:
            return spreadsheet.worksheet(title)
        except Exception:
            pass
        wanted = self._normalize_sheet_title(title)
        for ws in spreadsheet.worksheets():
            if ws.title == title:
                return ws
            if self._normalize_sheet_title(ws.title) == wanted:
                return ws
        return None

    def _worksheet(self, title: str, *, create_if_missing: bool = False):
        if title not in self._worksheets:
            spreadsheet = self._spreadsheet_obj()

            def _open():
                ws = self._find_worksheet(spreadsheet, title)
                if ws is not None:
                    return ws
                if not create_if_missing:
                    raise KeyError(
                        f"Aba '{title}' nao encontrada na planilha Google. "
                        "Crie a aba ou permita que a conta de servico a crie."
                    )
                try:
                    return spreadsheet.add_worksheet(title=title, rows=1000, cols=26)
                except Exception:
                    # Corrida / aba criada entre o find e o add.
                    ws = self._find_worksheet(spreadsheet, title)
                    if ws is not None:
                        return ws
                    raise

            self._worksheets[title] = sheets_call(_open, is_write=create_if_missing)
        return self._worksheets[title]

    def ensure_sheets(self, titles: list[str] | tuple[str, ...]) -> None:
        """Garante que as abas existam no Google Sheets (cria se faltar)."""
        pending = [
            t
            for t in titles
            if self._normalize_sheet_title(t)
            not in {self._normalize_sheet_title(cached) for cached in self._worksheets}
        ]
        if not pending and self._aux_sheets_ready:
            return
        for title in titles:
            try:
                self._worksheet(title, create_if_missing=True)
            except Exception as exc:
                print(f"[GoogleSheets] aviso: nao foi possivel garantir aba '{title}': {exc}")
        self._aux_sheets_ready = True

    @classmethod
    def _is_core_sheet(cls, title: str) -> bool:
        norm = cls._normalize_sheet_title(title)
        return any(hint in norm for hint in _CORE_SHEET_HINTS)

    @staticmethod
    def _date_as_sheets_text(value: date | datetime) -> str:
        """
        Mesmo formato do Assistente_Line na nuvem: 'YYYY-MM-DD HH:MM:SS'.
        Sem fórmula (=DATE/=DATA) — o Sheets interpreta o texto como data.
        """
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        now = datetime.now()
        return datetime.combine(value, now.time()).strftime("%Y-%m-%d %H:%M:%S")

    @classmethod
    def _serialize_value(cls, value: Any) -> str | int | float:
        if value is None:
            return ""
        if isinstance(value, bool):
            return "TRUE" if value else "FALSE"
        if isinstance(value, (int, float)):
            return value
        if isinstance(value, datetime):
            return cls._date_as_sheets_text(value)
        if isinstance(value, date):
            return cls._date_as_sheets_text(value)
        text = str(value).strip()
        # USER_ENTERED transforma "001" em número 1 — força texto para IDs com zero à esquerda.
        if text.isdigit() and len(text) >= 2 and text.startswith("0"):
            return f"'{text}"
        if text.startswith("'"):
            text = text[1:].strip()
        # Já está no formato Line / ISO datetime — mantém.
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}([ T]\d{2}:\d{2}(:\d{2})?)?", text):
            if "T" in text:
                text = text.replace("T", " ", 1)
            if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
                return cls._date_as_sheets_text(date.fromisoformat(text))
            return text
        # Datas texto DD/MM/YYYY → mesmo formato ISO do Line.
        m = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", text)
        if m:
            day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if year < 100:
                year += 2000
            try:
                return cls._date_as_sheets_text(date(year, month, day))
            except ValueError:
                pass
        # Fórmulas antigas DATE/DATA → converte para texto ISO (corrige #ERROR!).
        m_formula = re.fullmatch(
            r"=(?:DATE|DATA)\((\d{4})\s*[,;]\s*(\d{1,2})\s*[,;]\s*(\d{1,2})\)",
            text,
            flags=re.IGNORECASE,
        )
        if m_formula:
            try:
                return cls._date_as_sheets_text(
                    date(
                        int(m_formula.group(1)),
                        int(m_formula.group(2)),
                        int(m_formula.group(3)),
                    )
                )
            except ValueError:
                pass
        return str(value)

    @staticmethod
    def _cell_key(row: int, col: int) -> str:
        return f"{get_column_letter(col)}{row}"

    @staticmethod
    def _coerce_cell_value(value: Any) -> Any:
        """Normaliza valores do batchGet (números crus; datas serial ficam numéricas)."""
        if value in (None, ""):
            return None
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            serial = float(value)
            # Mantém número. Colunas de data usam _parse_date_cell (serial → date).
            return serial if serial != int(serial) else int(serial) if abs(serial) < 1e15 else serial
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return None
            return text
        return value

    def _snapshot_workbook(self, wb) -> dict[str, dict[str, Any]]:
        snap: dict[str, dict[str, Any]] = {}
        for name in wb.sheetnames:
            ws = wb[name]
            cells: dict[str, Any] = {}
            for row in ws.iter_rows(
                min_row=1,
                max_row=min(ws.max_row or 1, _MAX_SYNC_ROW),
                max_col=_MAX_SYNC_COL,
            ):
                for cell in row:
                    if cell.value is None or cell.value == "":
                        continue
                    cells[self._cell_key(cell.row, cell.column)] = cell.value
            snap[name] = cells
        return snap

    def _clone_workbook(self, wb) -> Workbook:
        """Cópia leve só com valores (suficiente para leitura/gravação do bot)."""
        clone = Workbook()
        clone.remove(clone.active)
        for name in wb.sheetnames:
            src = wb[name]
            dst = clone.create_sheet(title=name[:31])
            for row in src.iter_rows(
                min_row=1,
                max_row=min(src.max_row or 1, _MAX_SYNC_ROW),
                max_col=_MAX_SYNC_COL,
            ):
                for cell in row:
                    if cell.value not in (None, ""):
                        dst.cell(row=cell.row, column=cell.column, value=cell.value)
        return clone

    def _workbook_from_baseline(self) -> Workbook:
        wb = Workbook()
        wb.remove(wb.active)
        for name, cells in self._baseline.items():
            sheet = wb.create_sheet(title=str(name)[:31])
            for key, value in cells.items():
                if value in (None, ""):
                    continue
                sheet[key] = value
        return wb

    def _build_workbook_from_grids(self, *, data_only: bool = False, read_only: bool = False):
        """Lê abas do robô em 1–2 chamadas à API (batchGet) — sem baixar o resto da planilha."""
        spreadsheet = self._spreadsheet_obj()
        worksheets = sheets_call(spreadsheet.worksheets)
        if not worksheets:
            wb = Workbook()
            self._baseline = self._snapshot_workbook(wb)
            self._worksheets.clear()
            return wb

        # Preferência: só abas usadas pelo bot (mais rápido, no espírito do Line).
        core = [ws for ws in worksheets if self._is_core_sheet(ws.title)]
        selected = core if core else list(worksheets)

        for ws_meta in selected:
            self._worksheets[ws_meta.title] = ws_meta

        ranges = [
            f"'{ws_meta.title}'!A1:{_MAX_FETCH_COL}{_MAX_FETCH_ROW}"
            for ws_meta in selected
        ]

        def _batch_get():
            t0 = time.monotonic()
            payload = spreadsheet.values_batch_get(
                ranges,
                params={
                    # Números crus + datas como serial (evita "7/11/2026" ambíguo do locale US).
                    "valueRenderOption": "UNFORMATTED_VALUE",
                    "dateTimeRenderOption": "SERIAL_NUMBER",
                },
            )
            print(
                f"[GoogleSheets] batchGet {len(ranges)} aba(s) "
                f"A1:{_MAX_FETCH_COL}{_MAX_FETCH_ROW} em {time.monotonic() - t0:.2f}s"
            )
            return payload

        try:
            payload = sheets_call(_batch_get)
        except Exception as exc:
            print(f"[GoogleSheets] aviso: batchGet falhou ({exc}); fallback get_all_values")
            return self._build_workbook_from_grids_fallback(selected)

        value_ranges = payload.get("valueRanges", []) if isinstance(payload, dict) else []
        wb = Workbook()
        wb.remove(wb.active)

        for idx, ws_meta in enumerate(selected):
            sheet = wb.create_sheet(title=ws_meta.title[:31])
            rows = []
            if idx < len(value_ranges):
                rows = value_ranges[idx].get("values") or []
            for r_idx, row in enumerate(rows, start=1):
                for c_idx, value in enumerate(row, start=1):
                    coerced = self._coerce_cell_value(value)
                    if coerced not in (None, ""):
                        sheet.cell(row=r_idx, column=c_idx, value=coerced)

        self._baseline = self._snapshot_workbook(wb)
        return wb

    def _build_workbook_from_grids_fallback(self, worksheets) -> Workbook:
        wb = Workbook()
        wb.remove(wb.active)
        for ws_meta in worksheets:
            rows = sheets_call(ws_meta.get_all_values)
            sheet = wb.create_sheet(title=ws_meta.title[:31])
            for r_idx, row in enumerate(rows, start=1):
                for c_idx, value in enumerate(row, start=1):
                    if value not in (None, ""):
                        sheet.cell(row=r_idx, column=c_idx, value=value)
        self._baseline = self._snapshot_workbook(wb)
        return wb

    def invalidate_cache(self) -> None:
        self._memory_wb = None
        self._cache_until = 0.0

    def open_workbook(self, *, data_only: bool = False, read_only: bool = False):
        """Carrega a planilha. Reusa cache em memória na janela TTL (evita download a cada save)."""
        if self._memory_wb is not None and time.monotonic() < self._cache_until:
            return self._clone_workbook(self._memory_wb)
        wb = self._build_workbook_from_grids(data_only=data_only, read_only=read_only)
        self._memory_wb = self._clone_workbook(wb)
        self._cache_until = time.monotonic() + _CACHE_TTL_SEC
        return wb

    def save_workbook(self, wb, *, allow_clears: bool = False) -> None:
        """Envia apenas células alteradas em 1 batchUpdate (estilo Line: só o que mudou).

        Por padrão NÃO apaga valores existentes na nuvem quando a célula local está vazia
        (evita wipe acidental). Use allow_clears=True em exclusões explícitas.
        """
        if not self._baseline:
            self._baseline = self._snapshot_workbook(wb)
            self._memory_wb = self._clone_workbook(wb)
            self._cache_until = time.monotonic() + _CACHE_TTL_SEC
            return

        optional_failures: list[str] = []
        # Um único values.batchUpdate para todas as abas (menos round-trips à API).
        remote_data: list[dict[str, Any]] = []

        for name in wb.sheetnames:
            try:
                ws = wb[name]
                # Garante handle da aba (cria se for auxiliar nova).
                self._worksheet(name, create_if_missing=True)
                old_cells = self._baseline.get(name, {})

                baseline_max_row = 1
                for key in old_cells:
                    digits = "".join(ch for ch in key if ch.isdigit())
                    if digits:
                        baseline_max_row = max(baseline_max_row, int(digits))
                scan_max = min(
                    max(ws.max_row or 1, baseline_max_row + 5, 50),
                    2000,
                    _MAX_SYNC_ROW,
                )

                for row in ws.iter_rows(
                    min_row=1,
                    max_row=scan_max,
                    max_col=_MAX_SYNC_COL,
                ):
                    for cell in row:
                        key = self._cell_key(cell.row, cell.column)
                        new_val = cell.value
                        old_val = old_cells.get(key)
                        new_empty = new_val is None or new_val == ""
                        old_empty = old_val is None or old_val == ""
                        if new_empty and old_empty:
                            continue
                        if new_empty and not old_empty and not allow_clears:
                            continue
                        if not new_empty and not old_empty and new_val == old_val:
                            continue
                        serialized = self._serialize_value(new_val)
                        # Evita reescrever data que na nuvem já está como serial/texto equivalente.
                        if not new_empty and not old_empty:
                            if self._values_equivalent(old_val, new_val, serialized):
                                continue
                        remote_data.append(
                            {
                                "range": f"'{name}'!{key}",
                                "values": [[serialized]],
                            }
                        )
            except Exception as exc:
                if name in _OPTIONAL_SHEETS or self._normalize_sheet_title(name) in {
                    self._normalize_sheet_title(t) for t in _OPTIONAL_SHEETS
                }:
                    optional_failures.append(f"{name}: {exc}")
                    print(f"[GoogleSheets] aviso: falha ao sincronizar aba opcional '{name}': {exc}")
                    continue
                raise

        if remote_data:
            spreadsheet = self._spreadsheet_obj()
            chunk_size = 400
            for i in range(0, len(remote_data), chunk_size):
                chunk = remote_data[i : i + chunk_size]
                t0 = time.monotonic()

                def _write(c=chunk):
                    return spreadsheet.values_batch_update(
                        {
                            "valueInputOption": "USER_ENTERED",
                            "data": c,
                        }
                    )

                sheets_call(_write, is_write=True)
                print(
                    f"[GoogleSheets] batchUpdate {len(chunk)} célula(s) "
                    f"em {time.monotonic() - t0:.2f}s"
                )

        if optional_failures:
            print(
                "[GoogleSheets] venda/dados principais salvos; "
                f"abas auxiliares com aviso: {'; '.join(optional_failures)}"
            )

        self._baseline = self._snapshot_workbook(wb)
        self._memory_wb = self._clone_workbook(wb)
        self._cache_until = time.monotonic() + _CACHE_TTL_SEC

    @classmethod
    def _values_equivalent(cls, old_val: Any, new_val: Any, serialized_new: Any) -> bool:
        """True se o valor antigo (nuvem/baseline) já representa o mesmo dado que vamos gravar."""
        if old_val == new_val:
            return True
        if old_val == serialized_new:
            return True
        # Serial Sheets vs date/datetime
        if isinstance(new_val, (date, datetime)) and isinstance(old_val, (int, float)):
            try:
                d = new_val.date() if isinstance(new_val, datetime) else new_val
                serial = (d - date(1899, 12, 30)).days
                if abs(float(old_val) - serial) < 0.0001:
                    return True
            except Exception:
                pass
        # Texto ISO já na nuvem vs date local
        if isinstance(new_val, (date, datetime)) and isinstance(old_val, str):
            old_txt = old_val.strip().replace("T", " ", 1)
            new_txt = str(serialized_new).strip()
            if old_txt == new_txt:
                return True
            if old_txt[:10] == new_txt[:10]:
                return True
        return False

    def format_font_colors(
        self,
        sheet_title: str,
        cells: list[tuple[str, tuple[float, float, float]]],
    ) -> None:
        """Aplica cor de fonte via Sheets API (valores openpyxl de fonte não sobem sozinhos)."""
        if not cells:
            return
        worksheet = self._worksheet(sheet_title, create_if_missing=False)
        # Agrupa por cor para menos chamadas.
        by_color: dict[tuple[float, float, float], list[str]] = {}
        for a1, rgb in cells:
            by_color.setdefault(rgb, []).append(a1)
        for (r, g, b), ranges in by_color.items():
            style = {
                "textFormat": {
                    "foregroundColor": {"red": r, "green": g, "blue": b},
                    "bold": True,
                }
            }
            # gspread format aceita range A1; para várias células, formata uma a uma (poucas).
            for a1 in ranges:
                sheets_call(
                    lambda a1=a1, style=style, w=worksheet: w.format(a1, style),
                    is_write=True,
                )

    def copy_row_format(
        self,
        sheet_title: str,
        source_row: int,
        target_row: int,
        *,
        start_col: str = "A",
        end_col: str = "I",
    ) -> None:
        """Copia só a formatação da linha de cima para a nova (sem sobrescrever valores)."""
        if source_row == target_row or source_row < 1 or target_row < 1:
            return
        from openpyxl.utils import column_index_from_string as _col_idx

        worksheet = self._worksheet(sheet_title, create_if_missing=False)
        sheet_id = worksheet.id
        start_i = max(0, _col_idx(start_col) - 1)
        end_i = max(start_i + 1, _col_idx(end_col))
        body = {
            "requests": [
                {
                    "copyPaste": {
                        "source": {
                            "sheetId": sheet_id,
                            "startRowIndex": source_row - 1,
                            "endRowIndex": source_row,
                            "startColumnIndex": start_i,
                            "endColumnIndex": end_i,
                        },
                        "destination": {
                            "sheetId": sheet_id,
                            "startRowIndex": target_row - 1,
                            "endRowIndex": target_row,
                            "startColumnIndex": start_i,
                            "endColumnIndex": end_i,
                        },
                        "pasteType": "PASTE_FORMAT",
                        "pasteOrientation": "NORMAL",
                    }
                }
            ]
        }
        spreadsheet = self._spreadsheet_obj()
        sheets_call(lambda: spreadsheet.batch_update(body), is_write=True)


__all__ = ["GoogleSheetsBridge", "GoogleSheetsQuotaError"]
