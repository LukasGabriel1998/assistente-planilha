from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import os
import re
import unicodedata
import zipfile

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill

from .models import DeleteSaleCommand, FinancialCommand, RefundCommand, StatusUpdateCommand

SHEET_SALES = "TOTAL DE VENDAS DE 2026"
# Nome normalizado; no Excel a aba pode ser "Compras Matéria-Prima" (com acento).
SHEET_MATERIAL = "Compras Materia-Prima"
SHEET_FIXED = "Gastos Fixos"
SHEET_LOG = "Log_Agente"
SHEET_TEMPLATE_SALES = "_tmpl_vendas"
SHEET_TEMPLATE_MATERIAL = "_tmpl_material"
SHEET_TEMPLATE_FIXED = "_tmpl_fixos"
SHEET_REMINDERS = "Lembretes"
SUPPORTED_WORKBOOK_EXTS = (".xlsx", ".xlsm", ".xltx", ".xltm")
SALES_HEADER_ROW = 2
TEMPLATE_ROW = 3
ANCHOR_TEMPLATE_ROW = 1048576
DATA_START_ROW = 3
# Coluna J: caixas de totais (rótulos + fórmulas). O robô nunca deve escrever nem copiar estilo daqui.
SALES_SUMMARY_COL = "J"
# Última coluna da tabela de vendas (cabeçalhos na linha 2). Evita tratar J2 (fórmula) como cabeçalho.
SALES_TABLE_LAST_COL = "I"
SALES_REQUIRED_HEADERS = {
    # A aba TOTAL DE VENDAS DE 2026 agora usa explicitamente:
    # Coluna A: Data de venda
    # Coluna B: Data de Entrega
    # Coluna C: Cliente (antes "ID Cliente")
    "data de venda": "Data de venda",
    "data de entrega": "Data de Entrega",
    "cliente": "Cliente",
    "id produto": "ID produto",
    "total de vendas (pago)": "Total de vendas (pago)",
    "valor (pendente)": "Valor (pendente)",
    "id venda": "ID VENDA",
    "status de valor": "Status de valor",
}

FILL_PENDING = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")  # amarelo claro
FILL_DONE = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # verde claro
MATERIAL_HEADER_ROW = 2
MATERIAL_REQUIRED_HEADERS = {
    "data": "Data",
    "descricao": "Descricao",
    "valor": "Valor",
}


@dataclass
class WriteAction:
    sheet: str
    row: int
    amount: float
    label: str
    sale_id: str | None = None


class SpreadsheetService:
    def __init__(self, workbook_path: str | Path) -> None:
        self.workbook_path = Path(workbook_path)
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"Planilha nao encontrada: {self.workbook_path}")
        if self.workbook_path.is_dir():
            raise ValueError(
                "O caminho informado e uma pasta. Informe o arquivo da planilha "
                "(.xlsx, .xlsm, .xltx ou .xltm)."
            )
        if self.workbook_path.suffix.lower() not in SUPPORTED_WORKBOOK_EXTS:
            raise ValueError("Formato invalido de planilha. Use .xlsx, .xlsm, .xltx ou .xltm.")
        if self.workbook_path.name.startswith("~$"):
            raise ValueError(
                "Arquivo temporario do Excel detectado (~$...). "
                "Selecione o arquivo principal da planilha."
            )

    def _open_workbook(self, *, data_only: bool = False, read_only: bool = False):
        try:
            return load_workbook(
                self.workbook_path,
                data_only=data_only,
                read_only=read_only,
            )
        except PermissionError as exc:
            raise PermissionError(
                "Nao foi possivel abrir a planilha. Feche o arquivo no Excel e tente novamente."
            ) from exc
        except zipfile.BadZipFile as exc:
            from .workbook_paths import is_git_lfs_pointer

            if is_git_lfs_pointer(self.workbook_path):
                raise ValueError(
                    "A planilha no disco ainda e um ponteiro do Git LFS (arquivo real nao baixado). "
                    "Rode: sudo apt install git-lfs && git lfs install && git lfs pull "
                    "ou copie manualmente o .xlsx e ajuste WORKBOOK_PATH no .env."
                ) from exc
            raise ValueError(
                "Arquivo de planilha invalido ou corrompido. "
                "Abra e salve novamente no Excel como .xlsx/.xlsm."
            ) from exc
        except Exception as exc:
            msg = str(exc).lower()
            if "openpyxl does not support file format" in msg:
                raise ValueError(
                    "Formato de planilha nao suportado. Use .xlsx, .xlsm, .xltx ou .xltm."
                ) from exc
            raise

    def _save_workbook(self, wb) -> None:
        try:
            wb.save(self.workbook_path)
        except PermissionError as exc:
            raise PermissionError(
                "Sem permissao para salvar a planilha. Feche o arquivo no Excel e tente novamente."
            ) from exc
        except OSError as exc:
            msg = str(exc).lower()
            if ("permission denied" in msg) or ("being used by another process" in msg):
                raise PermissionError(
                    "Nao foi possivel salvar porque a planilha esta em uso. "
                    "Feche no Excel e tente novamente."
                ) from exc
            raise

    @staticmethod
    def _normalize_name(name: str) -> str:
        clean = name.strip().lower()
        clean = unicodedata.normalize("NFKD", clean)
        clean = "".join(ch for ch in clean if not unicodedata.combining(ch))
        return " ".join(clean.split())

    @classmethod
    def _resolve_sheet_name(cls, wb, expected_name: str) -> str:
        if expected_name in wb.sheetnames:
            return expected_name
        expected_norm = cls._normalize_name(expected_name).rsplit(" ", 1)[0]
        for real_name in wb.sheetnames:
            if cls._normalize_name(real_name).startswith(expected_norm):
                return real_name
        raise ValueError(f"Aba obrigatoria nao encontrada: {expected_name}")

    # Última linha da planilha guarda só o padrão (estilo, sem valores). O robô nunca escreve dados lá.
    MAX_DATA_ROW = ANCHOR_TEMPLATE_ROW - 1
    TELEGRAM_ROW_SCAN_LIMIT = 500

    @classmethod
    def _scan_row_cap(cls, ws, start_row: int = DATA_START_ROW, max_rows_scan: int | None = None) -> int:
        """Limite superior de varredura — evita percorrer milhões de linhas por causa da linha âncora."""
        limit = max_rows_scan if max_rows_scan is not None else cls.TELEGRAM_ROW_SCAN_LIMIT
        return min(ws.max_row, cls.MAX_DATA_ROW, start_row + max(1, limit) - 1)

    @classmethod
    def _effective_data_end_row(
        cls,
        ws,
        key_col: str,
        start_row: int = DATA_START_ROW,
        max_rows_scan: int | None = None,
    ) -> int:
        """Última linha com dado real, sem varrer a planilha inteira."""
        cap = cls._scan_row_cap(ws, start_row, max_rows_scan)
        if getattr(ws, "read_only", False):
            from openpyxl.utils import column_index_from_string

            col_idx = column_index_from_string(key_col)
            last = start_row
            for row in ws.iter_rows(min_row=start_row, max_row=cap, min_col=col_idx, max_col=col_idx):
                if row[0].value not in (None, ""):
                    last = row[0].row
            return last
        for row in range(cap, start_row - 1, -1):
            if ws[f"{key_col}{row}"].value not in (None, ""):
                return row
        return start_row

    def _sum_column_range(
        self,
        ws,
        col: str | None,
        start_row: int,
        end_row: int,
    ) -> float:
        if not col or end_row < start_row:
            return 0.0
        if getattr(ws, "read_only", False):
            from openpyxl.utils import column_index_from_string

            col_idx = column_index_from_string(col)
            total = 0.0
            for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=col_idx, max_col=col_idx):
                total += self._to_float(row[0].value)
            return total
        total = 0.0
        for row in range(start_row, end_row + 1):
            total += self._to_float(ws[f"{col}{row}"].value)
        return total

    def _count_filled_rows(
        self,
        ws,
        key_col: str | None,
        start_row: int,
        end_row: int,
    ) -> int:
        if not key_col or end_row < start_row:
            return 0
        if getattr(ws, "read_only", False):
            from openpyxl.utils import column_index_from_string

            col_idx = column_index_from_string(key_col)
            count = 0
            for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=col_idx, max_col=col_idx):
                if row[0].value not in (None, ""):
                    count += 1
            return count
        count = 0
        for row in range(start_row, end_row + 1):
            if ws[f"{key_col}{row}"].value not in (None, ""):
                count += 1
        return count

    @staticmethod
    def _next_row(ws, key_col: str = "A", start_row: int = DATA_START_ROW) -> int:
        row = start_row
        max_row = min(ws.max_row, SpreadsheetService.MAX_DATA_ROW)
        while row <= max_row and ws[f"{key_col}{row}"].value not in (None, ""):
            row += 1
        # Nunca devolver a linha âncora (última linha); ela guarda só o padrão, sem dados
        return min(row, SpreadsheetService.MAX_DATA_ROW)

    @staticmethod
    def _next_row_for_empty_cols(
        ws,
        empty_cols: tuple[str, ...],
        *,
        start_row: int = DATA_START_ROW,
        skip_rows: set[int] | frozenset[int] | None = None,
    ) -> int:
        row = start_row
        max_row = min(ws.max_row, SpreadsheetService.MAX_DATA_ROW)
        blocked = skip_rows or frozenset()
        while row <= max_row:
            if row in blocked:
                row += 1
                continue
            if all(ws[f"{col}{row}"].value in (None, "") for col in empty_cols):
                return row
            row += 1
        return row

    @staticmethod
    def _last_filled_row(ws, key_col: str = "A", start_row: int = DATA_START_ROW) -> int:
        for row in range(ws.max_row, start_row - 1, -1):
            if ws[f"{key_col}{row}"].value not in (None, ""):
                return row
        return start_row

    @staticmethod
    def _apply_currency_format(ws, col: str, row: int) -> None:
        ws[f"{col}{row}"].number_format = "R$ #,##0.00"

    @staticmethod
    def _apply_date_format(ws, col: str, row: int) -> None:
        ws[f"{col}{row}"].number_format = "DD/MM/YYYY"

    @staticmethod
    def _preserve_user_layout() -> bool:
        """
        Quando ativo (padrão), evita normalizar/reformatar a planilha inteira e não mexe na coluna J.
        Não impede copiar o padrão da linha 1048576 nem as cores de entrega/status.
        """
        val = os.getenv("SPREADSHEET_PRESERVE_LAYOUT", "1").strip().lower()
        return val not in ("0", "false", "no", "off")

    @classmethod
    def _apply_anchor_row_style(
        cls,
        wb,
        ws,
        row: int,
        logical_sheet_name: str,
        cols: tuple[str, ...],
    ) -> tuple:
        """Replica o padrão visual da linha 1048576 (ou template oculto) na linha alvo."""
        anchor = cls._anchor_template_row_for(logical_sheet_name)
        if anchor and (
            cls._row_has_padrao(ws, anchor, cols)
            or cls._row_has_values(ws, anchor, cols)
            or cls._row_has_any_border(ws, anchor, cols)
        ):
            cls._copy_row_style_between(ws, anchor, ws, row, cols, apply_row_dimensions=False)
            return ws, anchor
        template_ws, template_row = cls._template_source(wb, ws, logical_sheet_name, cols)
        if not (template_ws is ws and template_row == row):
            cls._copy_row_style_between(
                template_ws, template_row, ws, row, cols, apply_row_dimensions=False
            )
        return template_ws, template_row

    @classmethod
    def _is_service_delivered(cls, ws, sales_cols: dict[str, str], row: int) -> bool:
        delivery_col = sales_cols.get("data de entrega")
        if delivery_col:
            cell = ws[f"{delivery_col}{row}"]
            if cls._cell_fill_matches(cell, "C8E6C9"):
                return True
        legacy_status_col = sales_cols.get("status")
        if legacy_status_col:
            status_txt = cls._normalize_name(str(ws[f"{legacy_status_col}{row}"].value or ""))
            if status_txt == "finalizado":
                return True
        return False

    def is_sale_delivered(self, sale_id: str) -> bool:
        """True quando a venda já está marcada como entregue (Data de Entrega verde)."""
        sale_id = str(sale_id or "").strip()
        if not sale_id:
            return False
        wb = self._open_workbook(data_only=True)
        try:
            sales_name = self._resolve_sheet_name(wb, SHEET_SALES)
            ws_sales = wb[sales_name]
            sales_cols = self._sales_columns(ws_sales)
            try:
                row = self._find_sale_row(ws_sales, sales_cols["id venda"], sale_id)
            except ValueError:
                return False
            return self._is_service_delivered(ws_sales, sales_cols, row)
        finally:
            wb.close()

    @classmethod
    def _apply_sales_row_visual_status(
        cls,
        ws,
        sales_cols: dict[str, str],
        row: int,
        *,
        pending_amount: float = 0.0,
        service_delivered: bool | None = None,
    ) -> None:
        """
        Aplica cores conforme o padrão da planilha:
        - Status de valor (H): amarelo se pendente, verde se pago.
        - Data de Entrega (B): amarelo até entregar, verde quando entregue.
        """
        if service_delivered is None:
            service_delivered = cls._is_service_delivered(ws, sales_cols, row)

        status_col = sales_cols.get("status de valor")
        if status_col:
            ws[f"{status_col}{row}"].fill = (
                FILL_PENDING if abs(float(pending_amount or 0.0)) > 0.01 else FILL_DONE
            )

        delivery_col = sales_cols.get("data de entrega")
        if delivery_col:
            ws[f"{delivery_col}{row}"].fill = FILL_DONE if service_delivered else FILL_PENDING

    @classmethod
    def _maybe_set_cell_fill(cls, cell, fill) -> None:
        cell.fill = fill

    @staticmethod
    def _ensure_number_format_if_general(ws, col: str, row: int, fmt: str) -> None:
        """Só aplica formato se a célula estiver em General; não sobrescreve formatação da planilha."""
        if SpreadsheetService._preserve_user_layout():
            return
        cell = ws[f"{col}{row}"]
        current = str(getattr(cell, "number_format", None) or "General").strip()
        if not current or current == "General":
            cell.number_format = fmt

    @staticmethod
    def _excel_date_value(value) -> str:
        if hasattr(value, "strftime"):
            return value.strftime("%d/%m/%Y")
        return str(value)

    @classmethod
    def _header_map(
        cls,
        ws,
        header_row: int = SALES_HEADER_ROW,
        *,
        max_col: str | None = None,
    ) -> dict[str, str]:
        headers: dict[str, str] = {}
        last_col_idx = ws.max_column
        if max_col:
            last_col_idx = min(last_col_idx, column_index_from_string(max_col))
        for col_idx in range(1, last_col_idx + 1):
            value = ws.cell(row=header_row, column=col_idx).value
            if value in (None, ""):
                continue
            if not isinstance(value, str):
                continue
            headers[cls._normalize_name(value)] = get_column_letter(col_idx)
        return headers

    @staticmethod
    def _cell_has_content(value) -> bool:
        return value not in (None, "")

    @classmethod
    def _sales_summary_rows(cls, ws, *, scan_limit: int = 40) -> set[int]:
        """
        Linhas com conteúdo na coluna J (caixas de totais).
        Usado só para não sobrescrever estilo de J na normalização — não bloqueia novas vendas em A–I.
        """
        rows: set[int] = set()
        for row in range(1, min(ws.max_row, scan_limit) + 1):
            if cls._cell_has_content(ws[f"{SALES_SUMMARY_COL}{row}"].value):
                rows.add(row)
        return rows

    @classmethod
    def _is_sales_summary_row(cls, ws, row: int) -> bool:
        return cls._cell_has_content(ws[f"{SALES_SUMMARY_COL}{row}"].value)

    @classmethod
    def _sale_data_columns(cls, sales_cols: dict[str, str]) -> tuple[str, ...]:
        """Colunas da tabela de vendas (A–I). A coluna J (totais) fica de fora."""
        cols = [
            sales_cols["data de venda"],
            sales_cols["data de entrega"],
            sales_cols["cliente"],
            sales_cols["id produto"],
            sales_cols["total de vendas (pago)"],
            sales_cols["valor (pendente)"],
            sales_cols["id venda"],
            sales_cols["status de valor"],
        ]
        status_col = sales_cols.get("status")
        if status_col and status_col not in cols:
            cols.append(status_col)
        return tuple(cols)

    @classmethod
    def _sale_row_is_available(cls, ws, row: int, sales_cols: dict[str, str]) -> bool:
        """True se nenhuma célula da tabela de vendas (A–I) na linha tem dado de venda."""
        if row < DATA_START_ROW or row > cls.MAX_DATA_ROW:
            return False
        for col in cls._sale_data_columns(sales_cols):
            if cls._cell_has_content(ws[f"{col}{row}"].value):
                return False
        return True

    @classmethod
    def _next_sale_data_row(
        cls,
        ws,
        sales_cols: dict[str, str],
        *,
        start_row: int = DATA_START_ROW,
    ) -> int:
        """
        Primeira linha livre na tabela (3, 4, 5, 6…).
        Totais na coluna J não impedem usar a linha — o robô só grava em A–I.
        """
        row = start_row
        scan_until = min(
            cls.MAX_DATA_ROW,
            max(min(ws.max_row, cls.MAX_DATA_ROW) + 30, start_row + 30),
        )
        while row <= scan_until:
            if cls._sale_row_is_available(ws, row, sales_cols):
                return row
            row += 1
        return min(row, cls.MAX_DATA_ROW)

    @classmethod
    def _sales_columns(cls, ws) -> dict[str, str]:
        headers = cls._header_map(ws, header_row=SALES_HEADER_ROW, max_col=SALES_TABLE_LAST_COL)
        # Compatibilidade: algumas versões antigas da planilha usavam "Data".
        # Se existir "Data de venda", tratamos como a data principal.
        if "data" not in headers and "data de venda" in headers:
            headers["data"] = headers["data de venda"]
        if "data de venda" not in headers and "data" in headers:
            headers["data de venda"] = headers["data"]
        # Variações comuns de entrega
        if "data de entrega" not in headers:
            if "data entrega" in headers:
                headers["data de entrega"] = headers["data entrega"]
            elif "entrega" in headers:
                headers["data de entrega"] = headers["entrega"]
        # Compatibilidade: coluna "Cliente" substituiu "ID Cliente" em planilhas novas.
        if "cliente" not in headers and "id cliente" in headers:
            headers["cliente"] = headers["id cliente"]
        if "id cliente" not in headers and "cliente" in headers:
            headers["id cliente"] = headers["cliente"]
        missing = [label for key, label in SALES_REQUIRED_HEADERS.items() if key not in headers]
        if missing:
            raise ValueError(
                "Colunas obrigatorias da aba de vendas nao encontradas: " + ", ".join(missing)
            )
        return headers

    @staticmethod
    def _to_float(value) -> float:
        if value in (None, ""):
            return 0.0
        try:
            return float(value)
        except Exception:
            return 0.0

    @staticmethod
    def _status_text(pending_amount: float) -> str:
        return "pago" if abs(float(pending_amount or 0.0)) < 0.01 else "pendente"

    @staticmethod
    def _delivery_fill_for_pending(pending_amount: float):
        """Data de entrega segue status financeiro: pendente=amarelo, pago=verde."""
        return FILL_DONE if abs(float(pending_amount or 0.0)) < 0.01 else FILL_PENDING

    @staticmethod
    def _parse_date_cell(value):
        """Tenta interpretar valores da célula como data (string DD/MM/YYYY ou objeto date/datetime)."""
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.date()
        if hasattr(value, "strftime"):
            # date (sem .date()) ou datetime (já coberto acima)
            try:
                return value.date()  # type: ignore[attr-defined]
            except Exception:
                return value  # type: ignore[return-value]
        s = str(value).strip()
        if not s:
            return None
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                continue
        return None

    @staticmethod
    def _match_text_case(template_value, text: str) -> str:
        if not text:
            return text
        template = str(template_value or "").strip()
        if not template:
            return text
        letters = [ch for ch in template if ch.isalpha()]
        if letters and all(ch.isupper() for ch in letters):
            return text.upper()
        if template == template.capitalize():
            return text.capitalize()
        if template == template.lower():
            return text.lower()
        return text

    def _generate_sale_id(self, ws, sale_id_col: str) -> str:
        used_short_ids: set[int] = set()
        max_id = 0
        for row in range(DATA_START_ROW, ws.max_row + 1):
            value = ws[f"{sale_id_col}{row}"].value
            if value in (None, ""):
                continue
            digits = re.findall(r"\d+", str(value))
            if not digits:
                continue
            numeric = int(digits[-1])
            if 1 <= numeric <= 999:
                used_short_ids.add(numeric)
            max_id = max(max_id, numeric)
        for candidate in range(1, 1000):
            if candidate not in used_short_ids:
                return f"{candidate:03d}"
        return str(max_id + 1)

    def _sale_id_exists_in_sales(self, ws, sale_id_col: str, sale_id: str) -> bool:
        target = str(sale_id or "").strip()
        if not target:
            return False
        for row in range(DATA_START_ROW, min(ws.max_row + 1, self.MAX_DATA_ROW)):
            if str(ws[f"{sale_id_col}{row}"].value or "").strip() == target:
                return True
        return False

    @classmethod
    def _material_columns(cls, ws) -> dict[str, str]:
        headers = cls._header_map(ws, header_row=MATERIAL_HEADER_ROW)
        normalized_map = {
            "data": headers.get("data"),
            "fornecedor": headers.get("fornecedor"),
            "descricao": headers.get("descricao") or headers.get("descrição"),
            "valor": next((col for key, col in headers.items() if key.startswith("valor")), None),
            "id venda": headers.get("id venda"),
        }
        missing = [label for key, label in MATERIAL_REQUIRED_HEADERS.items() if not normalized_map.get(key)]
        if missing:
            raise ValueError(
                "Colunas obrigatorias da aba de materia-prima nao encontradas: " + ", ".join(missing)
            )
        return normalized_map

    @staticmethod
    def _column_style(ws, col: str):
        style_id = getattr(ws.column_dimensions[col], "style", 0) or 0
        if style_id <= 0:
            return None
        try:
            return copy(ws.parent._cell_styles[style_id])
        except Exception:
            return None

    @staticmethod
    def _copy_row_style(
        ws,
        source_row: int,
        target_row: int,
        cols: tuple[str, ...],
        *,
        apply_row_dimensions: bool = True,
    ) -> None:
        SpreadsheetService._copy_row_style_between(
            ws, source_row, ws, target_row, cols, apply_row_dimensions=apply_row_dimensions
        )

    @staticmethod
    def _copy_cell_style(source_cell, target_cell) -> None:
        """Copia o estilo da célula de origem para a de destino (ex.: da linha 1048576).
        Mantém exatamente o padrão: negrito, cor da fonte, fundo, número/data, bordas, alinhamento."""
        if getattr(source_cell, "style_id", 0):
            target_cell._style = copy(source_cell._style)
        num_fmt = getattr(source_cell, "number_format", None)
        if num_fmt is not None and str(num_fmt).strip():
            target_cell.number_format = num_fmt
        try:
            if getattr(source_cell, "font", None) is not None:
                target_cell.font = copy(source_cell.font)
        except Exception:
            pass
        try:
            if getattr(source_cell, "fill", None) is not None:
                target_cell.fill = copy(source_cell.fill)
        except Exception:
            pass
        try:
            if getattr(source_cell, "alignment", None) is not None:
                target_cell.alignment = copy(source_cell.alignment)
        except Exception:
            pass
        try:
            if getattr(source_cell, "border", None) is not None:
                target_cell.border = copy(source_cell.border)
        except Exception:
            pass

    @classmethod
    def _should_copy_row_dimensions_for_target(cls, target_ws, logical_sheet_name: str) -> bool:
        """
        Na aba de vendas visível, não copiar row_dimensions do template: isso afeta a linha inteira
        e desalinha/oculta blocos de resumo ao lado da tabela (ex.: coluna J).
        """
        if cls._normalize_name(logical_sheet_name) != cls._normalize_name(SHEET_SALES):
            return True
        tmpl_title = cls._template_sheet_name_for(SHEET_SALES)
        return cls._normalize_name(getattr(target_ws, "title", "")) == cls._normalize_name(tmpl_title)

    @staticmethod
    def _copy_row_style_between(
        source_ws,
        source_row: int,
        target_ws,
        target_row: int,
        cols: tuple[str, ...],
        *,
        apply_row_dimensions: bool = True,
    ) -> None:
        for col in cols:
            source = source_ws[f"{col}{source_row}"]
            target = target_ws[f"{col}{target_row}"]
            SpreadsheetService._copy_cell_style(source, target)
        if not apply_row_dimensions:
            return
        source_dim = source_ws.row_dimensions[source_row]
        target_dim = target_ws.row_dimensions[target_row]
        if source_dim.height is not None:
            target_dim.height = source_dim.height
        target_dim.hidden = source_dim.hidden
        target_dim.outlineLevel = source_dim.outlineLevel
        target_dim.collapsed = source_dim.collapsed

    @staticmethod
    def _sheet_style_cols(ws) -> tuple[str, ...]:
        return tuple(get_column_letter(col_idx) for col_idx in range(1, ws.max_column + 1))

    @classmethod
    def _layout_cols_for_sheet(cls, ws, logical_sheet_name: str) -> tuple[str, ...]:
        logical_norm = cls._normalize_name(logical_sheet_name)
        if logical_norm == cls._normalize_name(SHEET_SALES):
            # Vendas: aplicar layout APENAS nas colunas da tabela de vendas.
            # Não usar a linha inteira porque isso pode sobrescrever blocos de layout/resumo
            # fora da tabela (ex.: coluna J com totais/fórmulas/caixas).
            return cls._sales_style_cols(ws)
        if logical_norm == cls._normalize_name(SHEET_MATERIAL):
            material_cols = cls._material_columns(ws)
            cols = (
                material_cols["data"],
                material_cols["descricao"],
                material_cols["valor"],
            )
            if material_cols.get("fornecedor"):
                cols = (material_cols["fornecedor"],) + cols
            if material_cols.get("id venda"):
                cols = cols + (material_cols["id venda"],)
            return cols
        if logical_norm == cls._normalize_name(SHEET_FIXED):
            return ("A", "B", "C")
        return cls._sheet_style_cols(ws)

    @classmethod
    def _sales_style_cols(cls, ws) -> tuple[str, ...]:
        """
        Colunas de estilo/dados da tabela de vendas.

        Importante: não usar a linha inteira para evitar sobrescrever colunas
        auxiliares de layout/resumo (ex.: coluna J com totais/fórmulas).
        """
        sales_cols = cls._sales_columns(ws)
        cols = [
            sales_cols["data de venda"],
            sales_cols["data de entrega"],
            sales_cols["cliente"],
            sales_cols["id produto"],
            sales_cols["total de vendas (pago)"],
            sales_cols["valor (pendente)"],
            sales_cols["id venda"],
            sales_cols["status de valor"],
        ]
        # Algumas planilhas têm também a coluna "Status" (serviço) ao lado.
        status_col = sales_cols.get("status")
        if status_col:
            cols.append(status_col)
        # Remove duplicadas preservando ordem.
        return tuple(dict.fromkeys(cols))

    @staticmethod
    def _row_is_empty(ws, row: int, cols: tuple[str, ...]) -> bool:
        return all(ws[f"{col}{row}"].value in (None, "") for col in cols)

    @staticmethod
    def _row_has_values(ws, row: int, cols: tuple[str, ...]) -> bool:
        return any(ws[f"{col}{row}"].value not in (None, "") for col in cols)

    @staticmethod
    def _row_has_any_style(ws, row: int, cols: tuple[str, ...]) -> bool:
        return any((getattr(ws[f"{col}{row}"], "style_id", 0) or 0) > 0 for col in cols)

    @staticmethod
    def _row_has_any_border(ws, row: int, cols: tuple[str, ...]) -> bool:
        """
        Detecta se a linha já tem bordas aplicadas nas colunas alvo.
        Isso é importante porque alguns layouts no Excel podem ter bordas sem "style_id" não-zero.
        """
        for col in cols:
            cell = ws[f"{col}{row}"]
            border = getattr(cell, "border", None)
            if border is None:
                continue
            for side in (border.left, border.right, border.top, border.bottom):
                if getattr(side, "style", None):
                    return True
        return False

    @staticmethod
    def _clear_row_values(ws, row: int, cols: tuple[str, ...]) -> None:
        for col in cols:
            ws[f"{col}{row}"] = None

    @staticmethod
    def _style_row_for(ws, cols: tuple[str, ...], start_row: int = 3, search_limit: int = 20) -> int:
        last_row = min(ws.max_row, start_row + search_limit)
        for row in range(start_row, last_row + 1):
            if any(getattr(ws[f"{col}{row}"], "style_id", 0) for col in cols):
                return row
        return start_row

    @staticmethod
    def _empty_style_row_for(ws, cols: tuple[str, ...], start_row: int = 3, search_limit: int = 50) -> int:
        last_row = min(ws.max_row, start_row + search_limit)
        for row in range(start_row, last_row + 1):
            if not any(getattr(ws[f"{col}{row}"], "style_id", 0) for col in cols):
                continue
            if all(ws[f"{col}{row}"].value in (None, "") for col in cols):
                return row
        return SpreadsheetService._style_row_for(ws, cols, start_row=start_row, search_limit=search_limit)

    @staticmethod
    def _template_style_row_for(ws, cols: tuple[str, ...], start_row: int = 3, search_limit: int = 50) -> int:
        if any(getattr(ws[f"{col}{start_row}"], "style_id", 0) for col in cols):
            return start_row
        return SpreadsheetService._style_row_for(ws, cols, start_row=start_row, search_limit=search_limit)

    @staticmethod
    def _visible_template_row_for(ws, cols: tuple[str, ...], start_row: int = 3, search_limit: int = 50) -> int | None:
        last_row = min(max(ws.max_row, start_row), start_row + search_limit)
        for row in range(start_row, last_row + 1):
            if any(getattr(ws[f"{col}{row}"], "style_id", 0) for col in cols):
                return row
        return None

    @classmethod
    def _row_has_padrao(cls, ws, row: int, cols: tuple[str, ...]) -> bool:
        """True se a linha tiver formatação ou número/data (padrão definido)."""
        if cls._row_has_any_style(ws, row, cols):
            return True
        for col in cols:
            nf = str(getattr(ws[f"{col}{row}"], "number_format", None) or "General").strip()
            if nf and nf != "General":
                return True
        return False

    @classmethod
    def _row_3_has_padrao(cls, ws, cols: tuple[str, ...]) -> bool:
        """True se a linha 3 tiver formatação ou número/data (padrão que o usuário definiu)."""
        return cls._row_has_padrao(ws, TEMPLATE_ROW, cols)

    @classmethod
    def _save_padrao_row3_to_template(cls, wb, ws, logical_sheet_name: str) -> None:
        """Salva o padrão da linha 3 no template. Se a linha 1048576 já for o padrão (tem formatação ou dados), não sobrescreve."""
        if cls._preserve_user_layout():
            return
        # Para abas com linha-âncora (1048576), não usar linha 3 como fonte.
        if cls._anchor_template_row_for(logical_sheet_name) is not None:
            return
        style_cols = cls._layout_cols_for_sheet(ws, logical_sheet_name)
        if cls._anchor_template_row_for(logical_sheet_name) is not None:
            all_cols = cls._sheet_style_cols(ws)
            if cls._row_has_padrao(ws, ANCHOR_TEMPLATE_ROW, style_cols) or cls._row_has_values(ws, ANCHOR_TEMPLATE_ROW, all_cols):
                return
        if not cls._row_3_has_padrao(ws, style_cols):
            return
        cls._sync_template_store(wb, ws, logical_sheet_name, source_row=TEMPLATE_ROW)

    @classmethod
    def _load_padrao_from_anchor_row(cls, wb, ws, logical_sheet_name: str, cols: tuple[str, ...]):
        """Linha 1048576 = padrão único. Lê o estado atual dessa linha e replica o estilo para o template;
        todas as linhas que o robô preencher seguem esse padrão. A linha 1048576 não é apagada."""
        if cls._anchor_template_row_for(logical_sheet_name) is None:
            return None
        tem_padrao = cls._row_has_padrao(ws, ANCHOR_TEMPLATE_ROW, cols)
        tem_valor = cls._row_has_values(ws, ANCHOR_TEMPLATE_ROW, cols)
        if not tem_padrao and not tem_valor:
            return None
        if cls._preserve_user_layout():
            return ws, ANCHOR_TEMPLATE_ROW
        template_ws = cls._ensure_template_sheet(wb, logical_sheet_name)
        cls._copy_row_style_between(ws, ANCHOR_TEMPLATE_ROW, template_ws, TEMPLATE_ROW, cols)
        cls._clear_row_values(template_ws, TEMPLATE_ROW, cols)
        return template_ws, TEMPLATE_ROW

    @staticmethod
    def _append_template_row(ws, row: int, cols: tuple[str, ...], start_row: int = 3, search_limit: int = 50) -> int:
        template_row = SpreadsheetService._template_style_row_for(
            ws,
            cols,
            start_row=start_row,
            search_limit=search_limit,
        )
        if template_row:
            return template_row
        return SpreadsheetService._style_row_for(ws, cols, start_row=start_row, search_limit=search_limit)

    @classmethod
    def _template_sheet_name_for(cls, logical_sheet_name: str) -> str:
        logical_norm = cls._normalize_name(logical_sheet_name)
        if logical_norm == cls._normalize_name(SHEET_SALES):
            return SHEET_TEMPLATE_SALES
        if logical_norm == cls._normalize_name(SHEET_MATERIAL):
            return SHEET_TEMPLATE_MATERIAL
        if logical_norm == cls._normalize_name(SHEET_FIXED):
            return SHEET_TEMPLATE_FIXED
        raise ValueError(f"Aba sem template configurado: {logical_sheet_name}")

    @classmethod
    def _anchor_template_row_for(cls, logical_sheet_name: str) -> int | None:
        """Linha 1048576 como padrão para TOTAL DE VENDAS e Compras Matéria-Prima."""
        logical_norm = cls._normalize_name(logical_sheet_name)
        if logical_norm in (
            cls._normalize_name(SHEET_SALES),
            cls._normalize_name(SHEET_MATERIAL),
        ):
            return ANCHOR_TEMPLATE_ROW
        return None

    @classmethod
    def _ensure_template_sheet(cls, wb, logical_sheet_name: str):
        template_name = cls._template_sheet_name_for(logical_sheet_name)
        if template_name in wb.sheetnames:
            ws = wb[template_name]
        else:
            ws = wb.create_sheet(template_name)
        ws.sheet_state = "veryHidden"
        return ws

    @staticmethod
    def _row_empty_and_no_style(ws, row: int, cols: tuple[str, ...]) -> bool:
        """True se a linha não tem valores e não tem formatação (evita sobrescrever padrão)."""
        if any(ws[f"{col}{row}"].value not in (None, "") for col in cols):
            return False
        if any((getattr(ws[f"{col}{row}"], "style_id", 0) or 0) > 0 for col in cols):
            return False
        if any(str(getattr(ws[f"{col}{row}"], "number_format", None) or "General").strip() not in ("", "General") for col in cols):
            return False
        return True

    @staticmethod
    def _copy_sheet_layout(
        source_ws,
        target_ws,
        row_limit: int = 12,
        style_cols: tuple[str, ...] | None = None,
        *,
        layout_cols: tuple[str, ...] | None = None,
    ) -> None:
        max_col = source_ws.max_column
        all_cols = tuple(get_column_letter(col_idx) for col_idx in range(1, max_col + 1))
        cols_to_check = style_cols or all_cols
        copy_cols = layout_cols or style_cols or all_cols
        copy_col_idx = {column_index_from_string(col) for col in copy_cols}
        for merged in list(target_ws.merged_cells.ranges):
            target_ws.unmerge_cells(str(merged))
        for row in range(1, min(source_ws.max_row, row_limit) + 1):
            # Linhas de dados vazias e sem formato: não sobrescrever o template (mantém padrão da planilha)
            if row >= DATA_START_ROW and SpreadsheetService._row_empty_and_no_style(source_ws, row, cols_to_check):
                continue
            SpreadsheetService._copy_row_style_between(source_ws, row, target_ws, row, copy_cols)
            for col_idx in copy_col_idx:
                source = source_ws.cell(row=row, column=col_idx)
                target = target_ws.cell(row=row, column=col_idx)
                target.value = source.value
        for merged in source_ws.merged_cells.ranges:
            if merged.min_col in copy_col_idx and merged.max_col in copy_col_idx:
                target_ws.merge_cells(str(merged))
        for col_idx in copy_col_idx:
            col = get_column_letter(col_idx)
            source_dim = source_ws.column_dimensions[col]
            target_dim = target_ws.column_dimensions[col]
            target_dim.width = source_dim.width
            target_dim.hidden = source_dim.hidden
            target_dim.bestFit = source_dim.bestFit
            target_dim.outlineLevel = source_dim.outlineLevel
            target_dim.collapsed = source_dim.collapsed

    @staticmethod
    def _template_row_score(ws, row: int, cols: tuple[str, ...]) -> tuple[int, int, int, int]:
        style_ids = [getattr(ws[f"{col}{row}"], "style_id", 0) or 0 for col in cols]
        non_zero = [style_id for style_id in style_ids if style_id > 0]
        if not non_zero:
            return (0, 0, 0, 0)
        distinct_non_default_styles = len({style_id for style_id in non_zero if style_id != 1})
        special_numfmts = sum(1 for col in cols if str(ws[f"{col}{row}"].number_format or "General") != "General")
        special_fonts = sum(
            1
            for col in cols
            if (
                (getattr(ws[f"{col}{row}"], "style_id", 0) or 0) not in (0, 1)
                and ws[f"{col}{row}"].font.name not in (None, "", "Calibri")
            )
        )
        colored_fonts = sum(
            1
            for col in cols
            if (
                getattr(ws[f"{col}{row}"].font.color, "type", None) == "rgb"
                and getattr(ws[f"{col}{row}"].font.color, "rgb", None) not in (None, "00000000")
            )
        )
        return (distinct_non_default_styles, special_numfmts, special_fonts, colored_fonts)

    @classmethod
    def _row_has_rich_style(cls, ws, row: int, cols: tuple[str, ...]) -> bool:
        distinct_non_default_styles, special_numfmts, special_fonts, colored_fonts = cls._template_row_score(ws, row, cols)
        return (distinct_non_default_styles > 2) or (special_numfmts > 0) or (special_fonts > 1) or (colored_fonts > 0)

    @classmethod
    def _visible_template_is_rich(cls, ws, cols: tuple[str, ...]) -> bool:
        return cls._row_has_rich_style(ws, TEMPLATE_ROW, cols)

    @classmethod
    def _sync_template_store(cls, wb, source_ws, logical_sheet_name: str, source_row: int = TEMPLATE_ROW) -> None:
        if cls._preserve_user_layout():
            return
        template_ws = cls._ensure_template_sheet(wb, logical_sheet_name)
        style_cols = cls._layout_cols_for_sheet(source_ws, logical_sheet_name)
        cls._copy_sheet_layout(
            source_ws, template_ws, row_limit=12, style_cols=style_cols, layout_cols=style_cols
        )
        if source_row != TEMPLATE_ROW:
            cls._copy_row_style_between(source_ws, source_row, template_ws, TEMPLATE_ROW, style_cols)
        # Hidden template sheets should keep only layout/style, never previous business data.
        for row in range(DATA_START_ROW, min(template_ws.max_row, 12) + 1):
            cls._clear_row_values(template_ws, row, style_cols)
        # Guardar o padrão também na última linha da planilha (só estilo, sem valores), como referência permanente.
        # O robô nunca escreve dados lá; usa só para copiar o padrão e não sobrescrever fora dele.
        if (
            not cls._preserve_user_layout()
            and source_row == TEMPLATE_ROW
            and cls._anchor_template_row_for(logical_sheet_name) is not None
        ):
            cls._copy_row_style_between(source_ws, TEMPLATE_ROW, source_ws, ANCHOR_TEMPLATE_ROW, style_cols)
            cls._clear_row_values(source_ws, ANCHOR_TEMPLATE_ROW, style_cols)

    @classmethod
    def _template_source(cls, wb, ws, logical_sheet_name: str, cols: tuple[str, ...]):
        anchored_sheet = cls._anchor_template_row_for(logical_sheet_name) is not None
        anchor_row = cls._anchor_template_row_for(logical_sheet_name)

        if cls._preserve_user_layout():
            result = cls._load_padrao_from_anchor_row(wb, ws, logical_sheet_name, cols)
            if result is not None:
                return result
            if anchored_sheet and anchor_row:
                return ws, anchor_row
            if cls._row_3_has_padrao(ws, cols):
                return ws, TEMPLATE_ROW
            return ws, TEMPLATE_ROW

        template_ws = cls._ensure_template_sheet(wb, logical_sheet_name)
        anchored_sheet = cls._anchor_template_row_for(logical_sheet_name) is not None
        # Importante: em abas "ancoradas", o padrão é a linha 1048576.
        # O robô nunca deve tentar "atualizar" esse padrão automaticamente a partir da linha 3,
        # para não alterar bordas/layout que o usuário considera como referência fixa.
        # Padrão único: linha 1048576. Replicar esse estilo em todas as linhas que o robô preencher.
        # TOTAL DE VENDAS DE 2026 e Compras Matéria-Prima: mesma regra — usar só a linha 1048576 quando tiver conteúdo ou formatação.
        result = cls._load_padrao_from_anchor_row(wb, ws, logical_sheet_name, cols)
        if result is not None:
            return result

        # Em abas ancoradas, nunca cair para linha 3 visível como fonte.
        if anchored_sheet:
            if any(getattr(template_ws[f"{col}{TEMPLATE_ROW}"], "style_id", 0) for col in cols):
                return template_ws, TEMPLATE_ROW
            return ws, ANCHOR_TEMPLATE_ROW

        # Fallback: template já tem o padrão salvo (linha 3 com estilo).
        if cls._row_3_has_padrao(template_ws, cols):
            return template_ws, TEMPLATE_ROW

        # Prioridade 3: linha 3 visível tiver o padrão (salvar quando o usuário formatar a linha 3).
        if cls._row_3_has_padrao(ws, cols):
            cls._sync_template_store(wb, ws, logical_sheet_name, source_row=TEMPLATE_ROW)
            return template_ws, TEMPLATE_ROW

        visible_template_row = cls._visible_template_row_for(ws, cols)
        if visible_template_row is not None:
            # Só sincronizar se for a linha 3 (não sobrescrever o template com outras linhas)
            if visible_template_row == TEMPLATE_ROW:
                cls._sync_template_store(wb, ws, logical_sheet_name, source_row=TEMPLATE_ROW)
            return ws, visible_template_row

        anchor_row = cls._anchor_template_row_for(logical_sheet_name)
        if anchor_row and cls._row_has_padrao(ws, anchor_row, cols):
            all_cols = cls._sheet_style_cols(ws)
            template_ready = cls._row_has_any_style(template_ws, TEMPLATE_ROW, cols)
            if cls._row_has_values(ws, anchor_row, all_cols) or not template_ready:
                cls._sync_template_store(wb, ws, logical_sheet_name, source_row=anchor_row)
            return template_ws, TEMPLATE_ROW
        if any(getattr(template_ws[f"{col}{TEMPLATE_ROW}"], "style_id", 0) for col in cols):
            return template_ws, TEMPLATE_ROW
        cls._sync_template_store(wb, ws, logical_sheet_name)
        return ws, TEMPLATE_ROW

    @classmethod
    def _copy_sales_row_style_from_template(cls, wb, ws_sales, row: int) -> None:
        cols = cls._sales_style_cols(ws_sales)
        cls._apply_anchor_row_style(wb, ws_sales, row, SHEET_SALES, cols)

    @classmethod
    def _prepare_row_from_template(
        cls,
        wb,
        ws,
        row: int,
        logical_sheet_name: str,
        cols: tuple[str, ...],
        start_row: int = TEMPLATE_ROW,
        search_limit: int = 50,
    ):
        if cls._row_has_values(ws, row, cols):
            return cls._template_source(wb, ws, logical_sheet_name, cols)
        return cls._apply_anchor_row_style(wb, ws, row, logical_sheet_name, cols)

    @classmethod
    def _normalize_sheet_layout(
        cls,
        wb,
        ws,
        logical_sheet_name: str,
        *,
        row_limit: int,
    ) -> None:
        if cls._preserve_user_layout():
            return
        style_cols = cls._layout_cols_for_sheet(ws, logical_sheet_name)
        template_ws, template_row = cls._template_source(wb, ws, logical_sheet_name, style_cols)
        effective_limit = cls._effective_layout_row_limit(ws, logical_sheet_name, style_cols, minimum_rows=row_limit)
        summary_rows: set[int] = set()
        if cls._normalize_name(logical_sheet_name) == cls._normalize_name(SHEET_SALES):
            summary_rows = cls._sales_summary_rows(ws)
        for row in range(TEMPLATE_ROW, effective_limit + 1):
            if row in summary_rows:
                continue
            if template_ws is ws and template_row == row:
                continue
            if cls._row_has_values(ws, row, style_cols):
                continue
            # Nunca sobrescrever formatação existente: só aplicar template em linhas sem estilo
            if cls._row_has_any_style(ws, row, style_cols) or cls._row_has_any_border(ws, row, style_cols):
                continue
            cls._copy_row_style_between(
                template_ws,
                template_row,
                ws,
                row,
                style_cols,
                apply_row_dimensions=cls._should_copy_row_dimensions_for_target(ws, logical_sheet_name),
            )

    @classmethod
    def _effective_layout_row_limit(
        cls,
        ws,
        logical_sheet_name: str,
        cols: tuple[str, ...],
        *,
        minimum_rows: int,
        empty_streak: int = 25,
    ) -> int:
        anchor_row = cls._anchor_template_row_for(logical_sheet_name)
        max_scan_row = (anchor_row - 1) if anchor_row else ws.max_row
        last_non_empty = TEMPLATE_ROW
        streak = 0
        row = TEMPLATE_ROW
        minimum_target = max(TEMPLATE_ROW + minimum_rows - 1, TEMPLATE_ROW + empty_streak)
        while row <= max_scan_row:
            if cls._row_has_values(ws, row, cols):
                last_non_empty = row
                streak = 0
            else:
                streak += 1
                if row >= minimum_target and streak >= empty_streak:
                    break
            row += 1
        return max(last_non_empty + empty_streak, minimum_target)

    @classmethod
    def normalize_template_rows(
        cls,
        workbook_path: str | Path,
        *,
        row_limit: int = 12,
    ) -> None:
        service = cls(workbook_path)
        wb = service._open_workbook()
        try:
            sheet_names = (
                service._resolve_sheet_name(wb, SHEET_SALES),
                service._resolve_sheet_name(wb, SHEET_MATERIAL),
                service._resolve_sheet_name(wb, SHEET_FIXED),
            )
            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                service._normalize_sheet_layout(wb, ws, sheet_name, row_limit=row_limit)
            service._save_workbook(wb)
        finally:
            wb.close()

    @staticmethod
    def _restore_visible_sheet(ws) -> None:
        for row in range(TEMPLATE_ROW, min(ws.max_row, TEMPLATE_ROW + 2) + 1):
            ws.row_dimensions[row].hidden = False

    def restore_visible_layout(self) -> None:
        wb = self._open_workbook()
        try:
            for sheet_name in (
                self._resolve_sheet_name(wb, SHEET_SALES),
                self._resolve_sheet_name(wb, SHEET_MATERIAL),
                self._resolve_sheet_name(wb, SHEET_FIXED),
            ):
                self._restore_visible_sheet(wb[sheet_name])
            self._save_workbook(wb)
        finally:
            wb.close()

    def _find_sale_row(self, ws, sale_id_col: str, sale_id: str) -> int:
        wanted = self._normalize_name(str(sale_id))
        for row in range(DATA_START_ROW, ws.max_row + 1):
            value = ws[f"{sale_id_col}{row}"].value
            if value in (None, ""):
                continue
            if self._normalize_name(str(value)) == wanted:
                return row
        raise ValueError(f"ID VENDA nao encontrado: {sale_id}")

    def _sale_product_by_id(self, ws, sale_id: str) -> str | None:
        sales_cols = self._sales_columns(ws)
        row = self._find_sale_row(ws, sales_cols["id venda"], sale_id)
        value = ws[f"{sales_cols['id produto']}{row}"].value
        if value in (None, ""):
            return None
        return str(value).strip() or None

    def _cleanup_incomplete_sales_rows(self, ws, row_limit: int = 200) -> None:
        sales_cols = self._sales_columns(ws)
        legacy_status_col = sales_cols.get("status")
        tracked_cols = (
            sales_cols["data de venda"],
            sales_cols["cliente"],
            sales_cols["id produto"],
            sales_cols["total de vendas (pago)"],
            sales_cols["valor (pendente)"],
            sales_cols["id venda"],
            sales_cols["status de valor"],
        )
        if legacy_status_col:
            tracked_cols = tracked_cols + (legacy_status_col,)

        required_sale_cols = (
            sales_cols["id produto"],
            sales_cols["total de vendas (pago)"],
            sales_cols["valor (pendente)"],
            sales_cols["id venda"],
        )
        aux_only_cols = (
            sales_cols["data de venda"],
            sales_cols["cliente"],
            sales_cols["status de valor"],
        )
        if legacy_status_col:
            aux_only_cols = aux_only_cols + (legacy_status_col,)

        for row in range(DATA_START_ROW, min(ws.max_row, row_limit) + 1):
            has_main_data = any(ws[f"{col}{row}"].value not in (None, "") for col in required_sale_cols)
            if has_main_data:
                continue
            has_aux_data = any(ws[f"{col}{row}"].value not in (None, "") for col in aux_only_cols)
            if not has_aux_data:
                continue
            for col in tracked_cols:
                ws[f"{col}{row}"] = None

    def _cleanup_incomplete_material_rows(self, ws, row_limit: int = 200) -> None:
        material_cols = self._material_columns(ws)
        tracked_cols = (
            material_cols["data"],
            material_cols["descricao"],
            material_cols["valor"],
        )
        if material_cols.get("fornecedor"):
            tracked_cols = (material_cols["fornecedor"],) + tracked_cols
        if material_cols.get("id venda"):
            tracked_cols = tracked_cols + (material_cols["id venda"],)

        required_cols = (
            material_cols["descricao"],
            material_cols["valor"],
        )
        if material_cols.get("id venda"):
            required_cols = required_cols + (material_cols["id venda"],)
        aux_cols = (material_cols["data"],)
        if material_cols.get("fornecedor"):
            aux_cols = aux_cols + (material_cols["fornecedor"],)

        for row in range(DATA_START_ROW, min(ws.max_row, row_limit) + 1):
            has_main_data = any(ws[f"{col}{row}"].value not in (None, "") for col in required_cols)
            if has_main_data:
                continue
            has_aux_data = any(ws[f"{col}{row}"].value not in (None, "") for col in aux_cols)
            if not has_aux_data:
                continue
            for col in tracked_cols:
                ws[f"{col}{row}"] = None

    @staticmethod
    def _cleanup_incomplete_fixed_rows(ws, row_limit: int = 200) -> None:
        for row in range(DATA_START_ROW, min(ws.max_row, row_limit) + 1):
            has_main_data = any(ws[f"{col}{row}"].value not in (None, "") for col in ("B", "C"))
            if has_main_data:
                continue
            if ws[f"A{row}"].value in (None, ""):
                continue
            for col in ("A", "B", "C"):
                ws[f"{col}{row}"] = None

    def repair_layout(self, row_limit: int = 50) -> None:
        wb = self._open_workbook()
        try:
            sales_name = self._resolve_sheet_name(wb, SHEET_SALES)
            material_name = self._resolve_sheet_name(wb, SHEET_MATERIAL)
            fixed_name = self._resolve_sheet_name(wb, SHEET_FIXED)

            ws_sales = wb[sales_name]
            ws_material = wb[material_name]
            ws_fixed = wb[fixed_name]

            self._cleanup_incomplete_sales_rows(ws_sales, row_limit=max(row_limit, 200))
            self._cleanup_incomplete_material_rows(ws_material, row_limit=max(row_limit, 200))
            self._cleanup_incomplete_fixed_rows(ws_fixed, row_limit=max(row_limit, 200))

            for sheet_name, ws in (
                (sales_name, ws_sales),
                (material_name, ws_material),
                (fixed_name, ws_fixed),
            ):
                self._normalize_sheet_layout(wb, ws, sheet_name, row_limit=row_limit)

            self._save_workbook(wb)
        finally:
            wb.close()

    @staticmethod
    def _ensure_log_sheet(wb):
        if SHEET_LOG in wb.sheetnames:
            return wb[SHEET_LOG]
        ws = wb.create_sheet(SHEET_LOG)
        ws.append(
            [
                "id",
                "timestamp",
                "origem",
                "cliente",
                "descricao",
                "valor",
                "data_ref",
                "aba",
                "linha",
                "texto_original",
            ]
        )
        return ws

    @staticmethod
    def _ensure_reminders_sheet(wb):
        if SHEET_REMINDERS in wb.sheetnames:
            return wb[SHEET_REMINDERS]
        ws = wb.create_sheet(SHEET_REMINDERS)
        ws.append(
            [
                "ID VENDA",
                "Cliente",
                "Descricao",
                "Data venda",
                "Data entrega",
                "Valor pendente",
                "Status pagamento",
                "Status servico",
                "Chat ID",
                "Criado em",
                "Atualizado em",
            ]
        )
        return ws

    @staticmethod
    def _reminders_header_map(ws) -> dict[str, int]:
        headers: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col_idx).value
            if value in (None, ""):
                continue
            headers[SpreadsheetService._normalize_name(str(value))] = col_idx
        return headers

    def upsert_reminder(
        self,
        wb,
        *,
        sale_id: str,
        customer: str,
        description: str,
        sale_date,
        due_date,
        pending_amount: float,
        payment_status: str,
        service_status: str,
        chat_id: str,
    ) -> None:
        ws = self._ensure_reminders_sheet(wb)
        headers = self._reminders_header_map(ws)
        col_sale_id = headers.get(self._normalize_name("ID VENDA"), 1)
        target_row: int | None = None
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row=row, column=col_sale_id).value or "").strip() == str(sale_id).strip():
                target_row = row
                break
        if target_row is None:
            target_row = ws.max_row + 1

        now = datetime.now()

        def _set(label: str, value) -> None:
            idx = headers.get(self._normalize_name(label))
            if not idx:
                return
            ws.cell(row=target_row, column=idx).value = value

        _set("ID VENDA", sale_id)
        _set("Cliente", customer)
        _set("Descricao", description)
        _set("Data venda", self._excel_date_value(sale_date))
        _set("Data entrega", self._excel_date_value(due_date) if due_date else "")
        _set("Valor pendente", float(pending_amount or 0.0))
        _set("Status pagamento", payment_status)
        _set("Status servico", service_status)
        _set("Chat ID", str(chat_id))
        if not ws.cell(row=target_row, column=headers.get(self._normalize_name("Criado em"), 10)).value:
            _set("Criado em", now)
        _set("Atualizado em", now)

        fill = FILL_DONE if self._normalize_name(service_status) == "finalizado" else FILL_PENDING
        if not self._preserve_user_layout():
            for col in range(1, ws.max_column + 1):
                ws.cell(row=target_row, column=col).fill = fill

    def _existing_reminder_chat_id(self, wb, sale_id: str) -> str:
        if SHEET_REMINDERS not in wb.sheetnames:
            return ""
        ws = wb[SHEET_REMINDERS]
        headers = self._reminders_header_map(ws)
        col_sale_id = headers.get(self._normalize_name("ID VENDA"), 1)
        col_chat = headers.get(self._normalize_name("Chat ID"))
        if not col_chat:
            return ""
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row=row, column=col_sale_id).value or "").strip() == str(sale_id).strip():
                return str(ws.cell(row=row, column=col_chat).value or "").strip()
        return ""

    def refresh_delivery_reminder(
        self,
        sale_id: str,
        chat_id: str = "",
        *,
        mark_delivered: bool = False,
    ) -> None:
        """Sincroniza lembrete de entrega após atualização de pagamento/status."""
        wb = self._open_workbook()
        try:
            sales_name = self._resolve_sheet_name(wb, SHEET_SALES)
            ws_sales = wb[sales_name]
            snap = self._sale_snapshot(ws_sales, str(sale_id).strip())
            if not snap:
                return
            existing_chat = (chat_id or "").strip() or self._existing_reminder_chat_id(wb, sale_id)
            pending_amount = self._to_float(snap.get("pending_amount"))
            payment_status_raw = (snap.get("payment_status") or self._status_text(pending_amount)).strip()
            payment_status = payment_status_raw.upper() if payment_status_raw else ("PENDENTE" if pending_amount > 0.01 else "PAGO")
            due_date = self._parse_date_cell(snap.get("delivery_date"))
            sale_date = self._parse_date_cell(snap.get("sale_date")) or date.today()
            self.upsert_reminder(
                wb,
                sale_id=str(sale_id).strip(),
                customer=snap.get("customer", ""),
                description=snap.get("description", ""),
                sale_date=sale_date,
                due_date=due_date,
                pending_amount=pending_amount,
                payment_status=payment_status,
                service_status="FINALIZADO" if mark_delivered else "PENDENTE",
                chat_id=existing_chat,
            )
            self._save_workbook(wb)
        finally:
            wb.close()

    @classmethod
    def _cell_fill_matches(cls, cell, hex_color: str) -> bool:
        if cell is None or not cell.fill or cell.fill.fill_type != "solid":
            return False
        target = hex_color.upper().lstrip("#")
        for attr in ("start_color", "fgColor"):
            color = getattr(cell.fill, attr, None)
            if color is None:
                continue
            rgb = str(getattr(color, "rgb", "") or getattr(color, "value", "") or "").upper()
            if target in rgb.replace("#", ""):
                return True
        return False

    @classmethod
    def _delivery_date_is_yellow(cls, ws, sales_cols: dict[str, str], row: int) -> bool:
        """True quando a célula Data de Entrega está amarela (ainda não entregue)."""
        delivery_col = sales_cols.get("data de entrega")
        if not delivery_col:
            return False
        return cls._cell_fill_matches(ws[f"{delivery_col}{row}"], "FFF59D")

    @classmethod
    def _delivery_still_pending(
        cls,
        ws,
        sales_cols: dict[str, str],
        row: int,
        *,
        service_status: str = "",
    ) -> bool:
        """True se a entrega ainda não foi marcada (amarelo / não finalizado)."""
        if cls._normalize_name(service_status) == "finalizado":
            return False
        delivery_col = sales_cols.get("data de entrega")
        if delivery_col:
            cell = ws[f"{delivery_col}{row}"]
            if cls._cell_fill_matches(cell, "C8E6C9"):
                return False
            if cls._cell_fill_matches(cell, "FFF59D"):
                return True
        return not cls._is_service_delivered(ws, sales_cols, row)

    def _reminder_meta_by_sale_id(self, wb) -> dict[str, dict[str, str]]:
        """Mapa sale_id -> {chat_id, status_servico} na aba Lembretes."""
        meta: dict[str, dict[str, str]] = {}
        if SHEET_REMINDERS not in wb.sheetnames:
            return meta
        ws = wb[SHEET_REMINDERS]
        headers = self._reminders_header_map(ws)
        col_sale_id = headers.get(self._normalize_name("ID VENDA"), 1)
        col_chat = headers.get(self._normalize_name("Chat ID"))
        col_status = headers.get(self._normalize_name("Status servico"))
        for row in range(2, ws.max_row + 1):
            sale_id = str(ws.cell(row=row, column=col_sale_id).value or "").strip()
            if not sale_id:
                continue
            meta[sale_id] = {
                "chat_id": str(ws.cell(row=row, column=col_chat).value or "").strip() if col_chat else "",
                "status_servico": str(ws.cell(row=row, column=col_status).value or "").strip() if col_status else "",
            }
        return meta

    def list_overdue_deliveries(self, wb, reference_today: date, *, max_rows_scan: int = 500) -> list[dict]:
        """
        Vendas com Data de Entrega amarela (não entregue) e data anterior a reference_today.
        reference_today deve ser sempre date.today() no momento da verificação.
        """
        today = reference_today
        sales_name = self._resolve_sheet_name(wb, SHEET_SALES)
        ws_sales = wb[sales_name]
        sales_cols = self._sales_columns(ws_sales)
        col_sale_id = sales_cols["id venda"]
        col_delivery = sales_cols.get("data de entrega")
        col_customer = sales_cols.get("cliente")
        col_product = sales_cols.get("id produto")
        col_sale_date = sales_cols.get("data de venda")
        if not col_delivery:
            return []

        reminder_meta = self._reminder_meta_by_sale_id(wb)
        rows: list[dict] = []
        scan_until = min(ws_sales.max_row, self.MAX_DATA_ROW, DATA_START_ROW + max_rows_scan)
        for row in range(DATA_START_ROW, scan_until + 1):
            sale_id = str(ws_sales[f"{col_sale_id}{row}"].value or "").strip()
            if not sale_id:
                continue
            delivery_dt = self._parse_date_cell(ws_sales[f"{col_delivery}{row}"].value)
            # Só datas anteriores ao dia de hoje (ex.: hoje 23/06 → cobra 21/06, não 23/06).
            if delivery_dt is None or delivery_dt >= today:
                continue
            sale_dt = (
                self._parse_date_cell(ws_sales[f"{col_sale_date}{row}"].value)
                if col_sale_date
                else None
            )
            # Ignora entrega anterior à data da venda (dado inconsistente na planilha).
            if sale_dt is not None and delivery_dt < sale_dt:
                continue
            if not self._delivery_date_is_yellow(ws_sales, sales_cols, row):
                continue
            rem = reminder_meta.get(sale_id, {})
            customer = str(ws_sales[f"{col_customer}{row}"].value or "").strip() if col_customer else ""
            product = str(ws_sales[f"{col_product}{row}"].value or "").strip() if col_product else ""
            days_overdue = (today - delivery_dt).days
            rows.append(
                {
                    "id venda": sale_id,
                    "cliente": customer,
                    "descricao": product,
                    "data entrega": delivery_dt.strftime("%d/%m/%Y"),
                    "delivery_date": delivery_dt,
                    "days_overdue": days_overdue,
                    "reference_today": today.strftime("%d/%m/%Y"),
                    "chat id": rem.get("chat_id") or "",
                }
            )
        rows.sort(key=lambda item: (item.get("delivery_date") or today, str(item.get("id venda") or "")))
        return rows

    def list_due_reminders(self, wb, today) -> list[dict]:
        """
        Retorna lembretes do dia da entrega com consistência:
        - Entrega: enquanto o serviço não estiver FINALIZADO.
        - Cobrança: enquanto houver valor pendente no mesmo dia da entrega
          (mesmo após marcar como entregue).
        - Ignora IDs que não existam mais na aba de Vendas.
        """
        if SHEET_REMINDERS not in wb.sheetnames:
            return []
        if SHEET_SALES not in wb.sheetnames:
            return []

        ws_rem = wb[SHEET_REMINDERS]
        headers = self._reminders_header_map(ws_rem)
        idx_due = headers.get(self._normalize_name("Data entrega"))
        idx_status = headers.get(self._normalize_name("Status servico"))
        if not idx_due or not idx_status:
            return []

        sales_name = self._resolve_sheet_name(wb, SHEET_SALES)
        ws_sales = wb[sales_name]
        sales_cols = self._sales_columns(ws_sales)

        result: list[dict] = []
        today_txt = self._excel_date_value(today)

        for row in range(2, ws_rem.max_row + 1):
            due_raw = str(ws_rem.cell(row=row, column=idx_due).value or "").strip()
            if not (due_raw and due_raw == today_txt):
                continue

            sale_id = str(ws_rem.cell(row=row, column=headers.get(self._normalize_name("ID VENDA"), 1)).value or "").strip()
            if not sale_id:
                continue

            service_status = str(ws_rem.cell(row=row, column=idx_status).value or "").strip()
            service_finalized = self._normalize_name(service_status) == "finalizado"

            try:
                sales_row = self._find_sale_row(ws_sales, sales_cols["id venda"], sale_id)
            except Exception:
                continue

            customer = str(ws_sales[f"{sales_cols['cliente']}{sales_row}"].value or "").strip()
            product_desc = str(ws_sales[f"{sales_cols['id produto']}{sales_row}"].value or "").strip()
            pending_amount = self._to_float(ws_sales[f"{sales_cols['valor (pendente)']}{sales_row}"].value)
            paid_amount = self._to_float(ws_sales[f"{sales_cols['total de vendas (pago)']}{sales_row}"].value)
            total_amount = round(pending_amount + paid_amount, 2)

            needs_delivery = not service_finalized
            needs_payment = pending_amount > 0.01
            if not needs_delivery and not needs_payment:
                continue

            record: dict = {}
            for key, idx in headers.items():
                record[key] = str(ws_rem.cell(row=row, column=idx).value or "")

            record["id venda"] = sale_id
            record["cliente"] = customer
            record["descricao"] = product_desc
            record["pending_amount_num"] = pending_amount
            record["total_amount_num"] = total_amount
            record["needs_delivery_reminder"] = needs_delivery
            record["needs_payment_reminder"] = needs_payment
            record["service_finalized"] = service_finalized

            result.append(record)

        return result

    def finalize_reminder(self, wb, sale_id: str) -> bool:
        if SHEET_REMINDERS not in wb.sheetnames:
            return False
        ws = wb[SHEET_REMINDERS]
        headers = self._reminders_header_map(ws)
        col_sale_id = headers.get(self._normalize_name("ID VENDA"), 1)
        col_status = headers.get(self._normalize_name("Status servico"))
        col_updated = headers.get(self._normalize_name("Atualizado em"))
        if not col_status:
            return False
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row=row, column=col_sale_id).value or "").strip() != str(sale_id).strip():
                continue
            ws.cell(row=row, column=col_status).value = "FINALIZADO"
            if col_updated:
                ws.cell(row=row, column=col_updated).value = datetime.now()
            if not self._preserve_user_layout():
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = FILL_DONE
            return True
        return False

    def finalize_service(self, sale_id: str) -> bool:
        """Marca serviço como finalizado (Lembretes + cor verde na Data de Entrega da aba de vendas)."""
        wb = self._open_workbook()
        try:
            changed = False
            # 1) Aba Lembretes
            if self.finalize_reminder(wb, sale_id):
                changed = True

            # 2) Aba de vendas: pintar Data de Entrega de verde
            sales_name = self._resolve_sheet_name(wb, SHEET_SALES)
            ws_sales = wb[sales_name]
            sales_cols = self._sales_columns(ws_sales)
            col_sale_id = sales_cols["id venda"]
            col_delivery = sales_cols["data de entrega"]
            for row in range(DATA_START_ROW, min(ws_sales.max_row + 1, self.MAX_DATA_ROW)):
                if str(ws_sales[f"{col_sale_id}{row}"].value or "").strip() != str(sale_id).strip():
                    continue
                pending_col = sales_cols["valor (pendente)"]
                current_pending = self._to_float(ws_sales[f"{pending_col}{row}"].value)
                self._apply_sales_row_visual_status(
                    ws_sales, sales_cols, row, pending_amount=current_pending, service_delivered=True
                )
                changed = True
                break

            if changed:
                self._save_workbook(wb)
            return changed
        finally:
            wb.close()

    def _sale_snapshot(self, ws_sales, sale_id: str) -> dict[str, str]:
        """Lê dados principais da venda pelo ID VENDA para atualizar lembretes mesmo em mensagens de atualização."""
        sales_cols = self._sales_columns(ws_sales)
        col_sale_id = sales_cols["id venda"]
        row_found = None
        for row in range(DATA_START_ROW, min(ws_sales.max_row + 1, self.MAX_DATA_ROW)):
            if str(ws_sales[f"{col_sale_id}{row}"].value or "").strip() == str(sale_id).strip():
                row_found = row
                break
        if row_found is None:
            return {}
        def _v(key: str) -> str:
            col = sales_cols.get(key)
            if not col:
                return ""
            return str(ws_sales[f"{col}{row_found}"].value or "").strip()
        return {
            "sale_id": str(sale_id).strip(),
            "customer": _v("cliente"),
            "description": _v("id produto"),
            "sale_date": _v("data de venda") or _v("data"),
            "delivery_date": _v("data de entrega"),
            "pending_amount": _v("valor (pendente)"),
            "payment_status": _v("status de valor"),
        }

    def _delete_reminder(self, wb, sale_id: str) -> bool:
        if SHEET_REMINDERS not in wb.sheetnames:
            return False
        ws = wb[SHEET_REMINDERS]
        headers = self._reminders_header_map(ws)
        col_sale_id = headers.get(self._normalize_name("ID VENDA"), 1)
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row=row, column=col_sale_id).value or "").strip() != str(sale_id).strip():
                continue
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).value = None
            return True
        return False

    def _count_material_rows_for_sale(self, ws, sale_id: str) -> int:
        material_cols = self._material_columns(ws)
        col_sale_id = material_cols.get("id venda")
        if not col_sale_id:
            return 0
        count = 0
        for row in range(DATA_START_ROW, min(ws.max_row + 1, self.MAX_DATA_ROW)):
            if str(ws[f"{col_sale_id}{row}"].value or "").strip() == str(sale_id).strip():
                count += 1
        return count

    def material_costs_by_sale_id(self, wb, *, max_rows_scan: int = 2000) -> dict[str, float]:
        """Soma custos de material por ID VENDA (sem fornecedor — só o valor)."""
        try:
            material_name = self._resolve_sheet_name(wb, SHEET_MATERIAL)
        except Exception:
            return {}
        if material_name not in wb.sheetnames:
            return {}
        ws = wb[material_name]
        try:
            material_cols = self._material_columns(ws)
        except Exception:
            return {}
        col_sale_id = material_cols.get("id venda")
        col_valor = material_cols.get("valor")
        if not col_sale_id or not col_valor:
            return {}
        totals: dict[str, float] = {}
        end_row = min(ws.max_row, self.MAX_DATA_ROW, DATA_START_ROW + max_rows_scan - 1)
        if getattr(ws, "read_only", False):
            from openpyxl.utils import column_index_from_string

            sid_idx = column_index_from_string(col_sale_id)
            val_idx = column_index_from_string(col_valor)
            for row_cells in ws.iter_rows(
                min_row=DATA_START_ROW,
                max_row=end_row,
                min_col=min(sid_idx, val_idx),
                max_col=max(sid_idx, val_idx),
            ):
                values = {cell.column: cell.value for cell in row_cells}
                sale_id = str(values.get(sid_idx) or "").strip()
                if not sale_id:
                    continue
                amount = self._to_float(values.get(val_idx))
                if amount <= 0:
                    continue
                totals[sale_id] = round(totals.get(sale_id, 0.0) + amount, 2)
        else:
            for row in range(DATA_START_ROW, end_row + 1):
                sale_id = str(ws[f"{col_sale_id}{row}"].value or "").strip()
                if not sale_id:
                    continue
                amount = self._to_float(ws[f"{col_valor}{row}"].value)
                if amount <= 0:
                    continue
                totals[sale_id] = round(totals.get(sale_id, 0.0) + amount, 2)
        return totals

    def _clear_material_rows_for_sale(self, ws, sale_id: str) -> int:
        """Remove linhas de material vinculadas a um ID VENDA (antes de substituir custo)."""
        sale_id = str(sale_id or "").strip()
        if not sale_id:
            return 0
        material_cols = self._material_columns(ws)
        col_sid = material_cols.get("id venda")
        if not col_sid:
            return 0
        mat_cols = (
            material_cols["data"],
            material_cols["descricao"],
            material_cols["valor"],
        )
        if material_cols.get("fornecedor"):
            mat_cols = (material_cols["fornecedor"],) + mat_cols
        mat_cols = mat_cols + (col_sid,)
        cleared = 0
        for mat_row in range(DATA_START_ROW, min(ws.max_row + 1, self.MAX_DATA_ROW)):
            if str(ws[f"{col_sid}{mat_row}"].value or "").strip() != sale_id:
                continue
            self._clear_row_values(ws, mat_row, mat_cols)
            cleared += 1
        return cleared

    def update_sale_pending_amount(self, sale_id: str, new_pending: float) -> bool:
        """Atualiza o valor pendente de uma venda pelo ID VENDA."""
        sale_id = str(sale_id or "").strip()
        if not sale_id:
            return False
        wb = self._open_workbook()
        try:
            sales_name = self._resolve_sheet_name(wb, SHEET_SALES)
            ws_sales = wb[sales_name]
            sales_cols = self._sales_columns(ws_sales)
            try:
                row = self._find_sale_row(ws_sales, sales_cols["id venda"], sale_id)
            except ValueError:
                return False
            paid = self._to_float(ws_sales[f"{sales_cols['total de vendas (pago)']}{row}"].value)
            new_pending = round(max(float(new_pending), 0.0), 2)
            ws_sales[f"{sales_cols['valor (pendente)']}{row}"] = float(new_pending)
            pending_amount = new_pending
            value_status = self._status_text(pending_amount)
            template_ws, template_row = self._apply_anchor_row_style(
                wb, ws_sales, row, SHEET_SALES, self._sales_style_cols(ws_sales)
            )
            value_status_display = self._match_text_case(
                template_ws[f"{sales_cols['status de valor']}{template_row}"].value,
                value_status,
            )
            ws_sales[f"{sales_cols['status de valor']}{row}"] = value_status_display
            self._apply_sales_row_visual_status(
                ws_sales,
                sales_cols,
                row,
                pending_amount=pending_amount,
                service_delivered=False,
            )
            self._save_workbook(wb)
            return True
        finally:
            wb.close()

    def delete_sale(
        self,
        delete_cmd: DeleteSaleCommand,
        original_text: str = "",
        origin: str = "telegram",
    ) -> WriteAction:
        """Remove linha de venda (e material/lembrete vinculados) da planilha."""
        wb = self._open_workbook()
        try:
            sales_name = self._resolve_sheet_name(wb, SHEET_SALES)
            ws_sales = wb[sales_name]
            ws_log = self._ensure_log_sheet(wb)
            sales_cols = self._sales_columns(ws_sales)
            sale_id = str(delete_cmd.sale_id).strip()
            row = self._find_sale_row(ws_sales, sales_cols["id venda"], sale_id)
            snap = self._sale_snapshot(ws_sales, sale_id)
            style_cols = self._sales_style_cols(ws_sales)
            self._clear_row_values(ws_sales, row, style_cols)

            if delete_cmd.remove_material:
                material_name = self._resolve_sheet_name(wb, SHEET_MATERIAL)
                ws_material = wb[material_name]
                material_cols = self._material_columns(ws_material)
                col_sid = material_cols.get("id venda")
                if col_sid:
                    mat_cols = (
                        material_cols["data"],
                        material_cols["descricao"],
                        material_cols["valor"],
                    )
                    if material_cols.get("fornecedor"):
                        mat_cols = (material_cols["fornecedor"],) + mat_cols
                    mat_cols = mat_cols + (col_sid,)
                    for mat_row in range(DATA_START_ROW, min(ws_material.max_row + 1, self.MAX_DATA_ROW)):
                        if str(ws_material[f"{col_sid}{mat_row}"].value or "").strip() != sale_id:
                            continue
                        self._clear_row_values(ws_material, mat_row, mat_cols)

            self._delete_reminder(wb, sale_id)
            action = WriteAction(
                sheet=sales_name,
                row=row,
                amount=0.0,
                label="Exclusao",
                sale_id=sale_id,
            )
            self._append_log(
                ws=ws_log,
                log_id=f"{datetime.now().strftime('%Y%m%d%H%M%S')}-D-1",
                origin=origin,
                cmd=FinancialCommand(
                    customer=(snap or {}).get("customer") or "",
                    description=delete_cmd.reason,
                    sale_date=delete_cmd.ref_date,
                    total_value=0.0,
                    payments=[],
                    sale_id=sale_id,
                    product_id=(snap or {}).get("description") or "",
                ),
                amount=0.0,
                ref_date=delete_cmd.ref_date,
                sheet=sales_name,
                row=row,
                original_text=original_text or delete_cmd.reason,
            )
            self._save_workbook(wb)
            return action
        finally:
            wb.close()

    def _append_sale(self, wb, ws, cmd: FinancialCommand) -> tuple[int, str, float]:
        """Adiciona uma linha de venda. Só preenche valores; layout (fonte, cor, borda) fica como na planilha."""
        sales_cols = self._sales_columns(ws)
        style_cols = self._sales_style_cols(ws)
        row = self._next_sale_data_row(ws, sales_cols)
        template_ws, template_row = self._apply_anchor_row_style(
            wb, ws, row, SHEET_SALES, style_cols
        )

        paid_amount = round(sum(p.value for p in cmd.payments if p.status == "pago"), 2)
        pending_amount = round(sum(p.value for p in cmd.payments if p.status != "pago"), 2)
        if abs((paid_amount + pending_amount) - float(cmd.total_value or 0.0)) >= 0.01:
            pending_amount = round(max(float(cmd.total_value or 0.0) - paid_amount, 0.0), 2)
        sale_id_col = sales_cols["id venda"]
        sale_id = (cmd.sale_id or "").strip()
        if sale_id and self._sale_id_exists_in_sales(ws, sale_id_col, sale_id):
            sale_id = ""
        sale_id = sale_id or self._generate_sale_id(ws, sale_id_col)
        value_status = self._status_text(pending_amount)
        value_status_display = self._match_text_case(
            template_ws[f"{sales_cols['status de valor']}{template_row}"].value,
            value_status,
        )

        # Datas seguindo o padrão da planilha (Data de venda / Data de Entrega)
        ws[f"{sales_cols['data de venda']}{row}"] = self._excel_date_value(cmd.sale_date)
        # Data de Entrega:
        # - Preferência: campo explícito (cmd.service_due_date).
        # - Fallback: vence do "Saldo" (quando o usuário só fala "restante dia 10/04").
        delivery_date = cmd.service_due_date
        if delivery_date is None and cmd.payments:
            for p in cmd.payments:
                if str(getattr(p, "label", "") or "").strip().lower() == "saldo" and getattr(p, "due_date", None):
                    delivery_date = p.due_date
                    break

        if delivery_date:
            ws[f"{sales_cols['data de entrega']}{row}"] = self._excel_date_value(delivery_date)
        ws[f"{sales_cols['cliente']}{row}"] = cmd.customer
        ws[f"{sales_cols['id produto']}{row}"] = cmd.product_id or cmd.description
        ws[f"{sales_cols['total de vendas (pago)']}{row}"] = float(paid_amount)
        ws[f"{sales_cols['valor (pendente)']}{row}"] = float(pending_amount)
        ws[f"{sales_cols['id venda']}{row}"] = sale_id
        ws[f"{sales_cols['status de valor']}{row}"] = value_status_display

        service_delivered = (cmd.service_status or "").strip().lower() == "finalizado"
        self._apply_sales_row_visual_status(
            ws,
            sales_cols,
            row,
            pending_amount=pending_amount,
            service_delivered=service_delivered,
        )
        return row, sale_id, paid_amount

    def _append_material(
        self,
        wb,
        ws,
        supplier: str,
        description: str,
        amount: float,
        ref_date,
        sale_id: str | None = None,
    ) -> int:
        """Adiciona linha na aba Compras Matéria-Prima. Formatação = linha 1048576 dessa aba (Data, Fornecedor, Descrição, ID VENDA, Valor)."""
        material_cols = self._material_columns(ws)
        style_cols = (
            material_cols["data"],
            material_cols["descricao"],
            material_cols["valor"],
        )
        if material_cols.get("fornecedor"):
            style_cols = (material_cols["fornecedor"],) + style_cols
        if material_cols.get("id venda"):
            style_cols = style_cols + (material_cols["id venda"],)
        empty_cols = (material_cols["descricao"], material_cols["valor"])
        if material_cols.get("id venda"):
            empty_cols = empty_cols + (material_cols["id venda"],)
        row = self._next_row_for_empty_cols(ws, empty_cols, start_row=DATA_START_ROW)
        self._prepare_row_from_template(
            wb,
            ws,
            row,
            SHEET_MATERIAL,
            style_cols,
            start_row=TEMPLATE_ROW,
        )
        ws[f"{material_cols['data']}{row}"] = self._excel_date_value(ref_date)
        if material_cols.get("fornecedor"):
            ws[f"{material_cols['fornecedor']}{row}"] = supplier
        ws[f"{material_cols['descricao']}{row}"] = description
        ws[f"{material_cols['valor']}{row}"] = float(amount)
        if material_cols.get("id venda") and sale_id:
            ws[f"{material_cols['id venda']}{row}"] = sale_id
        self._ensure_number_format_if_general(ws, material_cols["data"], row, "DD/MM/YYYY")
        self._ensure_number_format_if_general(ws, material_cols["valor"], row, "R$ #,##0.00")
        return row

    def _append_fixed(self, wb, ws, label: str, amount: float, ref_date) -> int:
        row = self._next_row_for_empty_cols(ws, ("B", "C"), start_row=DATA_START_ROW)
        self._prepare_row_from_template(
            wb,
            ws,
            row,
            SHEET_FIXED,
            ("A", "B", "C"),
            start_row=TEMPLATE_ROW,
        )
        ws[f"A{row}"] = self._excel_date_value(ref_date)
        ws[f"B{row}"] = label
        ws[f"C{row}"] = float(amount)
        self._ensure_number_format_if_general(ws, "A", row, "DD/MM/YYYY")
        self._ensure_number_format_if_general(ws, "C", row, "R$ #,##0.00")
        return row

    @staticmethod
    def _append_log(ws, log_id: str, origin: str, cmd: FinancialCommand, amount: float, ref_date, sheet: str, row: int, original_text: str) -> None:
        ws.append(
            [
                log_id,
                datetime.now(),
                origin,
                cmd.customer,
                cmd.description,
                amount,
                ref_date,
                sheet,
                row,
                original_text,
            ]
        )

    def apply_command(
        self,
        cmd: FinancialCommand,
        original_text: str,
        origin: str = "texto",
        *,
        chat_id: str | None = None,
    ) -> list[WriteAction]:
        wb = self._open_workbook()
        try:
            sales_name = self._resolve_sheet_name(wb, SHEET_SALES)
            material_name = self._resolve_sheet_name(wb, SHEET_MATERIAL)
            fixed_name = self._resolve_sheet_name(wb, SHEET_FIXED)

            ws_sales = wb[sales_name]
            ws_material = wb[material_name]
            ws_fixed = wb[fixed_name]
            ws_log = self._ensure_log_sheet(wb)
            self._cleanup_incomplete_sales_rows(ws_sales)
            self._cleanup_incomplete_material_rows(ws_material)
            self._cleanup_incomplete_fixed_rows(ws_fixed)
            if not self._preserve_user_layout():
                self._save_padrao_row3_to_template(wb, ws_sales, sales_name)
                self._save_padrao_row3_to_template(wb, ws_material, material_name)
                self._save_padrao_row3_to_template(wb, ws_fixed, fixed_name)
                self._normalize_sheet_layout(wb, ws_sales, sales_name, row_limit=120)
                self._normalize_sheet_layout(wb, ws_material, material_name, row_limit=120)
                self._normalize_sheet_layout(wb, ws_fixed, fixed_name, row_limit=120)
            actions: list[WriteAction] = []
            log_root = datetime.now().strftime("%Y%m%d%H%M%S")
            created_sale_id: str | None = cmd.sale_id

            if cmd.total_value > 0:
                row, sale_id, paid_amount = self._append_sale(wb, ws_sales, cmd)
                created_sale_id = sale_id
                actions.append(
                    WriteAction(
                        sheet=sales_name,
                        row=row,
                        amount=cmd.total_value,
                        label="Venda",
                        sale_id=sale_id,
                    )
                )
                self._append_log(
                    ws=ws_log,
                    log_id=f"{log_root}-V-1",
                    origin=origin,
                    cmd=FinancialCommand(
                        customer=cmd.customer,
                        description=cmd.product_id or cmd.description,
                        sale_date=cmd.sale_date,
                        total_value=cmd.total_value,
                        payments=cmd.payments,
                        product_id=cmd.product_id,
                        sale_id=sale_id,
                    ),
                    amount=paid_amount,
                    ref_date=cmd.sale_date,
                    sheet=sales_name,
                    row=row,
                    original_text=original_text,
                )

                # Registrar/atualizar lembrete (aba Lembretes) quando o comando veio do bot.
                if created_sale_id:
                    pending_amount = 0.0
                    latest_due = None
                    for p in cmd.payments:
                        if (p.status or "").strip().lower() != "pago":
                            pending_amount += float(p.value or 0.0)
                            latest_due = max(latest_due, p.due_date) if latest_due else p.due_date
                    payment_status = "PENDENTE" if pending_amount > 0.01 else "PAGO"
                    due_date = cmd.service_due_date or latest_due
                    service_status = "FINALIZADO" if (cmd.service_status or "").strip().lower() == "finalizado" else "PENDENTE"
                    if due_date or pending_amount > 0.01 or (chat_id and str(chat_id).strip()):
                        self.upsert_reminder(
                            wb,
                            sale_id=str(created_sale_id),
                            customer=cmd.customer or "",
                            description=cmd.product_id or cmd.description or "",
                            sale_date=cmd.sale_date,
                            due_date=due_date,
                            pending_amount=pending_amount,
                            payment_status=payment_status,
                            service_status=service_status,
                            chat_id=str(chat_id or ""),
                        )

            if cmd.material_allocations:
                for idx, allocation in enumerate(cmd.material_allocations, start=1):
                    alloc_sale_id = str(allocation.sale_id).strip()
                    if getattr(cmd, "material_replace", False) and alloc_sale_id:
                        self._clear_material_rows_for_sale(ws_material, alloc_sale_id)
                    material_description = allocation.description or self._sale_product_by_id(ws_sales, allocation.sale_id)
                    if not material_description:
                        material_description = f"Material da venda {allocation.sale_id}"
                    mat_row = self._append_material(
                        wb,
                        ws_material,
                        supplier="",
                        description=material_description,
                        amount=allocation.amount,
                        ref_date=allocation.material_date,
                        sale_id=allocation.sale_id,
                    )
                    actions.append(
                        WriteAction(
                            sheet=material_name,
                            row=mat_row,
                            amount=allocation.amount,
                            label="Material",
                            sale_id=allocation.sale_id,
                        )
                    )
                    self._append_log(
                        ws=ws_log,
                        log_id=f"{log_root}-M-{idx}",
                        origin=origin,
                        cmd=FinancialCommand(
                            customer=cmd.customer,
                            description=material_description,
                            sale_date=cmd.sale_date,
                            total_value=0.0,
                            payments=[],
                            sale_id=allocation.sale_id,
                        ),
                        amount=allocation.amount,
                        ref_date=allocation.material_date,
                        sheet=material_name,
                        row=mat_row,
                        original_text=original_text,
                    )
            elif cmd.material_cost and cmd.material_cost > 0:
                material_description = cmd.product_id or cmd.description
                if created_sale_id and (not material_description or str(material_description).startswith("Material da venda ")):
                    material_description = self._sale_product_by_id(ws_sales, created_sale_id) or material_description
                mat_sale_id = created_sale_id
                if getattr(cmd, "material_replace", False) and mat_sale_id:
                    self._clear_material_rows_for_sale(ws_material, str(mat_sale_id))
                mat_row = self._append_material(
                    wb,
                    ws_material,
                    supplier="",
                    description=material_description,
                    amount=cmd.material_cost,
                    ref_date=cmd.material_date or cmd.sale_date,
                    sale_id=created_sale_id,
                )
                actions.append(
                    WriteAction(
                        sheet=material_name,
                        row=mat_row,
                        amount=cmd.material_cost,
                        label="Material",
                        sale_id=created_sale_id,
                    )
                )
                self._append_log(
                    ws=ws_log,
                    log_id=f"{log_root}-M-1",
                    origin=origin,
                    cmd=cmd,
                    amount=cmd.material_cost,
                    ref_date=cmd.material_date or cmd.sale_date,
                    sheet=material_name,
                    row=mat_row,
                    original_text=original_text,
                )

            if cmd.fixed_cost and cmd.fixed_cost > 0:
                fixed_row = self._append_fixed(
                    wb,
                    ws_fixed,
                    label=cmd.fixed_cost_label or "Gasto fixo via audio",
                    amount=cmd.fixed_cost,
                    ref_date=cmd.fixed_cost_date or cmd.sale_date,
                )
                actions.append(WriteAction(sheet=fixed_name, row=fixed_row, amount=cmd.fixed_cost, label="Gasto fixo"))
                self._append_log(
                    ws=ws_log,
                    log_id=f"{log_root}-F-1",
                    origin=origin,
                    cmd=cmd,
                    amount=cmd.fixed_cost,
                    ref_date=cmd.fixed_cost_date or cmd.sale_date,
                    sheet=fixed_name,
                    row=fixed_row,
                    original_text=original_text,
                )

            self._save_workbook(wb)
            return actions
        finally:
            wb.close()

    def apply_refund(
        self,
        refund: RefundCommand,
        original_text: str = "",
        origin: str = "audio",
    ) -> WriteAction:
        """Registra um estorno na aba de vendas (valor negativo) e no log."""
        wb = self._open_workbook()
        try:
            sales_name = self._resolve_sheet_name(wb, SHEET_SALES)
            ws_sales = wb[sales_name]
            ws_log = self._ensure_log_sheet(wb)
            if not self._preserve_user_layout():
                self._save_padrao_row3_to_template(wb, ws_sales, sales_name)
            self._cleanup_incomplete_sales_rows(ws_sales)
            if not self._preserve_user_layout():
                self._normalize_sheet_layout(wb, ws_sales, sales_name, row_limit=120)
            sales_cols = self._sales_columns(ws_sales)
            style_cols = self._sales_style_cols(ws_sales)
            row = self._next_sale_data_row(ws_sales, sales_cols)
            amount = -abs(float(refund.amount))
            sale_id = self._generate_sale_id(ws_sales, sales_cols["id venda"])
            template_ws, template_row = self._apply_anchor_row_style(
                wb, ws_sales, row, SHEET_SALES, style_cols
            )
            status_paid = self._match_text_case(
                template_ws[f"{sales_cols['status de valor']}{template_row}"].value,
                "pago",
            )
            ws_sales[f"{sales_cols['data de venda']}{row}"] = self._excel_date_value(refund.ref_date)
            ws_sales[f"{sales_cols['cliente']}{row}"] = refund.customer
            ws_sales[f"{sales_cols['id produto']}{row}"] = f"Estorno - {refund.reason}"
            ws_sales[f"{sales_cols['total de vendas (pago)']}{row}"] = amount
            ws_sales[f"{sales_cols['valor (pendente)']}{row}"] = 0.0
            ws_sales[f"{sales_cols['id venda']}{row}"] = sale_id
            ws_sales[f"{sales_cols['status de valor']}{row}"] = status_paid
            self._apply_sales_row_visual_status(ws_sales, sales_cols, row, pending_amount=0.0)
            action = WriteAction(sheet=sales_name, row=row, amount=amount, label="Estorno", sale_id=sale_id)
            self._append_log(
                ws=ws_log,
                log_id=f"{datetime.now().strftime('%Y%m%d%H%M%S')}-E-1",
                origin=origin,
                cmd=FinancialCommand(
                    customer=refund.customer,
                    description=refund.reason,
                    sale_date=refund.ref_date,
                    total_value=refund.amount,
                    payments=[],
                    sale_id=sale_id,
                ),
                amount=amount,
                ref_date=refund.ref_date,
                sheet=sales_name,
                row=row,
                original_text=original_text,
            )
            self._save_workbook(wb)
            return action
        finally:
            wb.close()

    def update_sale_status(
        self,
        status_update: StatusUpdateCommand,
        original_text: str = "",
        origin: str = "status-update",
    ) -> WriteAction:
        wb = self._open_workbook()
        try:
            sales_name = self._resolve_sheet_name(wb, SHEET_SALES)
            ws_sales = wb[sales_name]
            ws_log = self._ensure_log_sheet(wb)
            if not self._preserve_user_layout():
                self._save_padrao_row3_to_template(wb, ws_sales, sales_name)
                self._normalize_sheet_layout(wb, ws_sales, sales_name, row_limit=120)
            sales_cols = self._sales_columns(ws_sales)
            row = self._find_sale_row(ws_sales, sales_cols["id venda"], status_update.sale_id)

            paid_col = sales_cols["total de vendas (pago)"]
            pending_col = sales_cols["valor (pendente)"]
            value_status_col = sales_cols["status de valor"]

            current_paid = self._to_float(ws_sales[f"{paid_col}{row}"].value)
            current_pending = self._to_float(ws_sales[f"{pending_col}{row}"].value)
            status = status_update.status.strip().lower()
            amount_moved = 0.0

            if status == "pago" and current_pending > 0:
                amount_moved = current_pending
                current_paid = round(current_paid + current_pending, 2)
                current_pending = 0.0

            status_display = self._match_text_case(ws_sales[f"{value_status_col}{row}"].value, status)
            ws_sales[f"{paid_col}{row}"] = current_paid
            ws_sales[f"{pending_col}{row}"] = current_pending
            ws_sales[f"{value_status_col}{row}"] = status_display
            self._apply_sales_row_visual_status(
                ws_sales, sales_cols, row, pending_amount=current_pending
            )

            customer = status_update.customer or str(ws_sales[f"{sales_cols['cliente']}{row}"].value or "")
            product = str(ws_sales[f"{sales_cols['id produto']}{row}"].value or "")
            self._append_log(
                ws=ws_log,
                log_id=f"{datetime.now().strftime('%Y%m%d%H%M%S')}-S-1",
                origin=origin,
                cmd=FinancialCommand(
                    customer=customer,
                    description=product or f"Atualizacao de status {status_update.sale_id}",
                    sale_date=status_update.ref_date,
                    total_value=current_paid + current_pending,
                    payments=[],
                    product_id=product or None,
                    sale_id=status_update.sale_id,
                ),
                amount=amount_moved,
                ref_date=status_update.ref_date,
                sheet=sales_name,
                row=row,
                original_text=original_text,
            )

            self._save_workbook(wb)
            return WriteAction(
                sheet=sales_name,
                row=row,
                amount=amount_moved,
                label=f"Status {status}",
                sale_id=status_update.sale_id,
            )
        finally:
            wb.close()

    def update_sale_delivery_date(self, sale_id: str, delivery_date) -> None:
        """Atualiza apenas a coluna 'Data de Entrega' para um ID VENDA existente.

        Observação: alguns usuários enviam "cliente id 003" querendo dizer o ID VENDA.
        Por isso, se não achar por ID VENDA, tentamos também casar pelo nome na coluna Cliente.
        """
        wb = self._open_workbook()
        try:
            sales_name = self._resolve_sheet_name(wb, SHEET_SALES)
            ws_sales = wb[sales_name]
            sales_cols = self._sales_columns(ws_sales)
            col_sale_id = sales_cols["id venda"]
            col_delivery = sales_cols["data de entrega"]
            col_customer = sales_cols.get("cliente")
            target_row = None
            needle = str(sale_id).strip()
            # 1) Primeiro tenta por ID VENDA
            for row in range(DATA_START_ROW, min(ws_sales.max_row + 1, self.MAX_DATA_ROW)):
                if str(ws_sales[f"{col_sale_id}{row}"].value or "").strip() == needle:
                    target_row = row
                    break
            # 2) Fallback: tenta pelo nome na coluna Cliente
            if target_row is None and col_customer:
                for row in range(DATA_START_ROW, min(ws_sales.max_row + 1, self.MAX_DATA_ROW)):
                    if str(ws_sales[f"{col_customer}{row}"].value or "").strip() == needle:
                        target_row = row
                        break
            if target_row is None:
                return
            # Não reaplicar estilo da linha inteira em updates: preserva bordas já existentes.
            ws_sales[f"{col_delivery}{target_row}"] = self._excel_date_value(delivery_date)
            pending_col = sales_cols["valor (pendente)"]
            current_pending = self._to_float(ws_sales[f"{pending_col}{target_row}"].value)
            self._apply_sales_row_visual_status(
                ws_sales, sales_cols, target_row, pending_amount=current_pending
            )
            self._save_workbook(wb)
        finally:
            wb.close()

    def apply_partial_payment(self, sale_id: str, amount: float, ref_date: date, original_text: str = "", origin: str = "payment_update") -> WriteAction | None:
        """Move parte do valor pendente -> pago para um ID VENDA (correção de entrada parcial)."""
        if amount is None or float(amount) <= 0:
            return None
        wb = self._open_workbook()
        try:
            sales_name = self._resolve_sheet_name(wb, SHEET_SALES)
            ws_sales = wb[sales_name]
            ws_log = self._ensure_log_sheet(wb)
            sales_cols = self._sales_columns(ws_sales)
            row = self._find_sale_row(ws_sales, sales_cols["id venda"], str(sale_id).strip())
            paid_col = sales_cols["total de vendas (pago)"]
            pending_col = sales_cols["valor (pendente)"]
            status_col = sales_cols["status de valor"]
            delivery_col = sales_cols.get("data de entrega")

            current_paid = self._to_float(ws_sales[f"{paid_col}{row}"].value)
            current_pending = self._to_float(ws_sales[f"{pending_col}{row}"].value)
            move = min(float(amount), float(current_pending))
            current_paid = round(current_paid + move, 2)
            current_pending = round(max(current_pending - move, 0.0), 2)
            ws_sales[f"{paid_col}{row}"] = current_paid
            ws_sales[f"{pending_col}{row}"] = current_pending
            ws_sales[f"{status_col}{row}"] = self._match_text_case(
                ws_sales[f"{status_col}{row}"].value,
                self._status_text(current_pending),
            )
            self._apply_sales_row_visual_status(
                ws_sales, sales_cols, row, pending_amount=current_pending
            )

            customer = str(ws_sales[f"{sales_cols['cliente']}{row}"].value or "")
            product = str(ws_sales[f"{sales_cols['id produto']}{row}"].value or "")
            self._append_log(
                ws=ws_log,
                log_id=f"{datetime.now().strftime('%Y%m%d%H%M%S')}-P-1",
                origin=origin,
                cmd=FinancialCommand(
                    customer=customer,
                    description=product or f"Pagamento parcial {sale_id}",
                    sale_date=ref_date,
                    total_value=current_paid + current_pending,
                    payments=[],
                    product_id=product or None,
                    sale_id=str(sale_id).strip(),
                ),
                amount=move,
                ref_date=ref_date,
                sheet=sales_name,
                row=row,
                original_text=original_text,
            )
            self._save_workbook(wb)
            return WriteAction(sheet=sales_name, row=row, amount=move, label="Pagamento parcial", sale_id=str(sale_id).strip())
        except Exception:
            return None
        finally:
            wb.close()

    def update_row(
        self,
        sheet_name: str,
        row: int,
        ref_date,
        party: str,
        description: str,
        amount: float,
        original_text: str = "",
        origin: str = "correcao",
    ) -> None:
        if row < DATA_START_ROW:
            raise ValueError(f"Linha invalida. Use linhas a partir da {DATA_START_ROW}.")

        wb = self._open_workbook()
        try:
            sales_name = self._resolve_sheet_name(wb, SHEET_SALES)
            material_name = self._resolve_sheet_name(wb, SHEET_MATERIAL)
            fixed_name = self._resolve_sheet_name(wb, SHEET_FIXED)
            target_name = self._resolve_sheet_name(wb, sheet_name)
            ws = wb[target_name]
            ws_log = self._ensure_log_sheet(wb)
            if not self._preserve_user_layout():
                self._save_padrao_row3_to_template(wb, ws, target_name)
                self._normalize_sheet_layout(wb, ws, target_name, row_limit=max(120, row))

            target_norm = self._normalize_name(target_name)
            if target_norm in (self._normalize_name(sales_name), self._normalize_name(material_name)):
                if target_norm == self._normalize_name(sales_name):
                    sales_cols = self._sales_columns(ws)
                    style_cols = self._sales_style_cols(ws)

            # Se foi uma atualização por ID (ex.: "ID VENDA 1001, gastei 800 de material"),
            # atualiza/garante lembrete com base no que já está na planilha de vendas.
            target_sale_id = created_sale_id
            if not target_sale_id:
                if cmd.sale_id:
                    target_sale_id = str(cmd.sale_id).strip()
                elif cmd.material_allocations:
                    target_sale_id = str(cmd.material_allocations[0].sale_id).strip()
            if chat_id and target_sale_id:
                snap = self._sale_snapshot(ws_sales, target_sale_id)
                if snap:
                    pending_amount = self._to_float(snap.get("pending_amount"))
                    payment_status = snap.get("payment_status") or ("PENDENTE" if pending_amount > 0.01 else "PAGO")
                    due_date = snap.get("delivery_date") or ""
                    self.upsert_reminder(
                        wb,
                        sale_id=target_sale_id,
                        customer=snap.get("customer", ""),
                        description=snap.get("description", ""),
                        sale_date=snap.get("sale_date") or cmd.sale_date,
                        due_date=due_date or cmd.service_due_date,
                        pending_amount=pending_amount,
                        payment_status=payment_status,
                        service_status="PENDENTE",
                        chat_id=str(chat_id),
                    )
                    self._prepare_row_from_template(
                        wb,
                        ws,
                        row,
                        SHEET_SALES,
                        style_cols,
                        start_row=TEMPLATE_ROW,
                    )
                    ws[f"{sales_cols['data de venda']}{row}"] = self._excel_date_value(ref_date)
                    ws[f"{sales_cols['cliente']}{row}"] = party
                    ws[f"{sales_cols['id produto']}{row}"] = description
                    ws[f"{sales_cols['total de vendas (pago)']}{row}"] = float(amount)
                    if ws[f"{sales_cols['valor (pendente)']}{row}"].value in (None, ""):
                        ws[f"{sales_cols['valor (pendente)']}{row}"] = 0.0
                    pending_amount = self._to_float(ws[f"{sales_cols['valor (pendente)']}{row}"].value)
                    status = self._status_text(pending_amount)
                    ws[f"{sales_cols['status de valor']}{row}"] = self._match_text_case(
                        ws[f"{sales_cols['status de valor']}{row}"].value,
                        status,
                    )
                    if not self._preserve_user_layout():
                        template_ws, template_row = self._template_source(
                            wb,
                            ws,
                            SHEET_SALES,
                            self._sales_style_cols(ws),
                        )
                        self._copy_row_style_between(
                            template_ws,
                            template_row,
                            ws,
                            row,
                            (
                                sales_cols["total de vendas (pago)"],
                                sales_cols["valor (pendente)"],
                                sales_cols["status de valor"],
                            ),
                            apply_row_dimensions=False,
                        )
                        ws[f"{sales_cols['total de vendas (pago)']}{row}"] = float(amount)
                        if ws[f"{sales_cols['valor (pendente)']}{row}"].value in (None, ""):
                            ws[f"{sales_cols['valor (pendente)']}{row}"] = 0.0
                        pending_amount = self._to_float(ws[f"{sales_cols['valor (pendente)']}{row}"].value)
                        status = self._status_text(pending_amount)
                        ws[f"{sales_cols['status de valor']}{row}"] = self._match_text_case(
                            ws[f"{sales_cols['status de valor']}{row}"].value,
                            status,
                        )
                    self._ensure_number_format_if_general(ws, sales_cols["data de venda"], row, "DD/MM/YYYY")
                else:
                    material_cols = self._material_columns(ws)
                    style_cols = (
                        material_cols["data"],
                        material_cols["descricao"],
                        material_cols["valor"],
                    )
                    if material_cols.get("fornecedor"):
                        style_cols = (material_cols["fornecedor"],) + style_cols
                    if material_cols.get("id venda"):
                        style_cols = style_cols + (material_cols["id venda"],)
                    self._prepare_row_from_template(
                        wb,
                        ws,
                        row,
                        SHEET_MATERIAL,
                        style_cols,
                        start_row=TEMPLATE_ROW,
                    )
                    ws[f"{material_cols['data']}{row}"] = self._excel_date_value(ref_date)
                    if material_cols.get("fornecedor"):
                        ws[f"{material_cols['fornecedor']}{row}"] = party
                    ws[f"{material_cols['descricao']}{row}"] = description
                    ws[f"{material_cols['valor']}{row}"] = float(amount)
                    self._ensure_number_format_if_general(ws, material_cols["data"], row, "DD/MM/YYYY")
                    self._ensure_number_format_if_general(ws, material_cols["valor"], row, "R$ #,##0.00")
            elif target_norm == self._normalize_name(fixed_name):
                self._prepare_row_from_template(
                    wb,
                    ws,
                    row,
                    SHEET_FIXED,
                    ("A", "B", "C"),
                    start_row=TEMPLATE_ROW,
                )
                ws[f"A{row}"] = self._excel_date_value(ref_date)
                ws[f"B{row}"] = description
                ws[f"C{row}"] = float(amount)
                self._ensure_number_format_if_general(ws, "A", row, "DD/MM/YYYY")
                self._ensure_number_format_if_general(ws, "C", row, "R$ #,##0.00")
            else:
                raise ValueError(f"Aba nao suportada para correcao: {target_name}")

            self._append_log(
                ws=ws_log,
                log_id=f"{datetime.now().strftime('%Y%m%d%H%M%S')}-U-1",
                origin=origin,
                cmd=FinancialCommand(
                    customer=party or "",
                    description=description,
                    sale_date=ref_date,
                    total_value=float(amount),
                    payments=[],
                ),
                amount=float(amount),
                ref_date=ref_date,
                sheet=target_name,
                row=row,
                original_text=original_text,
            )

            self._save_workbook(wb)
        finally:
            wb.close()

    def read_last_rows(self, sheet_name: str, limit: int = 8) -> list[dict[str, str]]:
        wb = self._open_workbook(data_only=True)
        try:
            target_name = self._resolve_sheet_name(wb, sheet_name)
            sales_name = self._resolve_sheet_name(wb, SHEET_SALES)
            fixed_name = self._resolve_sheet_name(wb, SHEET_FIXED)
            ws = wb[target_name]

            if self._normalize_name(target_name) == self._normalize_name(fixed_name):
                cols = ("A", "B", "C")
            elif self._normalize_name(target_name) == self._normalize_name(sales_name):
                sales_cols = self._sales_columns(ws)
                cols = tuple(
                    sales_cols[key]
                    for key in (
                        "data de venda",
                        "data de entrega",
                        "cliente",
                        "id produto",
                        "total de vendas (pago)",
                        "valor (pendente)",
                        "id venda",
                        "status de valor",
                    )
                )
            else:
                material_cols = self._material_columns(ws)
                cols = (
                    material_cols["data"],
                    material_cols["descricao"],
                    material_cols["valor"],
                )
                if material_cols.get("fornecedor"):
                    cols = (material_cols["data"], material_cols["fornecedor"], material_cols["descricao"], material_cols["valor"])
                if material_cols.get("id venda"):
                    cols = cols + (material_cols["id venda"],)

            headers = [ws[f"{col}{SALES_HEADER_ROW if self._normalize_name(target_name) == self._normalize_name(sales_name) else 2}"].value for col in cols]
            rows: list[dict[str, str]] = []
            for row in range(ws.max_row, DATA_START_ROW - 1, -1):
                vals = [ws[f"{col}{row}"].value for col in cols]
                if all(v in (None, "") for v in vals):
                    continue
                rows.append({str(headers[i]): str(vals[i]) for i in range(len(cols))})
                if len(rows) >= limit:
                    break
            return rows
        finally:
            wb.close()

    def get_planilha_summary(self, max_rows_scan: int = 500, wb=None) -> str:
        """Resumo amigável de vendas, matéria-prima, gastos fixos e lucro para o bot."""
        own_wb = wb is None
        if own_wb:
            wb = self._open_workbook(data_only=True)
        try:
            def _format_currency_pt(value: float) -> str:
                txt = f"{float(value):,.2f}"
                # pt-BR: separador decimal vírgula e milhares ponto
                txt = txt.replace(",", "X").replace(".", ",").replace("X", ".")
                return f"R$ {txt}"

            # Totais de vendas (pago/pendente)
            sales_name = self._resolve_sheet_name(wb, SHEET_SALES)
            ws_sales = wb[sales_name]
            sales_cols = self._sales_columns(ws_sales)
            col_pago = sales_cols.get("total de vendas (pago)")
            col_pendente = sales_cols.get("valor (pendente)")
            col_id = sales_cols.get("id venda")
            n_sales = 0
            total_pago = 0.0
            total_pendente = 0.0
            sales_end = (
                self._effective_data_end_row(ws_sales, col_id, max_rows_scan=max_rows_scan)
                if col_id
                else self._scan_row_cap(ws_sales, max_rows_scan=max_rows_scan)
            )
            n_sales = self._count_filled_rows(ws_sales, col_id, DATA_START_ROW, sales_end)
            total_pago = self._sum_column_range(ws_sales, col_pago, DATA_START_ROW, sales_end)
            total_pendente = self._sum_column_range(ws_sales, col_pendente, DATA_START_ROW, sales_end)

            # Totais de matéria-prima
            material_name = self._resolve_sheet_name(wb, SHEET_MATERIAL)
            ws_mat = wb[material_name]
            mat_cols = self._material_columns(ws_mat)
            col_val = mat_cols.get("valor")
            col_desc = mat_cols.get("descricao")
            n_mat = 0
            total_mat = 0.0
            mat_end = (
                self._effective_data_end_row(ws_mat, col_desc, max_rows_scan=max_rows_scan)
                if col_desc
                else self._scan_row_cap(ws_mat, max_rows_scan=max_rows_scan)
            )
            n_mat = self._count_filled_rows(ws_mat, col_desc, DATA_START_ROW, mat_end)
            total_mat = self._sum_column_range(ws_mat, col_val, DATA_START_ROW, mat_end)

            # Totais de gastos fixos (somando a coluna Valor)
            fixed_name = self._resolve_sheet_name(wb, SHEET_FIXED)
            ws_fixed = wb[fixed_name]
            total_fixos = 0.0
            fix_end = self._effective_data_end_row(ws_fixed, "D", max_rows_scan=max_rows_scan)
            total_fixos = self._sum_column_range(ws_fixed, "D", DATA_START_ROW, fix_end)

            # Fallback (se o modelo mudar e a aba não existir/estragar leitura).
            # Lucro deve usar vendas totais (pago + pendente), para bater com a aba.
            lucro = (total_pago + total_pendente) - total_mat - total_fixos

            # Formato (padrão da aba "Lucro Mensal e Anual"):
            # Total Vendas | total de vendas (pago) | Valor (pendente) | Total Compras Matéria Prima | Gastos Fixos | Lucro
            total_vendas = total_pago + total_pendente
            return "\n".join(
                [
                    f"• Total Vendas: {_format_currency_pt(total_vendas)}",
                    f"• Total de vendas (pago): {_format_currency_pt(total_pago)}",
                    f"• Valor (pendente): {_format_currency_pt(total_pendente)}",
                    f"• Total Compras Materia Prima: {_format_currency_pt(total_mat)}",
                    f"• Gastos Fixos: {_format_currency_pt(total_fixos)}",
                    f"• Lucro: {_format_currency_pt(lucro)}",
                ]
            )
        except Exception as e:
            return f"Erro ao ler planilha: {e}"
        finally:
            if own_wb:
                wb.close()

    def get_sales_preview(self, limit: int = 12, wb=None) -> str:
        """Prévia curta para o Telegram: últimas vendas com pendência e/ou prazo de entrega."""
        own_wb = wb is None
        if own_wb:
            wb = self._open_workbook(data_only=True)
        try:
            sales_name = self._resolve_sheet_name(wb, SHEET_SALES)
            ws = wb[sales_name]
            sales_cols = self._sales_columns(ws)

            col_id = sales_cols["id venda"]
            col_customer = sales_cols["cliente"]
            col_desc = sales_cols["id produto"]
            col_pending = sales_cols["valor (pendente)"]
            col_status = sales_cols["status de valor"]
            col_delivery = sales_cols["data de entrega"]
            col_sale_date = sales_cols.get("data de venda")
            material_totals = self.material_costs_by_sale_id(wb)

            rows: list[dict[str, str]] = []
            end_row = self._effective_data_end_row(ws, col_id)
            if getattr(ws, "read_only", False):
                from openpyxl.utils import column_index_from_string

                idx_map = {
                    "id": column_index_from_string(col_id),
                    "customer": column_index_from_string(col_customer),
                    "desc": column_index_from_string(col_desc),
                    "pending": column_index_from_string(col_pending),
                    "status": column_index_from_string(col_status),
                    "delivery": column_index_from_string(col_delivery),
                }
                if col_sale_date:
                    idx_map["sale_date"] = column_index_from_string(col_sale_date)
                min_col = min(idx_map.values())
                max_col = max(idx_map.values())
                collected: list[dict[str, str]] = []
                for row_cells in ws.iter_rows(
                    min_row=DATA_START_ROW,
                    max_row=end_row,
                    min_col=min_col,
                    max_col=max_col,
                ):
                    values = {cell.column: cell.value for cell in row_cells}

                    def _val(key: str):
                        col_i = idx_map.get(key)
                        return values.get(col_i) if col_i else None

                    sale_id = str(_val("id") or "").strip()
                    if not sale_id:
                        continue
                    pending = self._to_float(_val("pending"))
                    delivery = str(_val("delivery") or "").strip()
                    if (pending <= 0.01) and (not delivery):
                        continue
                    collected.append(
                        {
                            "id": sale_id,
                            "cliente": str(_val("customer") or "").strip(),
                            "desc": str(_val("desc") or "").strip(),
                            "pendente": pending,
                            "status": str(_val("status") or "").strip(),
                            "entrega": delivery,
                            "data": str(_val("sale_date") or "").strip() if col_sale_date else "",
                            "material": material_totals.get(sale_id, 0.0),
                        }
                    )
                rows = list(reversed(collected[-limit:]))
            else:
                for row in range(end_row, DATA_START_ROW - 1, -1):
                    sale_id = str(ws[f"{col_id}{row}"].value or "").strip()
                    if not sale_id:
                        continue
                    customer = str(ws[f"{col_customer}{row}"].value or "").strip()
                    desc = str(ws[f"{col_desc}{row}"].value or "").strip()
                    pending = self._to_float(ws[f"{col_pending}{row}"].value)
                    status = str(ws[f"{col_status}{row}"].value or "").strip()
                    delivery = str(ws[f"{col_delivery}{row}"].value or "").strip()
                    sale_date = str(ws[f"{col_sale_date}{row}"].value or "").strip()

                    if (pending <= 0.01) and (not delivery):
                        continue
                    rows.append(
                        {
                            "id": sale_id,
                            "cliente": customer,
                            "desc": desc,
                            "pendente": pending,
                            "status": status,
                            "entrega": delivery,
                            "data": sale_date,
                            "material": material_totals.get(sale_id, 0.0),
                        }
                    )
                    if len(rows) >= limit:
                        break

            if not rows:
                return "📄 *Prévia da planilha*\n\nNenhuma venda com pendência ou prazo de entrega encontrada nas últimas linhas."

            lines = ["📄 *Prévia da planilha (pendências e entregas)*", ""]
            for item in rows:
                pend_txt = f"R$ {item['pendente']:,.2f}" if item["pendente"] > 0.01 else "R$ 0,00"
                entrega = item["entrega"] or "-"
                cliente = item["cliente"] or "-"
                material_val = float(item.get("material") or 0.0)
                material_txt = (
                    f" | Material: R$ {material_val:,.2f}" if material_val > 0.01 else ""
                )
                lines.append(
                    f"• ID {item['id']} | {cliente}\n"
                    f"  Venda: {item['data'] or '-'} | Pendente: {pend_txt} ({item['status'] or '-'}) | "
                    f"Entrega: {entrega}{material_txt}"
                )
            lines.append("")
            # Dica baseada em um exemplo real de pendência da prévia.
            example_customer = None
            for item in rows:
                if item.get("cliente") and self._to_float(item.get("pendente", 0.0)) > 0.01:
                    example_customer = item.get("cliente")
                    break
            if not example_customer and rows:
                example_customer = rows[0].get("cliente")

            if example_customer:
                lines.append(
                    f"Dica: se o cliente {example_customer} tem pendência, envie `ID VENDA 001 pagou` "
                    f"(use o ID VENDA da linha) para marcar a venda como paga."
                )
            else:
                lines.append("Dica: envie `ID VENDA 001 pagou` para atualizar o status.")
            return "\n".join(lines)
        finally:
            if own_wb:
                wb.close()

    def get_pending_sales_by_customer(
        self,
        customer_id: str,
        max_rows_scan: int = 500,
        *,
        include_paid: bool = False,
    ) -> list[dict[str, str]]:
        """
        Retorna vendas pendentes (ou com status pendente) pelo nome na coluna Cliente.
        Usado no Telegram quando o usuário menciona o nome do cliente.
        """
        wb = self._open_workbook(data_only=True)
        try:
            import re

            def _norm_name(v: object) -> str:
                return self._normalize_name(str(v or "").strip())

            sales_name = self._resolve_sheet_name(wb, SHEET_SALES)
            ws = wb[sales_name]
            sales_cols = self._sales_columns(ws)

            col_id = sales_cols.get("id venda")
            col_customer = sales_cols.get("cliente")
            col_desc = sales_cols.get("id produto")
            col_pending = sales_cols.get("valor (pendente)")
            col_status = sales_cols.get("status de valor")
            col_delivery = sales_cols.get("data de entrega")
            col_sale_date = sales_cols.get("data de venda")

            if not col_id or not col_customer:
                return []

            normalized_customer = _norm_name(customer_id)
            rows: list[dict[str, str]] = []
            start_row = min(ws.max_row, self.MAX_DATA_ROW, DATA_START_ROW + max_rows_scan)
            for row in range(start_row, DATA_START_ROW - 1, -1):
                cust_val_raw = ws[f"{col_customer}{row}"].value
                cust_val = str(cust_val_raw or "").strip()
                if _norm_name(cust_val) != normalized_customer:
                    continue

                pending = self._to_float(ws[f"{col_pending}{row}"].value) if col_pending else 0.0
                status_txt = str(ws[f"{col_status}{row}"].value or "").strip().lower() if col_status else ""
                if not include_paid:
                    if pending <= 0.01 and status_txt != "pendente":
                        continue

                delivery = str(ws[f"{col_delivery}{row}"].value or "").strip() if col_delivery else ""
                sale_date = str(ws[f"{col_sale_date}{row}"].value or "").strip() if col_sale_date else ""
                desc = str(ws[f"{col_desc}{row}"].value or "").strip() if col_desc else ""

                rows.append(
                    {
                        "sale_id": str(ws[f"{col_id}{row}"].value or "").strip(),
                        "cliente": cust_val,
                        "desc": desc,
                        "pendente": f"{pending:.2f}",
                        "status": status_txt or ("pendente" if pending > 0.01 else "pago"),
                        "delivery": delivery,
                        "sale_date": sale_date,
                    }
                )
                if len(rows) >= 5:
                    break
            return rows
        finally:
            wb.close()

    def get_pending_sale_by_sale_id(
        self,
        sale_id: str,
        max_rows_scan: int = 500,
        *,
        include_paid: bool = False,
    ) -> list[dict[str, str]]:
        """
        Retorna uma venda pendente (status pendente/pagamento pendente) por ID VENDA.
        Usado quando o usuário informa o ID VENDA (ex.: "id venda 004 pagou").
        """
        wb = self._open_workbook(data_only=True)
        try:
            import re

            def _norm_sale_id(v: object) -> str:
                raw = str(v or "").strip()
                digits = re.sub(r"\D", "", raw)
                if digits and len(digits) <= 3:
                    return digits.zfill(3)
                return digits or raw.strip().upper()

            sales_name = self._resolve_sheet_name(wb, SHEET_SALES)
            ws = wb[sales_name]
            sales_cols = self._sales_columns(ws)

            col_id = sales_cols.get("id venda")
            col_customer = sales_cols.get("cliente")
            col_desc = sales_cols.get("id produto")
            col_pending = sales_cols.get("valor (pendente)")
            col_status = sales_cols.get("status de valor")
            col_delivery = sales_cols.get("data de entrega")
            col_sale_date = sales_cols.get("data de venda")

            if not col_id:
                return []

            normalized_sale_id = _norm_sale_id(sale_id)
            rows: list[dict[str, str]] = []
            start_row = min(ws.max_row, self.MAX_DATA_ROW, DATA_START_ROW + max_rows_scan)
            for row in range(start_row, DATA_START_ROW - 1, -1):
                sale_val_raw = ws[f"{col_id}{row}"].value
                sale_val = _norm_sale_id(sale_val_raw)
                if sale_val != normalized_sale_id:
                    continue

                pending = self._to_float(ws[f"{col_pending}{row}"].value) if col_pending else 0.0
                status_txt = str(ws[f"{col_status}{row}"].value or "").strip().lower() if col_status else ""
                if not include_paid:
                    if pending <= 0.01 and status_txt != "pendente":
                        # Já está pago/sem pendência
                        return []

                delivery = str(ws[f"{col_delivery}{row}"].value or "").strip() if col_delivery else ""
                sale_date = str(ws[f"{col_sale_date}{row}"].value or "").strip() if col_sale_date else ""
                desc = str(ws[f"{col_desc}{row}"].value or "").strip() if col_desc else ""
                cliente = str(ws[f"{col_customer}{row}"].value or "").strip() if col_customer else ""

                rows.append(
                    {
                        "sale_id": sale_val,
                        "cliente": cliente,
                        "desc": desc,
                        "pendente": f"{pending:.2f}",
                        "status": status_txt or ("pendente" if pending > 0.01 else "pago"),
                        "delivery": delivery,
                        "sale_date": sale_date,
                    }
                )
                break
            return rows
        finally:
            wb.close()
